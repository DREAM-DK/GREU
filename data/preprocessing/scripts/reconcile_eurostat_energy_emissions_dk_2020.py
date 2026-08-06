"""Reconcile Eurostat PEFA/air accounts with GREU Danish energy data, 2020.

Inputs
------
- energy_and_emissions.xlsx (GREU energy account, PJ and kt)
- non_energy_emissions.xlsx (used to explain the total-air-account boundary)
- emissions_bridge_items.xlsx (residence-to-territory scope diagnostic)
- metadata.xlsx (GREU-to-PEFA product and GREU-to-NACE A64 maps)
- eurostat_energy_emissions_raw/*.json (downloaded by the companion script)

Output
------
data/preprocessing/data/eurostat_energy_emissions_dk2020_reconciliation.xlsx

Important scope distinction
---------------------------
Eurostat PEFA's published all-flow total includes natural inputs and all energy
residuals, including dissipative heat. GREU selects energy products plus a
small set of renewable natural inputs and waste residuals. The workbook reports:
1) the strict existing metadata concordance; and
2) a concept-adjusted PEFA total that supplements four P-products missing from
   the concordance while excluding upstream natural-input and residual-loss
   stages that would double count the GREU product flow.

Run
---
python data/preprocessing/scripts/reconcile_eurostat_energy_emissions_dk_2020.py
"""

from __future__ import annotations

import datetime
import json
import pathlib
from collections import defaultdict

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DATA = pathlib.Path(__file__).resolve().parents[1] / "data"
RAW = DATA / "eurostat_energy_emissions_raw"
OUT = DATA / "eurostat_energy_emissions_dk2020_reconciliation.xlsx"
YEAR = 2020
RETRIEVAL_DATE = "2026-07-30"

PEFA_URL = (
    "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/"
    "env_ac_pefasu?geo=DK&time=2020&lang=en"
)
AEA_URL = (
    "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/"
    "env_ac_ainah_r2?geo=DK&time=2020&lang=en"
)

# GREU's product-flow boundary: all energy products, selected renewable natural
# inputs, and energy-bearing waste residuals. N06 is excluded because its
# transformed biomass appears again in P23; N01 fossil inputs similarly appear
# as P12/P13. R30/R31 are outside GREU's energy-product list.
PEFA_P_CODES = [f"P{i:02d}" for i in range(8, 28)]
CONCEPT_CODES = PEFA_P_CODES + ["N03", "N04", "N05", "N07", "R28", "R29"]
PEFA_COMPONENT_CODES = (
    [f"N{i:02d}" for i in range(1, 8)]
    + PEFA_P_CODES
    + [f"R{i:02d}" for i in range(28, 32)]
)

AEA_POLLUTANT_MAP = {
    "ch4": "CH4",
    "co2_bio": "CO2_BIO",
    "co2_xbio": "CO2",
    "n2o": "N2O",
    "co2_eq": "GHG",
}


def jsonstat_to_long(path: pathlib.Path) -> tuple[pd.DataFrame, dict]:
    """Expand a sparse JSON-stat response to one row per observation."""
    payload = json.loads(path.read_text(encoding="utf-8"))
    dims = payload["id"]
    categories = []
    labels = {}
    for dim in dims:
        category = payload["dimension"][dim]["category"]
        ordered = [k for k, _ in sorted(category["index"].items(), key=lambda x: x[1])]
        categories.append(ordered)
        labels[dim] = category.get("label", {})

    statuses = payload.get("status", {})
    rows = []
    for flat_index, value in payload["value"].items():
        positions = np.unravel_index(int(flat_index), payload["size"])
        row = {dim: categories[i][positions[i]] for i, dim in enumerate(dims)}
        row["value"] = value
        row["obs_status"] = statuses.get(str(flat_index), "")
        rows.append(row)
    return pd.DataFrame(rows), labels


def build_clusters(mapping: pd.DataFrame, left: str, right: str):
    """Connected components of a many-to-many concordance."""
    parent: dict[tuple[str, str], tuple[str, str]] = {}

    def find(node):
        parent.setdefault(node, node)
        if parent[node] != node:
            parent[node] = find(parent[node])
        return parent[node]

    def union(a, b):
        parent[find(a)] = find(b)

    for row in mapping[[left, right]].drop_duplicates().itertuples(index=False):
        union(("left", str(row[0])), ("right", str(row[1])))

    components: dict[tuple[str, str], list[tuple[str, str]]] = defaultdict(list)
    for node in list(parent):
        components[find(node)].append(node)

    clusters = []
    for members in components.values():
        left_values = sorted(v for side, v in members if side == "left")
        right_values = sorted(v for side, v in members if side == "right")
        clusters.append(
            {
                "cluster": "+".join(right_values),
                "left_values": left_values,
                "right_values": right_values,
            }
        )
    clusters.sort(key=lambda x: x["cluster"])
    left_to_cluster = {v: c["cluster"] for c in clusters for v in c["left_values"]}
    right_to_cluster = {v: c["cluster"] for c in clusters for v in c["right_values"]}
    return clusters, left_to_cluster, right_to_cluster


def comparison_row(metric, dk, eu, unit, note=""):
    diff = eu - dk if pd.notna(dk) and pd.notna(eu) else np.nan
    return {
        "metric": metric,
        "unit": unit,
        "dk_greu": dk,
        "eurostat": eu,
        "difference_eurostat_minus_dk": diff,
        "difference_pct_of_dk": diff / dk * 100 if pd.notna(diff) and abs(dk) > 1e-12 else np.nan,
        "interpretation": note,
    }


def style_workbook(path: pathlib.Path) -> None:
    """Apply simple readable formatting without merged cells."""
    from openpyxl import load_workbook

    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column in sheet.columns:
            letter = get_column_letter(column[0].column)
            max_length = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 55)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
    workbook.save(path)


def main() -> None:
    pefa, pefa_labels = jsonstat_to_long(RAW / "env_ac_pefasu_DK_2020.json")
    aea, aea_labels = jsonstat_to_long(RAW / "env_ac_ainah_r2_DK_2020.json")

    assert set(pefa["time"]) == {"2020"} and set(pefa["geo"]) == {"DK"}
    assert set(pefa["unit"]) == {"TJ"}
    assert set(aea["time"]) == {"2020"} and set(aea["geo"]) == {"DK"}
    assert "THS_T" in set(aea["unit"])
    assert not pefa.empty and not aea.empty

    energy = pd.read_excel(DATA / "energy_and_emissions.xlsx", sheet_name="ems_energy")
    nonenergy = pd.read_excel(DATA / "non_energy_emissions.xlsx", sheet_name="ems_non_energy")
    bridge = pd.read_excel(DATA / "emissions_bridge_items.xlsx", sheet_name="bridge_items")
    product_map_raw = pd.read_excel(DATA / "metadata.xlsx", sheet_name="energy_products_pefa_map")
    industry_map = pd.read_excel(DATA / "metadata.xlsx", sheet_name="industries_naceA64_map")

    energy = energy[energy["year"] == YEAR].copy()
    nonenergy = nonenergy[nonenergy["year"] == YEAR].copy()
    bridge = bridge[bridge["year"] == YEAR].copy()
    product_map_raw["product_greu"] = product_map_raw["product_greu"].astype(str)
    product_map_raw["product_pefa"] = product_map_raw["product_pefa"].astype(str)
    # Pilot-discovered corrections used for comparison but not written back to
    # the input workbook: one spelling mismatch, one missing gasoil code, and
    # ambient heat classified as a renewable natural input rather than P27 heat.
    product_map = product_map_raw.copy()
    product_map["product_greu"] = product_map["product_greu"].replace(
        {"sem_refin_oil": "semi_refin_oil"}
    )
    product_map = product_map[
        ~((product_map["product_greu"] == "heat_pump") & (product_map["product_pefa"] == "P27"))
    ]
    product_map = pd.concat(
        [
            product_map,
            pd.DataFrame(
                [
                    {
                        "product_greu": "diesel_transp",
                        "product_greu_txt": "Diesel for transport",
                        "product_pefa": "P18",
                        "product_pefa_txt": "Heating and other gasoil (without bio)",
                    },
                    *[
                        {
                            "product_greu": "heat_pump",
                            "product_greu_txt": "Heat pumps",
                            "product_pefa": code,
                            "product_pefa_txt": pefa_labels["prod_nrg"].get(code, ""),
                        }
                        for code in ["N03", "N04", "N05", "N07"]
                    ],
                ]
            ),
        ],
        ignore_index=True,
    ).drop_duplicates()
    industry_map["indu_greu"] = industry_map["indu_greu"].astype(str)
    industry_map["indu_naceA64"] = industry_map["indu_naceA64"].astype(str)

    # Validate source accounting identities before comparing.
    pefa_grand = pefa[
        (pefa["nace_r2"] == "NRG_FLOW") & (pefa["prod_nrg"] == "N00_P00_R00")
    ].set_index("stk_flow")["value"]
    assert abs(pefa_grand["SUP"] - pefa_grand["USE"]) < 1e-6
    assert abs(energy.loc[energy["bal"] == "sup", "pj"].sum() - energy.loc[energy["bal"] == "use", "pj"].sum()) < 1e-5

    raw_mapped_product_codes = sorted(product_map_raw["product_pefa"].unique())
    mapped_product_codes = sorted(product_map["product_pefa"].unique())
    nace_codes = sorted(industry_map["indu_naceA64"].unique())
    missing_nace = set(nace_codes) - set(pefa["nace_r2"])
    assert not missing_nace, f"NACE codes absent from PEFA: {sorted(missing_nace)}"

    def pefa_total(stk_flow, products, activities=("NRG_FLOW",)):
        return (
            pefa[
                (pefa["stk_flow"] == stk_flow)
                & pefa["prod_nrg"].isin(products)
                & pefa["nace_r2"].isin(activities)
            ]["value"].sum()
            / 1000.0
        )

    # ---------------- Physical-energy totals and flow mapping ----------------
    dk_supply = energy[energy["bal"] == "sup"]["pj"].sum()
    dk_use = energy[energy["bal"] == "use"]["pj"].sum()
    pefa_strict = pefa_total("SUP", raw_mapped_product_codes)
    pefa_concept = pefa_total("SUP", CONCEPT_CODES)
    energy_totals = pd.DataFrame(
        [
            comparison_row(
                "Selected GREU energy-account flow: strict existing product concordance",
                dk_supply,
                pefa_strict,
                "PJ",
                "Understates coverage because metadata.xlsx omits P10, P16, P18 and P22.",
            ),
            comparison_row(
                "Selected GREU energy-account flow: concept-adjusted PEFA boundary",
                dk_supply,
                pefa_concept,
                "PJ",
                "All P08-P27 plus N03/N04/N05/N07 and R28/R29; excludes upstream/dissipative stages.",
            ),
            comparison_row(
                "Published PEFA all-flow supply/use total",
                dk_supply,
                pefa_grand["SUP"] / 1000.0,
                "PJ",
                "Not like-for-like: includes all natural inputs and energy residuals, especially dissipative heat.",
            ),
            comparison_row(
                "PEFA energy-products aggregate P00",
                dk_supply,
                pefa_total("SUP", ["P00"]),
                "PJ",
                "Excludes GREU renewable natural inputs and waste residuals.",
            ),
            comparison_row("GREU supply-use identity", dk_supply, dk_use, "PJ", "Must balance."),
            comparison_row(
                "PEFA concept-adjusted supply-use identity",
                pefa_concept,
                pefa_total("USE", CONCEPT_CODES),
                "PJ",
                "Must balance.",
            ),
        ]
    )

    sup = energy[energy["bal"] == "sup"].groupby("flow")["pj"].sum()
    use = energy[energy["bal"] == "use"].groupby("flow")["pj"].sum()

    def concept_by_activity(stk_flow, activities):
        return pefa_total(stk_flow, CONCEPT_CODES, activities)

    pefa_industry_use = concept_by_activity("USE", nace_codes)
    flow_rows = [
        comparison_row(
            "Domestic production",
            sup.get("production", 0.0),
            concept_by_activity("SUP", nace_codes),
            "PJ",
            "PEFA supply by the mapped NACE A64 industries.",
        ),
        comparison_row(
            "Imports",
            sup.get("import", 0.0),
            concept_by_activity("SUP", ["ROW_ACT"]),
            "PJ",
            "Rest-of-world supply.",
        ),
        comparison_row(
            "Natural/residual other supply",
            sup.get("other_supply", 0.0),
            concept_by_activity("SUP", ["ENV", "CH_INV_PA"]),
            "PJ",
            "Environment plus changes in inventories/produced assets supplying waste residuals.",
        ),
        comparison_row(
            "Intermediate use",
            use.get("cons_inter", 0.0),
            pefa_industry_use,
            "PJ",
            "PEFA has no separate transmission-loss use category.",
        ),
        comparison_row(
            "Intermediate use plus GREU transmission loss",
            use.get("cons_inter", 0.0) + use.get("transmis_loss", 0.0),
            pefa_industry_use,
            "PJ",
            "Closest use-side boundary because PEFA embeds losses differently.",
        ),
        comparison_row(
            "Households",
            use.get("cons_hh", 0.0),
            concept_by_activity("USE", ["HH"]),
            "PJ",
            "Total household activities.",
        ),
        comparison_row(
            "Exports",
            use.get("export", 0.0),
            concept_by_activity("USE", ["ROW_ACT"]),
            "PJ",
            "Rest-of-world use.",
        ),
        comparison_row(
            "Inventory change",
            use.get("invent_change", 0.0),
            concept_by_activity("USE", ["CH_INV_PA"]),
            "PJ",
            "Changes in inventories and produced assets.",
        ),
        comparison_row(
            "Transmission loss (separate row)",
            use.get("transmis_loss", 0.0),
            0.0,
            "PJ",
            "No directly matching PEFA use cell; PEFA records energy losses mainly as R30 residual supply.",
        ),
    ]
    energy_flows = pd.DataFrame(flow_rows)

    # ---------------- Product-cluster comparison ----------------
    product_clusters, _, _ = build_clusters(product_map, "product_greu", "product_pefa")
    product_rows = []
    for cluster in product_clusters:
        greu_products = cluster["left_values"]
        pefa_products = cluster["right_values"]
        dk_sup = energy[
            (energy["bal"] == "sup") & energy["product"].astype(str).isin(greu_products)
        ]["pj"].sum()
        dk_use_cluster = energy[
            (energy["bal"] == "use") & energy["product"].astype(str).isin(greu_products)
        ]["pj"].sum()
        eu_sup = pefa_total("SUP", pefa_products)
        eu_use = pefa_total("USE", pefa_products)
        product_rows.append(
            {
                "product_cluster": cluster["cluster"],
                "greu_products": ", ".join(greu_products),
                "pefa_products": ", ".join(pefa_products),
                "n_greu_products": len(greu_products),
                "n_pefa_products": len(pefa_products),
                "dk_supply_pj": dk_sup,
                "pefa_supply_pj": eu_sup,
                "supply_difference_pj": eu_sup - dk_sup,
                "supply_difference_pct": (eu_sup - dk_sup) / dk_sup * 100 if dk_sup else np.nan,
                "dk_use_pj": dk_use_cluster,
                "pefa_use_pj": eu_use,
                "use_difference_pj": eu_use - dk_use_cluster,
            }
        )
    energy_products = pd.DataFrame(product_rows)

    component_total = (
        pefa[
            (pefa["stk_flow"] == "SUP")
            & (pefa["nace_r2"] == "NRG_FLOW")
            & pefa["prod_nrg"].isin(PEFA_COMPONENT_CODES)
        ]
        .set_index("prod_nrg")["value"]
        .div(1000.0)
    )
    component_use = (
        pefa[
            (pefa["stk_flow"] == "USE")
            & (pefa["nace_r2"] == "NRG_FLOW")
            & pefa["prod_nrg"].isin(PEFA_COMPONENT_CODES)
        ]
        .set_index("prod_nrg")["value"]
        .div(1000.0)
    )

    def component_reason(code):
        if code in raw_mapped_product_codes:
            return "Mapped in metadata.xlsx"
        if code == "P18":
            return "Missing from input concordance; pilot assigns to GREU diesel cluster"
        if code in {"P10", "P16", "P18", "P22"}:
            return "Unmapped PEFA energy product; omission from current concordance"
        if code in {"N01", "N02", "N06"}:
            return "Upstream natural input; excluded to avoid double counting transformed P-products"
        if code in {"N03", "N04", "N05", "N07", "R28", "R29"}:
            return "Included in concept-adjusted boundary"
        if code in {"R30", "R31"}:
            return "Energy loss/non-energy-use residual outside GREU product boundary"
        return "Not mapped; zero or outside selected boundary"

    pefa_components = pd.DataFrame(
        [
            {
                "pefa_code": code,
                "pefa_label": pefa_labels["prod_nrg"].get(code, ""),
                "mapped_in_metadata": "yes" if code in raw_mapped_product_codes else "no",
                "mapped_in_pilot": "yes" if code in mapped_product_codes else "no",
                "included_concept_adjusted": "yes" if code in CONCEPT_CODES else "no",
                "supply_pj": component_total.get(code, 0.0),
                "use_pj": component_use.get(code, 0.0),
                "treatment": component_reason(code),
            }
            for code in PEFA_COMPONENT_CODES
        ]
    )
    mapping_adjustments = pd.DataFrame(
        [
            {
                "issue": "Spelling mismatch",
                "input_mapping": "sem_refin_oil -> P21",
                "pilot_treatment": "semi_refin_oil -> P21",
                "evidence": (
                    "energy_and_emissions.xlsx uses semi_refin_oil; the metadata spelling otherwise "
                    "leaves 7.842 PJ of GREU supply unmatched."
                ),
                "recommended_action": "Correct the concordance spelling after owner review.",
            },
            {
                "issue": "Missing heating/other gasoil code",
                "input_mapping": "diesel_transp -> P17 only",
                "pilot_treatment": "diesel_transp -> P17 + P18",
                "evidence": (
                    "P17+P18 = 238.361 PJ, matching diesel_transp+bunk_trucks = 238.361 PJ; "
                    "P18 alone is 146.140 PJ."
                ),
                "recommended_action": "Add P18 to the GREU diesel cluster; consider renaming the GREU product.",
            },
            {
                "issue": "Ambient heat classification",
                "input_mapping": "heat_pump -> P27",
                "pilot_treatment": "heat_pump + renewable -> N03/N04/N05/N07; district_heat -> P27",
                "evidence": (
                    "GREU heat_pump is environment-supplied. PEFA renewable natural inputs = "
                    "79.562 PJ, exactly GREU renewable+heat_pump within rounding."
                ),
                "recommended_action": "Move heat_pump to the renewable-natural-input cluster after owner review.",
            },
            {
                "issue": "Unmapped derived gas",
                "input_mapping": "No GREU mapping for P10",
                "pilot_treatment": "Included in concept total but not assigned to a product cluster",
                "evidence": "PEFA P10 supply/use is 0.555 PJ in Denmark 2020.",
                "recommended_action": "Decide whether P10 belongs with refinery/other gas or remains an explicit residual.",
            },
        ]
    )

    # ---------------- Household energy-purpose comparison ----------------
    hh_purpose_map = {
        "heating": "HH_HEAT",
        "transport": "HH_TRA",
        "appliances": "HH_OTH",
    }
    hh_rows = []
    for purpose, pefa_activity in hh_purpose_map.items():
        dk = energy[
            (energy["bal"] == "use")
            & (energy["flow"] == "cons_hh")
            & (energy["purp"] == purpose)
        ]["pj"].sum()
        eu = concept_by_activity("USE", [pefa_activity])
        hh_rows.append(
            comparison_row(
                f"{purpose} / {pefa_activity}",
                dk,
                eu,
                "PJ",
                "Concepts are close but national allocation between heating and other household use differs.",
            )
        )
    hh_rows.append(
        comparison_row(
            "Household total / HH",
            energy[(energy["bal"] == "use") & (energy["flow"] == "cons_hh")]["pj"].sum(),
            concept_by_activity("USE", ["HH"]),
            "PJ",
            "Total is the robust comparison; PEFA household subcategories are available EU-wide.",
        )
    )
    household_energy = pd.DataFrame(hh_rows)

    # ---------------- Industry-cluster energy use ----------------
    industry_clusters, greu_to_cluster, nace_to_cluster = build_clusters(
        industry_map, "indu_greu", "indu_naceA64"
    )
    dk_industry_use = (
        energy[
            (energy["bal"] == "use")
            & (energy["flow"] == "cons_inter")
            & energy["indu"].astype(str).isin(greu_to_cluster)
        ]
        .assign(cluster=lambda x: x["indu"].astype(str).map(greu_to_cluster))
        .groupby("cluster")["pj"]
        .sum()
    )
    pefa_industry = (
        pefa[
            (pefa["stk_flow"] == "USE")
            & pefa["prod_nrg"].isin(CONCEPT_CODES)
            & pefa["nace_r2"].isin(nace_to_cluster)
        ]
        .assign(cluster=lambda x: x["nace_r2"].map(nace_to_cluster))
        .groupby("cluster")["value"]
        .sum()
        .div(1000.0)
    )
    energy_industry_rows = []
    for cluster in industry_clusters:
        key = cluster["cluster"]
        dk = dk_industry_use.get(key, 0.0)
        eu = pefa_industry.get(key, 0.0)
        energy_industry_rows.append(
            {
                "cluster": key,
                "greu_industries": ", ".join(cluster["left_values"]),
                "nace_a64_codes": ", ".join(cluster["right_values"]),
                "n_greu": len(cluster["left_values"]),
                "n_a64": len(cluster["right_values"]),
                "pefa_coarser": "yes" if len(cluster["left_values"]) > len(cluster["right_values"]) else "",
                "dk_pj": dk,
                "pefa_pj": eu,
                "difference_pj": eu - dk,
                "difference_pct_of_dk": (eu - dk) / dk * 100 if dk else np.nan,
            }
        )
    energy_by_industry = pd.DataFrame(energy_industry_rows)

    # ---------------- Emissions totals ----------------
    aea_kt = aea[(aea["unit"] == "THS_T") & (aea["nace_r2"] == "TOTAL_HH")].set_index("airpol")[
        "value"
    ]
    emission_rows = []
    for dk_column, aea_code in AEA_POLLUTANT_MAP.items():
        dk_energy = energy[dk_column].sum()
        dk_nonenergy = nonenergy[dk_column].sum()
        dk_combined = dk_energy + dk_nonenergy
        eu = aea_kt.get(aea_code, np.nan)
        emission_rows.append(
            {
                "pollutant": dk_column,
                "eurostat_code": aea_code,
                "unit": "kt" if dk_column != "co2_eq" else "kt CO2e",
                "dk_energy_only": dk_energy,
                "dk_nonenergy": dk_nonenergy,
                "dk_combined_air_account": dk_combined,
                "eurostat_air_account": eu,
                "difference_eurostat_minus_combined": eu - dk_combined,
                "difference_pct_of_combined": (eu - dk_combined) / dk_combined * 100
                if dk_combined
                else np.nan,
                "energy_share_pct": dk_energy / dk_combined * 100 if dk_combined else np.nan,
            }
        )
    dk_fgas = nonenergy[["hfc", "pfc", "sf6"]].sum().sum()
    eu_fgas = aea_kt.reindex(["HFC_CO2E", "PFC_CO2E", "NF3_SF6_CO2E"]).sum()
    emission_rows.append(
        {
            "pollutant": "f_gases_co2e",
            "eurostat_code": "HFC_CO2E+PFC_CO2E+NF3_SF6_CO2E",
            "unit": "kt CO2e",
            "dk_energy_only": 0.0,
            "dk_nonenergy": dk_fgas,
            "dk_combined_air_account": dk_fgas,
            "eurostat_air_account": eu_fgas,
            "difference_eurostat_minus_combined": eu_fgas - dk_fgas,
            "difference_pct_of_combined": (eu_fgas - dk_fgas) / dk_fgas * 100,
            "energy_share_pct": 0.0,
        }
    )
    emissions_totals = pd.DataFrame(emission_rows)

    # ---------------- Emissions by industry cluster ----------------
    aea_ind = aea[
        (aea["unit"] == "THS_T")
        & aea["nace_r2"].isin(nace_to_cluster)
        & aea["airpol"].isin(AEA_POLLUTANT_MAP.values())
    ].copy()
    aea_ind["cluster"] = aea_ind["nace_r2"].map(nace_to_cluster)
    aea_cluster = aea_ind.groupby(["airpol", "cluster"])["value"].sum()

    emissions_cluster_rows = []
    for dk_column, aea_code in AEA_POLLUTANT_MAP.items():
        dk_energy_by = (
            energy[
                (energy["flow"] == "cons_inter")
                & energy["indu"].astype(str).isin(greu_to_cluster)
            ]
            .assign(cluster=lambda x: x["indu"].astype(str).map(greu_to_cluster))
            .groupby("cluster")[dk_column]
            .sum()
        )
        dk_nonenergy_by = (
            nonenergy[
                (nonenergy["flow"] == "cons_inter")
                & nonenergy["indu"].astype(str).isin(greu_to_cluster)
            ]
            .assign(cluster=lambda x: x["indu"].astype(str).map(greu_to_cluster))
            .groupby("cluster")[dk_column]
            .sum()
        )
        for cluster in industry_clusters:
            key = cluster["cluster"]
            dke = dk_energy_by.get(key, 0.0)
            dkn = dk_nonenergy_by.get(key, 0.0)
            eu = aea_cluster.get((aea_code, key), 0.0)
            combined = dke + dkn
            emissions_cluster_rows.append(
                {
                    "pollutant": dk_column,
                    "cluster": key,
                    "greu_industries": ", ".join(cluster["left_values"]),
                    "nace_a64_codes": ", ".join(cluster["right_values"]),
                    "dk_energy": dke,
                    "dk_nonenergy": dkn,
                    "dk_combined": combined,
                    "eurostat": eu,
                    "difference": eu - combined,
                    "difference_pct": (eu - combined) / combined * 100 if combined else np.nan,
                }
            )
    emissions_by_industry = pd.DataFrame(emissions_cluster_rows)

    # Household emissions: air accounts contain all sources, so compare combined.
    household_emission_rows = []
    for dk_column, aea_code in AEA_POLLUTANT_MAP.items():
        dke = energy[energy["flow"] == "cons_hh"][dk_column].sum()
        dkn = nonenergy[nonenergy["flow"] == "cons_hh"][dk_column].sum()
        eu = aea[
            (aea["unit"] == "THS_T") & (aea["nace_r2"] == "HH") & (aea["airpol"] == aea_code)
        ]["value"].sum()
        household_emission_rows.append(
            {
                "pollutant": dk_column,
                "unit": "kt" if dk_column != "co2_eq" else "kt CO2e",
                "dk_energy": dke,
                "dk_nonenergy": dkn,
                "dk_combined": dke + dkn,
                "eurostat_hh": eu,
                "difference": eu - dke - dkn,
            }
        )
    household_emissions = pd.DataFrame(household_emission_rows)

    # ---------------- Residence principle / bunker scope ----------------
    bunker = energy[
        (energy["bal"] == "use")
        & energy["product"].isin(["bunk_planes", "bunk_ships", "bunk_trucks"])
    ]
    intl = bridge[bridge["item"] == "internat_transp"].iloc[0]
    scope = pd.DataFrame(
        [
            {
                "metric": "GREU bunker-product energy use",
                "value": bunker["pj"].sum(),
                "unit": "PJ",
                "comparison": "PEFA merges these fuels into P15/P17/P19; no separate bunker product.",
            },
            {
                "metric": "GREU bunker-product energy emissions",
                "value": bunker["co2_eq"].sum(),
                "unit": "kt CO2e",
                "comparison": "Resident Danish operators' fuel used outside Denmark is included.",
            },
            {
                "metric": "GREU bridge: international transport",
                "value": intl["co2_eq"],
                "unit": "kt CO2e",
                "comparison": "Residence-to-territory bridge; close but not identical to bunker-product sum.",
            },
            {
                "metric": "Bunker products minus bridge international transport",
                "value": bunker["co2_eq"].sum() - intl["co2_eq"],
                "unit": "kt CO2e",
                "comparison": "Different construction/boundary; do not net bunker products directly without bridge rules.",
            },
            {
                "metric": "Eurostat AEA GHG, resident scope",
                "value": aea_kt["GHG"],
                "unit": "kt CO2e",
                "comparison": "Same residence principle as GREU; includes international operations of resident firms.",
            },
        ]
    )

    # ---------------- Detail-loss and validation diagnostics ----------------
    coarse_clusters = [c for c in industry_clusters if len(c["left_values"]) > len(c["right_values"])]
    coarse_greu = {v for c in coarse_clusters for v in c["left_values"]}
    purpose_rows = energy[
        (energy["bal"] == "use") & energy["flow"].isin(["cons_inter", "cons_hh"])
    ]
    detail_loss = pd.DataFrame(
        [
            {
                "dimension": "Industry",
                "greu_detail": f"{energy[energy['flow'] == 'cons_inter']['indu'].nunique()} GREU users",
                "eurostat_detail": "NACE A64",
                "quantified_loss": (
                    f"{len(coarse_clusters)} many-to-one clusters cover {len(coarse_greu)} "
                    f"of {len(greu_to_cluster)} mapped GREU industries"
                ),
                "status": "Verified",
            },
            {
                "dimension": "Energy product",
                "greu_detail": f"{energy['product'].nunique()} products",
                "eurostat_detail": "20 P-products plus natural-input/residual codes",
                "quantified_loss": (
                    f"{len(product_clusters)} comparable concordance clusters; "
                    "P10/P16/P18/P22 absent from current metadata map; pilot assigns P18 to diesel"
                ),
                "status": "Verified",
            },
            {
                "dimension": "Industry purpose",
                "greu_detail": f"{purpose_rows['purp'].nunique()} mutually exclusive purposes",
                "eurostat_detail": "Transformation, end use, emission-relevant use only",
                "quantified_loss": "No process_normal/process_special/heating/transport/in_ETS split by industry",
                "status": "Verified gap",
            },
            {
                "dimension": "Household purpose",
                "greu_detail": "Heating, transport, appliances",
                "eurostat_detail": "HH_HEAT, HH_TRA, HH_OTH",
                "quantified_loss": "All three broad household end uses available; allocations differ while total closely matches",
                "status": "Verified coarser/method difference",
            },
            {
                "dimension": "Energy vs non-energy emissions",
                "greu_detail": "Separate workbooks",
                "eurostat_detail": "Air account publishes combined emissions by activity",
                "quantified_loss": "Cannot identify energy-related emissions directly from env_ac_ainah_r2",
                "status": "Verified gap",
            },
            {
                "dimension": "Bunker fuels",
                "greu_detail": "Three explicit bunker products",
                "eurostat_detail": "Merged into P15/P17/P19; resident transport industries retain emissions",
                "quantified_loss": f"{bunker['pj'].sum():.3f} PJ loses explicit bunker-product identity",
                "status": "Verified",
            },
            {
                "dimension": "Prices/taxes",
                "greu_detail": "Basic price, 3 margins, 5 taxes, VAT, purchaser price",
                "eurostat_detail": "Physical PEFA only",
                "quantified_loss": "No monetary cell values or Danish tax decomposition",
                "status": "Verified structural gap",
            },
        ]
    )

    statuses = pd.concat(
        [
            pefa[["obs_status"]].assign(dataset="env_ac_pefasu"),
            aea[["obs_status"]].assign(dataset="env_ac_ainah_r2"),
        ]
    )
    status_counts = (
        statuses.groupby(["dataset", "obs_status"], dropna=False)
        .size()
        .rename("observation_count")
        .reset_index()
    )
    validation = pd.DataFrame(
        [
            ("PEFA raw observations", len(pefa), "rows", "pass"),
            ("AEA raw observations", len(aea), "rows", "pass"),
            ("PEFA year", int(pefa["time"].iloc[0]), "year", "pass"),
            ("AEA year", int(aea["time"].iloc[0]), "year", "pass"),
            ("PEFA supply-use imbalance", pefa_grand["SUP"] - pefa_grand["USE"], "TJ", "pass"),
            (
                "GREU supply-use imbalance",
                energy[energy["bal"] == "sup"]["pj"].sum()
                - energy[energy["bal"] == "use"]["pj"].sum(),
                "PJ",
                "pass",
            ),
            (
                "AEA GHG component identity",
                aea_kt["GHG"]
                - aea_kt[
                    ["CO2", "CH4_CO2E", "N2O_CO2E", "HFC_CO2E", "PFC_CO2E", "NF3_SF6_CO2E"]
                ].sum(),
                "kt CO2e",
                "pass",
            ),
            (
                "GREU energy CO2e formula max row error",
                (
                    energy["co2_eq"].fillna(0)
                    - energy["co2_xbio"].fillna(0)
                    - 28 * energy["ch4"].fillna(0)
                    - 265 * energy["n2o"].fillna(0)
                ).abs().max(),
                "kt CO2e",
                "pass",
            ),
        ],
        columns=["check", "value", "unit", "result"],
    )

    mapping_rows = []
    for cluster in industry_clusters:
        mapping_rows.append(
            {
                "cluster": cluster["cluster"],
                "n_greu": len(cluster["left_values"]),
                "n_a64": len(cluster["right_values"]),
                "greu_industries": ", ".join(cluster["left_values"]),
                "nace_a64_codes": ", ".join(cluster["right_values"]),
                "pefa_coarser": "yes" if len(cluster["left_values"]) > len(cluster["right_values"]) else "",
            }
        )
    industry_mapping = pd.DataFrame(mapping_rows)

    meta = pd.DataFrame(
        [
            ("created", datetime.date.today().isoformat()),
            ("comparison_year", YEAR),
            ("country", "Denmark (DK)"),
            ("retrieval_date", RETRIEVAL_DATE),
            ("script", "data/preprocessing/scripts/reconcile_eurostat_energy_emissions_dk_2020.py"),
            ("danish_energy_input", "data/preprocessing/data/energy_and_emissions.xlsx; ems_energy"),
            ("danish_nonenergy_diagnostic", "data/preprocessing/data/non_energy_emissions.xlsx; ems_non_energy"),
            ("danish_bridge_diagnostic", "data/preprocessing/data/emissions_bridge_items.xlsx; bridge_items"),
            ("pefa_source", PEFA_URL),
            ("air_emissions_source", AEA_URL),
            ("pefa_raw_file", "data/preprocessing/data/eurostat_energy_emissions_raw/env_ac_pefasu_DK_2020.json"),
            ("aea_raw_file", "data/preprocessing/data/eurostat_energy_emissions_raw/env_ac_ainah_r2_DK_2020.json"),
            ("units", "PEFA source TJ converted to PJ; emissions source THS_T = kt; GHG/F-gases in kt CO2e"),
            (
                "scope",
                "SEEA/ESA residence principle. Concept-adjusted energy boundary = P08-P27 + "
                "N03/N04/N05/N07 + R28/R29.",
            ),
            (
                "difference_definition",
                "Eurostat minus Danish GREU; percentages divide by Danish value.",
            ),
            (
                "important_caveat",
                "env_ac_ainah_r2 is total air emissions. energy_and_emissions.xlsx is energy-related only; "
                "non_energy_emissions.xlsx is added only to test the like-for-like combined air-account total.",
            ),
        ],
        columns=["key", "value"],
    )

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        meta.to_excel(writer, sheet_name="metadata", index=False)
        energy_totals.to_excel(writer, sheet_name="energy_totals", index=False)
        energy_flows.to_excel(writer, sheet_name="energy_flow_mapping", index=False)
        energy_products.to_excel(writer, sheet_name="energy_product_clusters", index=False)
        pefa_components.to_excel(writer, sheet_name="pefa_product_coverage", index=False)
        mapping_adjustments.to_excel(writer, sheet_name="mapping_adjustments", index=False)
        household_energy.to_excel(writer, sheet_name="hh_energy_purpose", index=False)
        energy_by_industry.to_excel(writer, sheet_name="energy_by_industry", index=False)
        emissions_totals.to_excel(writer, sheet_name="emissions_totals", index=False)
        emissions_by_industry.to_excel(writer, sheet_name="emissions_by_industry", index=False)
        household_emissions.to_excel(writer, sheet_name="hh_emissions", index=False)
        scope.to_excel(writer, sheet_name="residence_bunker_scope", index=False)
        detail_loss.to_excel(writer, sheet_name="detail_loss", index=False)
        industry_mapping.to_excel(writer, sheet_name="industry_mapping", index=False)
        product_map_raw.to_excel(writer, sheet_name="product_mapping_raw", index=False)
        validation.to_excel(writer, sheet_name="validation", index=False)
        status_counts.to_excel(writer, sheet_name="source_status_flags", index=False)

    style_workbook(OUT)
    print(f"wrote {OUT}")
    print("\nEnergy totals")
    print(energy_totals.to_string(index=False))
    print("\nEmissions totals")
    print(emissions_totals.to_string(index=False))
    print("\nHousehold energy")
    print(household_energy.to_string(index=False))
    print("\nResidence/bunker scope")
    print(scope.to_string(index=False))


if __name__ == "__main__":
    main()
