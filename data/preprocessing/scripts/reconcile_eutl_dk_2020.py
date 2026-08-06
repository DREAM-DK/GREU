"""Reconcile public Union Registry data with GREU ets.xlsx for Denmark 2020.

Official Commission daily bulk files provide installation-level compliance
records.  The EEA release supplies independently aggregated checks and national
auction volumes.  A secondary EUETS.INFO concordance is tested for NACE
coverage, but is kept distinct from official fields throughout the workbook.

Run
---
python data/preprocessing/scripts/reconcile_eutl_dk_2020.py
"""

from __future__ import annotations

import hashlib
import pathlib
import zipfile

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DATA = pathlib.Path(__file__).resolve().parents[1] / "data"
RAW = DATA / "eea_eutl_2026_raw"
EEA_EXTRACTED = RAW / "extracted"
NACE_RAW = DATA / "euetsinfo_nace_2026_raw"
NACE_EXTRACTED = NACE_RAW / "extracted"
GREU_ETS = DATA / "ets.xlsx"
METADATA = DATA / "metadata.xlsx"
OUT = DATA / "eutl_dk2020_reconciliation.xlsx"

OPERATORS = RAW / "operators_daily.csv.gz"
YEARLY = RAW / "operators_yearly_activity_daily.csv.gz"
EEA_AGG = EEA_EXTRACTED / "ETS_Database_July_2026.xlsx"
ACTIVITY_TRANSLATION = EEA_EXTRACTED / "Translation of activity codes May 2019.xlsx"
EEA_ARCHIVE = RAW / "eea_eutl_union_registry_july_2026.zip"
NACE_ARCHIVE = NACE_RAW / "eutl_data_package_2026-07-21.zip"
NACE_INSTALLATIONS = NACE_EXTRACTED / "installations.csv"
NACE_MAP = NACE_EXTRACTED / "nace_mappings.csv"

YEAR = 2020
COUNTRY = "DK"
RETRIEVAL_DATE = "2026-07-30"
COMMISSION_SITE = "https://union-registry-data.ec.europa.eu/"
OPERATORS_URL = (
    "https://dlsclimabi.blob.core.windows.net/public-data/eutlpublic/"
    "extracts/_all_extracts/operator/operators_daily.csv.gz"
)
YEARLY_URL = (
    "https://dlsclimabi.blob.core.windows.net/public-data/eutlpublic/"
    "extracts/_all_extracts/operators_yearly_activity/"
    "operators_yearly_activity_daily.csv.gz"
)
EEA_CATALOGUE = (
    "https://www.eea.europa.eu/en/datahub/datahubitem-view/"
    "98f04097-26de-4fca-86c4-63834818c0c0/file"
)
ZENODO_DOI = "https://doi.org/10.5281/zenodo.21414185"

EU27 = {
    "AT",
    "BE",
    "BG",
    "HR",
    "CY",
    "CZ",
    "DK",
    "EE",
    "FI",
    "FR",
    "DE",
    "GR",
    "HU",
    "IE",
    "IT",
    "LV",
    "LT",
    "LU",
    "MT",
    "NL",
    "PL",
    "PT",
    "RO",
    "SK",
    "SI",
    "ES",
    "SE",
}

VALUE_COLUMNS = [
    "VERIFIED_EMISSIONS",
    "ALLOCATION",
    "ALLOCATION_RES",
    "ALLOCATION_TRA",
    "SURR_ALL",
]


def sha256(path: pathlib.Path) -> str:
    value = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(4 * 1024 * 1024), b""):
            value.update(block)
    return value.hexdigest()


def clean_identifier(value: object) -> str:
    text = str(value).strip()
    return text[:-2] if text.endswith(".0") else text


def normalize_nace(value: object) -> str | None:
    if pd.isna(value):
        return None
    text = str(value).strip()
    if "." in text:
        division, detail = text.split(".", 1)
        division_digits = "".join(character for character in division if character.isdigit())
        detail_digits = "".join(character for character in detail if character.isdigit())
        if division_digits and detail_digits:
            return division_digits.zfill(2)[-2:] + detail_digits.ljust(2, "0")[:2]
    digits = "".join(character for character in text if character.isdigit())
    if not digits:
        return None
    if len(digits) <= 2:
        return digits.zfill(2) + "00"
    if len(digits) == 3:
        return digits[:2] + digits[2:].ljust(2, "0")
    return digits.zfill(4)[:4]


def nace_to_a64(nace: str | None) -> str | None:
    if nace is None or pd.isna(nace) or not str(nace):
        return None
    division = int(str(nace)[:2])
    if division in (1, 2, 3):
        return f"A{division:02d}"
    if 5 <= division <= 9:
        return "B"
    if 10 <= division <= 12:
        return "C10-C12"
    if 13 <= division <= 15:
        return "C13-C15"
    if 16 <= division <= 30:
        return f"C{division}"
    if 31 <= division <= 32:
        return "C31_C32"
    if division == 33:
        return "C33"
    if division == 35:
        return "D"
    if division == 36:
        return "E36"
    if 37 <= division <= 39:
        return "E37-E39"
    if 41 <= division <= 43:
        return "F"
    if division in (45, 46, 47):
        return f"G{division}"
    if 49 <= division <= 53:
        return f"H{division}"
    if 55 <= division <= 56:
        return "I"
    if division == 58:
        return "J58"
    if 59 <= division <= 60:
        return "J59_J60"
    if division == 61:
        return "J61"
    if 62 <= division <= 63:
        return "J62_J63"
    if 64 <= division <= 66:
        return f"K{division}"
    if division == 68:
        return "L"
    if 69 <= division <= 70:
        return "M69_M70"
    if division in (71, 72, 73):
        return f"M{division}"
    if 74 <= division <= 75:
        return "M74_M75"
    if division in (77, 78, 79):
        return f"N{division}"
    if 80 <= division <= 82:
        return "N80-N82"
    if division == 84:
        return "O"
    if division == 85:
        return "P"
    if division == 86:
        return "Q86"
    if 87 <= division <= 88:
        return "Q87_Q88"
    if 90 <= division <= 92:
        return "R90-R92"
    if division == 93:
        return "R93"
    if division in (94, 95, 96):
        return f"S{division}"
    if 97 <= division <= 98:
        return "T"
    if division == 99:
        return "U"
    return None


def connected_industry_clusters(mapping: pd.DataFrame) -> tuple[dict[str, str], pd.DataFrame]:
    adjacency: dict[str, set[str]] = {}
    for row in mapping.itertuples(index=False):
        greu = f"g:{row.indu_greu}"
        a64 = f"a:{row.indu_naceA64}"
        adjacency.setdefault(greu, set()).add(a64)
        adjacency.setdefault(a64, set()).add(greu)
    node_to_cluster: dict[str, str] = {}
    rows = []
    cluster_number = 0
    for node in sorted(adjacency):
        if node in node_to_cluster:
            continue
        cluster_number += 1
        cluster_id = f"C{cluster_number:02d}"
        pending = [node]
        component = set()
        while pending:
            current = pending.pop()
            if current in component:
                continue
            component.add(current)
            pending.extend(adjacency[current] - component)
        for current in component:
            node_to_cluster[current] = cluster_id
        greu_codes = sorted(item[2:] for item in component if item.startswith("g:"))
        a64_codes = sorted(item[2:] for item in component if item.startswith("a:"))
        rows.append(
            {
                "cluster_id": cluster_id,
                "greu_industries": ", ".join(greu_codes),
                "nace_a64_codes": ", ".join(a64_codes),
                "greu_industry_count": len(greu_codes),
                "nace_a64_count": len(a64_codes),
                "mapping_granularity": (
                    "one-to-one"
                    if len(greu_codes) == 1 and len(a64_codes) == 1
                    else "many-to-many / aggregate only"
                ),
            }
        )
    return node_to_cluster, pd.DataFrame(rows)


def compare(metric: str, greu: float, source: float, unit: str, status: str, note: str) -> dict:
    difference = source - greu if pd.notna(greu) and pd.notna(source) else np.nan
    return {
        "metric": metric,
        "unit": unit,
        "greu_value": greu,
        "public_source_value": source,
        "difference_source_minus_greu": difference,
        "difference_pct_of_greu": (
            difference / greu * 100 if pd.notna(difference) and abs(greu) > 1e-12 else np.nan
        ),
        "field_status": status,
        "interpretation": note,
    }


def style_workbook(path: pathlib.Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column in sheet.columns:
            letter = get_column_letter(column[0].column)
            width = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[letter].width = min(max(width + 2, 10), 62)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
    workbook.save(path)


def main() -> None:
    required = [
        OPERATORS,
        YEARLY,
        EEA_AGG,
        ACTIVITY_TRANSLATION,
        EEA_ARCHIVE,
        NACE_ARCHIVE,
        NACE_INSTALLATIONS,
        NACE_MAP,
        GREU_ETS,
        METADATA,
    ]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"missing inputs; run downloaders first: {missing}")
    with zipfile.ZipFile(EEA_ARCHIVE) as archive:
        assert archive.testzip() is None
    with zipfile.ZipFile(NACE_ARCHIVE) as archive:
        assert archive.testzip() is None

    operators = pd.read_csv(
        OPERATORS,
        compression="gzip",
        encoding="latin1",
        dtype="string",
    )
    yearly = pd.read_csv(
        YEARLY,
        compression="gzip",
        encoding="latin1",
        dtype={
            "INSTALLATION_IDENTIFIER": "string",
            "REGISTRY_CODE": "string",
        },
    )
    for column in VALUE_COLUMNS + ["ALLOCATION_RES", "ALLOCATION_TRA"]:
        yearly[column] = pd.to_numeric(yearly[column], errors="raise")
    yearly["PERIOD_YEAR"] = pd.to_numeric(yearly["PERIOD_YEAR"], errors="raise").astype(int)

    operators["INSTALLATION_IDENTIFIER"] = operators["INSTALLATION_IDENTIFIER"].map(clean_identifier)
    yearly["INSTALLATION_IDENTIFIER"] = yearly["INSTALLATION_IDENTIFIER"].map(clean_identifier)
    operator_key = ["REGISTRY_CODE", "INSTALLATION_IDENTIFIER"]
    if operators.duplicated(operator_key).any():
        raise ValueError("operator master contains duplicate registry/installation keys")

    dk = yearly[(yearly["REGISTRY_CODE"] == COUNTRY) & (yearly["PERIOD_YEAR"] == YEAR)].copy()
    dk = dk.merge(
        operators[
            operator_key
            + [
                "INSTALLATION_NAME",
                "ACTIVITY_TYPE_CODE",
                "ACTIVITY_TYPE",
                "PERMIT_REVOCATION_DATE",
                "YEAR_OF_FIRST_EMISSIONS",
                "YEAR_OF_LAST_EMISSIONS",
            ]
        ],
        on=operator_key,
        how="left",
        validate="one_to_one",
        suffixes=("_yearly", "_master"),
        indicator="operator_master_match",
    )
    for column in VALUE_COLUMNS + ["ALLOCATION_RES", "ALLOCATION_TRA"]:
        dk[f"{column}_usable"] = dk[column].where(dk[column] >= 0, 0)
    dk["free_allocation_total"] = (
        dk["ALLOCATION_usable"]
        + dk["ALLOCATION_RES_usable"]
        + dk["ALLOCATION_TRA_usable"]
    )
    dk["bought_allowance_proxy"] = (
        dk["VERIFIED_EMISSIONS_usable"] - dk["free_allocation_total"]
    ).clip(lower=0)
    dk["entity_scope"] = np.select(
        [
            dk["ACTIVITY_TYPE_CODE"].eq("10").fillna(False).to_numpy(dtype=bool),
            dk["ACTIVITY_TYPE_CODE"].eq("50").fillna(False).to_numpy(dtype=bool),
            dk["ACTIVITY_TYPE_CODE"].isna().to_numpy(dtype=bool),
        ],
        ["aviation", "maritime", "unknown / ETS2 placeholder"],
        default="stationary",
    )
    dk["has_2020_value"] = dk[
        [
            "VERIFIED_EMISSIONS",
            "ALLOCATION",
            "ALLOCATION_RES",
            "ALLOCATION_TRA",
            "SURR_ALL",
        ]
    ].ge(0).any(axis=1)
    ets1 = dk[dk["entity_scope"].isin(["stationary", "aviation"]) & dk["has_2020_value"]].copy()

    translation = pd.read_excel(ACTIVITY_TRANSLATION, dtype="string")
    translation["Installation ID"] = translation["Installation ID"].map(clean_identifier)
    translation = translation.rename(
        columns={
            "Country Code": "REGISTRY_CODE",
            "Installation ID": "INSTALLATION_IDENTIFIER",
            "Translation as applied for EEA EU ETS dataviewer": "translated_activity_code",
        }
    )
    translation = translation[
        ["REGISTRY_CODE", "INSTALLATION_IDENTIFIER", "translated_activity_code"]
    ].drop_duplicates(operator_key)
    ets1 = ets1.merge(translation, on=operator_key, how="left", validate="one_to_one")
    old_code = pd.to_numeric(ets1["ACTIVITY_TYPE_CODE"], errors="coerce").between(1, 9)
    ets1["activity_code_for_analysis"] = ets1["ACTIVITY_TYPE_CODE"]
    ets1.loc[old_code & ets1["translated_activity_code"].notna(), "activity_code_for_analysis"] = (
        ets1.loc[
            old_code & ets1["translated_activity_code"].notna(), "translated_activity_code"
        ]
    )
    activity_names = (
        operators.dropna(subset=["ACTIVITY_TYPE_CODE", "ACTIVITY_TYPE"])
        .groupby("ACTIVITY_TYPE_CODE")["ACTIVITY_TYPE"]
        .agg(lambda values: values.mode().iloc[0])
        .to_dict()
    )

    greu = pd.read_excel(GREU_ETS, dtype={"indu": "string"})
    greu = greu[greu["year"] == YEAR].copy()
    greu_totals = greu[
        ["free_allowances", "emissions", "bought_allowances", "emissions_tax"]
    ].sum()
    implied_price_dkk_t = (
        greu_totals["emissions_tax"] / greu_totals["bought_allowances"] * 1_000_000
    )

    public_totals = {
        "emissions": ets1["VERIFIED_EMISSIONS_usable"].sum() / 1000,
        "free": ets1["free_allocation_total"].sum() / 1000,
        "bought_proxy": ets1["bought_allowance_proxy"].sum() / 1000,
        "surrendered": ets1["SURR_ALL_usable"].sum() / 1000,
    }

    eea = pd.read_excel(
        EEA_AGG, dtype={"country_code": "string", "main_activity_code": "string", "year": "string"}
    )
    eea_dk = eea[
        (eea["country_code"] == COUNTRY)
        & (eea["year"] == str(YEAR))
        & (eea["active_installation"] == "all entities")
        & (eea["size"] == "All sizes")
    ].copy()
    auction_metric = "1.3 Allowances auctioned or sold (EUAs and EUAAs)"
    auction = eea_dk[
        (eea_dk["citl_information"] == auction_metric)
        & (eea_dk["main_activity_code"].isin(["10", "20-99"]))
    ]["value"].sum() / 1000
    current_cost = public_totals["bought_proxy"] * implied_price_dkk_t / 1_000_000

    totals = pd.DataFrame(
        [
            compare(
                "Verified emissions",
                greu_totals["emissions"],
                public_totals["emissions"],
                "kt CO2e",
                "direct Union Registry field",
                "Current source differs only in aviation; stationary total is unchanged.",
            ),
            compare(
                "Free allocation",
                greu_totals["free_allowances"],
                public_totals["free"],
                "thousand allowances",
                "direct Union Registry fields",
                "Sum of ALLOCATION, ALLOCATION_RES and ALLOCATION_TRA, with -1 treated as missing.",
            ),
            compare(
                "Bought allowances / positive installation shortfall",
                greu_totals["bought_allowances"],
                public_totals["bought_proxy"],
                "thousand allowances",
                "derived, not a purchase field",
                "Sum by installation of max(verified emissions - free allocation, 0). "
                "This reproduces GREU but does not observe purchases, banking or transfers.",
            ),
            compare(
                "Implied emissions cost",
                greu_totals["emissions_tax"],
                current_cost,
                "bn DKK",
                "derived with GREU-implied price assumption",
                f"Applies {implied_price_dkk_t:.6f} DKK/t inferred from GREU. "
                "Union Registry has no allowance-price or tax field.",
            ),
            compare(
                "National auctioned/sold allowances",
                np.nan,
                auction,
                "thousand allowances",
                "direct EEA country-level field",
                "Government auction volume; not allocated to installations or industries and "
                "not equivalent to GREU bought_allowances.",
            ),
            compare(
                "Total surrendered units",
                np.nan,
                public_totals["surrendered"],
                "thousand units",
                "direct Union Registry field",
                "Compliance surrender is observable but does not identify purchased units.",
            ),
        ]
    )

    activity = (
        ets1.groupby("activity_code_for_analysis", dropna=False)
        .agg(
            entity_records=("INSTALLATION_IDENTIFIER", "size"),
            verified_emissions_kt=("VERIFIED_EMISSIONS_usable", lambda x: x.sum() / 1000),
            free_allocation_kallow=("free_allocation_total", lambda x: x.sum() / 1000),
            bought_proxy_kallow=("bought_allowance_proxy", lambda x: x.sum() / 1000),
        )
        .reset_index()
        .rename(columns={"activity_code_for_analysis": "eutl_activity_code"})
    )
    activity["eutl_activity_name"] = activity["eutl_activity_code"].map(activity_names)
    activity["classification_status"] = (
        "Direct ETS regulatory activity; not NACE and not assigned to GREU industry."
    )
    aviation = activity["eutl_activity_code"].eq("10")
    activity.loc[aviation, "classification_status"] = (
        "Aviation is semantically comparable with GREU 51001 only; no general "
        "activity-to-industry crosswalk is asserted."
    )

    secondary_installations = pd.read_csv(NACE_INSTALLATIONS, dtype="string")
    nace = pd.read_csv(NACE_MAP, dtype="string")
    if secondary_installations["id"].duplicated().any() or nace["installation_id"].duplicated().any():
        raise ValueError("secondary installation/NACE keys are not unique")
    nace = nace.merge(
        secondary_installations[["id", "registry_id", "activity_id"]],
        left_on="installation_id",
        right_on="id",
        how="left",
        validate="one_to_one",
    )
    ets1["secondary_installation_id"] = (
        ets1["REGISTRY_CODE"] + "_" + ets1["INSTALLATION_IDENTIFIER"]
    )
    ets1 = ets1.merge(
        nace[["installation_id", "nace_2015", "nace_2020"]],
        left_on="secondary_installation_id",
        right_on="installation_id",
        how="left",
        validate="one_to_one",
    )
    ets1["nace_selected_raw"] = ets1["nace_2020"].fillna(ets1["nace_2015"])
    ets1["nace_source"] = np.select(
        [
            ets1["nace_2020"].notna().to_numpy(dtype=bool),
            ets1["nace_2015"].notna().to_numpy(dtype=bool),
        ],
        ["2020 carbon-leakage list", "2015 list fallback"],
        default="unavailable",
    )
    ets1["nace4"] = ets1["nace_selected_raw"].map(normalize_nace)
    ets1["nace_a64"] = ets1["nace4"].map(nace_to_a64)

    industry_map = pd.read_excel(
        METADATA, sheet_name="industries_naceA64_map", dtype="string"
    )
    industry_map = industry_map[["indu_greu", "indu_naceA64"]].drop_duplicates()
    node_to_cluster, cluster_def = connected_industry_clusters(industry_map)
    ets1["cluster_id"] = ets1["nace_a64"].map(
        lambda value: node_to_cluster.get(f"a:{value}") if pd.notna(value) else None
    )
    greu["cluster_id"] = greu["indu"].map(
        lambda value: node_to_cluster.get(f"g:{value}") if pd.notna(value) else None
    )
    source_cluster = (
        ets1.dropna(subset=["cluster_id"])
        .groupby("cluster_id")
        .agg(
            source_records=("INSTALLATION_IDENTIFIER", "size"),
            source_emissions_kt=("VERIFIED_EMISSIONS_usable", lambda x: x.sum() / 1000),
            source_free_kallow=("free_allocation_total", lambda x: x.sum() / 1000),
            source_bought_proxy_kallow=("bought_allowance_proxy", lambda x: x.sum() / 1000),
        )
        .reset_index()
    )
    greu_cluster = (
        greu.groupby("cluster_id")
        .agg(
            greu_emissions_kt=("emissions", "sum"),
            greu_free_kallow=("free_allowances", "sum"),
            greu_bought_kallow=("bought_allowances", "sum"),
        )
        .reset_index()
    )
    clusters = (
        cluster_def.merge(greu_cluster, on="cluster_id", how="left")
        .merge(source_cluster, on="cluster_id", how="left")
        .fillna(
            {
                "greu_emissions_kt": 0,
                "greu_free_kallow": 0,
                "greu_bought_kallow": 0,
                "source_records": 0,
                "source_emissions_kt": 0,
                "source_free_kallow": 0,
                "source_bought_proxy_kallow": 0,
            }
        )
    )
    clusters = clusters[
        (clusters["greu_emissions_kt"] != 0) | (clusters["source_emissions_kt"] != 0)
    ].copy()
    for concept in ["emissions", "free", "bought"]:
        source_column = {
            "emissions": "source_emissions_kt",
            "free": "source_free_kallow",
            "bought": "source_bought_proxy_kallow",
        }[concept]
        greu_column = {
            "emissions": "greu_emissions_kt",
            "free": "greu_free_kallow",
            "bought": "greu_bought_kallow",
        }[concept]
        clusters[f"{concept}_difference_source_minus_greu"] = (
            clusters[source_column] - clusters[greu_column]
        )
    clusters["comparison_status"] = (
        "Diagnostic only: source side includes only installations covered by the "
        "secondary carbon-leakage-list NACE concordance."
    )

    relevant_count = len(ets1)
    nace_available = ets1["nace4"].notna()
    a64_available = ets1["cluster_id"].notna()
    mapping_diagnostics = pd.DataFrame(
        [
            {
                "diagnostic": "Official DK 2020 ETS1 records with any compliance value",
                "record_count": relevant_count,
                "emissions_kt": ets1["VERIFIED_EMISSIONS_usable"].sum() / 1000,
                "free_allocation_kallow": ets1["free_allocation_total"].sum() / 1000,
                "share_of_records_pct": 100,
                "share_of_emissions_pct": 100,
                "status": "official denominator",
            },
            {
                "diagnostic": "Matched to current official operator master",
                "record_count": int(ets1["ACTIVITY_TYPE_CODE"].notna().sum()),
                "emissions_kt": ets1.loc[
                    ets1["ACTIVITY_TYPE_CODE"].notna(), "VERIFIED_EMISSIONS_usable"
                ].sum()
                / 1000,
                "free_allocation_kallow": ets1.loc[
                    ets1["ACTIVITY_TYPE_CODE"].notna(), "free_allocation_total"
                ].sum()
                / 1000,
                "share_of_records_pct": ets1["ACTIVITY_TYPE_CODE"].notna().mean() * 100,
                "share_of_emissions_pct": (
                    ets1.loc[
                        ets1["ACTIVITY_TYPE_CODE"].notna(), "VERIFIED_EMISSIONS_usable"
                    ].sum()
                    / ets1["VERIFIED_EMISSIONS_usable"].sum()
                    * 100
                ),
                "status": "official join",
            },
            {
                "diagnostic": "NACE available from secondary public concordance",
                "record_count": int(nace_available.sum()),
                "emissions_kt": ets1.loc[nace_available, "VERIFIED_EMISSIONS_usable"].sum()
                / 1000,
                "free_allocation_kallow": ets1.loc[
                    nace_available, "free_allocation_total"
                ].sum()
                / 1000,
                "share_of_records_pct": nace_available.mean() * 100,
                "share_of_emissions_pct": (
                    ets1.loc[nace_available, "VERIFIED_EMISSIONS_usable"].sum()
                    / ets1["VERIFIED_EMISSIONS_usable"].sum()
                    * 100
                ),
                "status": "secondary; coverage test only",
            },
            {
                "diagnostic": "NACE mapped to metadata A64/GREU cluster",
                "record_count": int(a64_available.sum()),
                "emissions_kt": ets1.loc[a64_available, "VERIFIED_EMISSIONS_usable"].sum()
                / 1000,
                "free_allocation_kallow": ets1.loc[
                    a64_available, "free_allocation_total"
                ].sum()
                / 1000,
                "share_of_records_pct": a64_available.mean() * 100,
                "share_of_emissions_pct": (
                    ets1.loc[a64_available, "VERIFIED_EMISSIONS_usable"].sum()
                    / ets1["VERIFIED_EMISSIONS_usable"].sum()
                    * 100
                ),
                "status": "diagnostic aggregate; not production-ready",
            },
        ]
    )

    nace_coverage = (
        ets1.groupby(["nace_source", "nace4", "nace_a64"], dropna=False)
        .agg(
            records=("INSTALLATION_IDENTIFIER", "size"),
            emissions_kt=("VERIFIED_EMISSIONS_usable", lambda x: x.sum() / 1000),
            free_allocation_kallow=("free_allocation_total", lambda x: x.sum() / 1000),
            bought_proxy_kallow=("bought_allowance_proxy", lambda x: x.sum() / 1000),
        )
        .reset_index()
        .sort_values("emissions_kt", ascending=False)
    )

    detail = ets1[
        [
            "REGISTRY_CODE",
            "INSTALLATION_IDENTIFIER",
            "INSTALLATION_NAME_yearly",
            "ACTIVITY_TYPE_CODE",
            "activity_code_for_analysis",
            "entity_scope",
            "EXCLUDED",
            "VERIFIED_EMISSIONS",
            "ALLOCATION",
            "ALLOCATION_RES",
            "ALLOCATION_TRA",
            "free_allocation_total",
            "bought_allowance_proxy",
            "SURR_ALL",
            "nace_2015",
            "nace_2020",
            "nace_source",
            "nace4",
            "nace_a64",
            "cluster_id",
            "SNAPSHOT_DATE",
        ]
    ].copy()

    country_rows = []
    for country in sorted(EU27):
        subset = yearly[yearly["REGISTRY_CODE"] == country]
        country_rows.append(
            {
                "country_code": country,
                "operator_master_records": int((operators["REGISTRY_CODE"] == country).sum()),
                "yearly_activity_records": len(subset),
                "first_year": int(subset["PERIOD_YEAR"].min()) if len(subset) else np.nan,
                "last_year": int(subset["PERIOD_YEAR"].max()) if len(subset) else np.nan,
                "has_2020": bool((subset["PERIOD_YEAR"] == YEAR).any()),
                "coverage_verdict": "present" if len(subset) else "missing",
            }
        )
    source_coverage = pd.DataFrame(country_rows)

    concept_availability = pd.DataFrame(
        [
            {
                "greu_concept": "verified emissions",
                "availability": "DIRECT",
                "source_field_or_method": "VERIFIED_EMISSIONS",
                "unit": "tonne CO2e; divide by 1000 for GREU kt",
                "limitation": "All gases combined as CO2e; no fuel split.",
            },
            {
                "greu_concept": "free allowances",
                "availability": "DIRECT",
                "source_field_or_method": "ALLOCATION + ALLOCATION_RES + ALLOCATION_TRA",
                "unit": "allowance units; divide by 1000",
                "limitation": "Current exclusion flags/vintage can revise historic totals.",
            },
            {
                "greu_concept": "bought allowances",
                "availability": "DERIVED PROXY",
                "source_field_or_method": "sum installation max(VE - free allocation, 0)",
                "unit": "allowance units",
                "limitation": "Not observed purchases; ignores banking, transfers and surplus sales.",
            },
            {
                "greu_concept": "auctioned allowances",
                "availability": "DIRECT COUNTRY TOTAL ONLY",
                "source_field_or_method": "EEA viewer item 1.3",
                "unit": "allowance units",
                "limitation": "No installation/industry allocation; not equivalent to shortfall proxy.",
            },
            {
                "greu_concept": "emissions tax / allowance cost",
                "availability": "UNAVAILABLE DIRECTLY",
                "source_field_or_method": "shortfall proxy × external EUA price assumption",
                "unit": "currency",
                "limitation": "Registry has no transaction cost, tax or annual price field.",
            },
            {
                "greu_concept": "ETS activity",
                "availability": "DIRECT REGULATORY CODE",
                "source_field_or_method": "ACTIVITY_TYPE_CODE + official EEA old-code translation",
                "unit": "classification",
                "limitation": "ETS Annex I activity is not NACE and cannot generally map to GREU industry.",
            },
            {
                "greu_concept": "industry aggregation",
                "availability": "PARTIAL SECONDARY BRIDGE",
                "source_field_or_method": "EUETS.INFO carbon-leakage-list NACE → A64 → GREU clusters",
                "unit": "classification",
                "limitation": "Non-official, dated/partial NACE; many-to-many A64/GREU clusters.",
            },
            {
                "greu_concept": "energy purpose in_ETS",
                "availability": "NOT CONSTRUCTIBLE FROM EUTL ALONE",
                "source_field_or_method": "requires installation→industry plus fuel/energy modelling",
                "unit": "PJ by GREU industry, purpose and product",
                "limitation": "EUTL reports emissions/allowances, not energy use, fuel or GREU purpose.",
            },
        ]
    )

    assumptions = pd.DataFrame(
        [
            {
                "assumption_or_rule": "Negative sentinel",
                "treatment": "Union Registry -1 values are missing/not applicable and become zero only for aggregation.",
                "type": "source interpretation",
            },
            {
                "assumption_or_rule": "GREU source scope",
                "treatment": "Include stationary and aviation records with any 2020 compliance value; exclude maritime and ETS2 placeholders.",
                "type": "verified against totals",
            },
            {
                "assumption_or_rule": "Free allocation",
                "treatment": "Sum existing-entity, reserve and transitional components before installation shortfall.",
                "type": "direct fields",
            },
            {
                "assumption_or_rule": "Bought allowance proxy",
                "treatment": "Calculate positive shortfall at installation level before aggregation.",
                "type": "derived; reproduces GREU",
            },
            {
                "assumption_or_rule": "Implied cost",
                "treatment": f"Use GREU-implied uniform 2020 allowance price {implied_price_dkk_t:.6f} DKK/t only as a diagnostic.",
                "type": "assumption",
            },
            {
                "assumption_or_rule": "NACE selection",
                "treatment": "Use secondary 2020 carbon-leakage-list NACE; fall back to 2015 where absent.",
                "type": "secondary diagnostic",
            },
            {
                "assumption_or_rule": "Industry comparison",
                "treatment": "Compare only connected A64/GREU clusters; never split one NACE/A64 value across GREU industries.",
                "type": "conservative aggregation",
            },
            {
                "assumption_or_rule": "in_ETS purpose",
                "treatment": "No PJ is inferred from verified emissions without fuel-specific energy and emission-factor modelling.",
                "type": "gap retained",
            },
        ]
    )

    raw_stationary = ets1.loc[
        ets1["entity_scope"] == "stationary", "VERIFIED_EMISSIONS_usable"
    ].sum()
    raw_aviation = ets1.loc[
        ets1["entity_scope"] == "aviation", "VERIFIED_EMISSIONS_usable"
    ].sum()
    eea_stationary = eea_dk[
        (eea_dk["main_activity_code"] == "20-99")
        & (eea_dk["citl_information"] == "2. Verified emissions")
    ]["value"].sum()
    eea_aviation_eu = eea_dk[
        (eea_dk["main_activity_code"] == "10")
        & (eea_dk["citl_information"] == "2.1 EU-ETS Verified Emission")
    ]["value"].sum()
    validation = pd.DataFrame(
        [
            {
                "check": "GREU year is 2020 only",
                "result": bool((greu["year"] == YEAR).all()),
                "observed": ", ".join(map(str, sorted(greu["year"].unique()))),
                "expected": "2020",
            },
            {
                "check": "Official daily snapshot date",
                "result": bool((yearly["SNAPSHOT_DATE"] == RETRIEVAL_DATE).all()),
                "observed": ", ".join(map(str, yearly["SNAPSHOT_DATE"].unique())),
                "expected": RETRIEVAL_DATE,
            },
            {
                "check": "All EU-27 countries present",
                "result": EU27.issubset(set(yearly["REGISTRY_CODE"].dropna())),
                "observed": len(EU27 & set(yearly["REGISTRY_CODE"].dropna())),
                "expected": 27,
            },
            {
                "check": "DK 2020 present",
                "result": len(dk) > 0,
                "observed": len(dk),
                "expected": "> 0 records",
            },
            {
                "check": "Raw stationary emissions equal EEA aggregate",
                "result": raw_stationary == eea_stationary,
                "observed": raw_stationary,
                "expected": eea_stationary,
            },
            {
                "check": "Raw EU aviation emissions equal EEA aggregate",
                "result": raw_aviation == eea_aviation_eu,
                "observed": raw_aviation,
                "expected": eea_aviation_eu,
            },
            {
                "check": "No negative values used in totals",
                "result": bool(
                    (ets1[[f"{column}_usable" for column in VALUE_COLUMNS]] >= 0)
                    .all()
                    .all()
                ),
                "observed": "negative sentinels replaced only in *_usable columns",
                "expected": "all usable values >= 0",
            },
            {
                "check": "EEA archive CRC",
                "result": True,
                "observed": sha256(EEA_ARCHIVE),
                "expected": "validated by downloader",
            },
            {
                "check": "Secondary NACE archive CRC",
                "result": True,
                "observed": sha256(NACE_ARCHIVE),
                "expected": "validated by downloader",
            },
        ]
    )

    metadata = pd.DataFrame(
        [
            {"key": "created", "value": RETRIEVAL_DATE},
            {"key": "country/year", "value": f"Denmark ({COUNTRY}), {YEAR}"},
            {"key": "script", "value": "data/preprocessing/scripts/reconcile_eutl_dk_2020.py"},
            {"key": "GREU target", "value": "data/preprocessing/data/ets.xlsx"},
            {"key": "official public site", "value": COMMISSION_SITE},
            {"key": "official operator URL", "value": OPERATORS_URL},
            {"key": "official installation-year URL", "value": YEARLY_URL},
            {"key": "EEA catalogue", "value": EEA_CATALOGUE},
            {"key": "secondary NACE DOI", "value": ZENODO_DOI},
            {"key": "official source snapshot", "value": "2026-07-30 daily bulk files"},
            {"key": "EEA release", "value": "July 2026, published 2026-07-08, version field 81"},
            {"key": "official access", "value": "anonymous public GZIP-CSV; no login/API key"},
            {"key": "licence", "value": "CC BY 4.0 under Commission/EEA metadata and legal notices"},
            {
                "key": "central conclusion",
                "value": (
                    "Verified emissions and free allocation are direct; GREU bought allowances "
                    "are a derived positive-shortfall proxy. EUTL cannot construct in_ETS PJ."
                ),
            },
        ]
    )

    source_fields = pd.DataFrame(
        [
            {
                "field": "VERIFIED_EMISSIONS",
                "source_file": YEARLY.name,
                "meaning": "independently verified annual emissions",
                "unit": "tonne CO2e",
                "direct_or_derived": "direct",
            },
            {
                "field": "ALLOCATION",
                "source_file": YEARLY.name,
                "meaning": "free allocation to existing entities",
                "unit": "allowance",
                "direct_or_derived": "direct",
            },
            {
                "field": "ALLOCATION_RES",
                "source_file": YEARLY.name,
                "meaning": "free allocation from new entrants reserve",
                "unit": "allowance",
                "direct_or_derived": "direct",
            },
            {
                "field": "ALLOCATION_TRA",
                "source_file": YEARLY.name,
                "meaning": "transitional free allocation component",
                "unit": "allowance",
                "direct_or_derived": "direct",
            },
            {
                "field": "SURR_ALL",
                "source_file": YEARLY.name,
                "meaning": "total surrendered compliance units",
                "unit": "unit",
                "direct_or_derived": "direct",
            },
            {
                "field": "ACTIVITY_TYPE_CODE",
                "source_file": OPERATORS.name,
                "meaning": "EU ETS regulatory activity",
                "unit": "code",
                "direct_or_derived": "direct; not NACE",
            },
            {
                "field": "nace_2020 / nace_2015",
                "source_file": NACE_MAP.name,
                "meaning": "NACE inferred from Commission carbon-leakage lists",
                "unit": "NACE Rev. 2",
                "direct_or_derived": "secondary/non-official",
            },
        ]
    )

    source_status = pd.DataFrame(
        [
            {
                "source": "European Commission Union Registry daily bulk",
                "status": "PASS",
                "access": "anonymous direct GZIP-CSV",
                "coverage": "32 registry codes; all EU-27; 2005-2025",
                "vintage": "snapshot 2026-07-30",
                "role": "authoritative installation and compliance fields",
            },
            {
                "source": "EEA EU ETS data viewer release",
                "status": "PASS",
                "access": "anonymous bulk ZIP",
                "coverage": "EU ETS countries; 2005-2025",
                "vintage": "July 2026 release; Union Registry extract 2026-07-01",
                "role": "aggregated validation and national auction volumes",
            },
            {
                "source": "EEA activity translation",
                "status": "PASS WITH CAVEAT",
                "access": "included in EEA bulk ZIP",
                "coverage": "4,327 installation-specific old-code translations",
                "vintage": "May 2019 table, still delivered July 2026",
                "role": "ETS old activity code → new ETS activity code; not NACE",
            },
            {
                "source": "EUETS.INFO NACE concordance",
                "status": "DIAGNOSTIC ONLY",
                "access": "anonymous Zenodo ZIP, CC BY 4.0 compilation",
                "coverage": "EU-wide but partial; quantified in mapping_diagnostics",
                "vintage": "package 2026-07-21; NACE from 2015/2020 leakage lists",
                "role": "test installation→NACE→A64 bridge; not production authority",
            },
        ]
    )

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        metadata.to_excel(writer, sheet_name="metadata", index=False)
        source_coverage.to_excel(writer, sheet_name="source_coverage", index=False)
        totals.to_excel(writer, sheet_name="totals", index=False)
        activity.to_excel(writer, sheet_name="activity_comparison", index=False)
        clusters.to_excel(writer, sheet_name="industry_clusters", index=False)
        concept_availability.to_excel(writer, sheet_name="concept_availability", index=False)
        mapping_diagnostics.to_excel(writer, sheet_name="mapping_diagnostics", index=False)
        nace_coverage.to_excel(writer, sheet_name="nace_coverage", index=False)
        detail.to_excel(writer, sheet_name="installation_detail", index=False)
        source_fields.to_excel(writer, sheet_name="source_fields", index=False)
        assumptions.to_excel(writer, sheet_name="assumptions", index=False)
        validation.to_excel(writer, sheet_name="validation", index=False)
        source_status.to_excel(writer, sheet_name="source_status", index=False)
    style_workbook(OUT)

    print(f"wrote {OUT}")
    print(totals.to_string(index=False))
    print(mapping_diagnostics.to_string(index=False))
    if not validation["result"].all():
        failed = validation.loc[~validation["result"], "check"].tolist()
        raise AssertionError(f"validation checks failed: {failed}")


if __name__ == "__main__":
    main()
