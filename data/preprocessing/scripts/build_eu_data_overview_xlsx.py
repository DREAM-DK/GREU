"""Build docs/EU_data_overview.xlsx — a boss-friendly traffic-light overview.

Generates a three-sheet workbook (README, Summary, Detail) summarizing, for
every Danish input the GREU model consumes, where the EU-wide replacement
lives and how well it matches. All content is transcribed from the verified
findings in docs/eu_data_mapping.md (status and verdicts) and
docs/eu_data_pilots.md (pilot evidence), plus the pilot reconciliation
workbooks under data/preprocessing/data/ — this script performs no new
analysis and computes no new numbers.

Re-run after future pilots change any verdict:

    python data/preprocessing/scripts/build_eu_data_overview_xlsx.py
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[3]
OUTPUT_PATH = REPO_ROOT / "docs" / "EU_data_overview.xlsx"
SOURCE_DOC = "docs/eu_data_mapping.md and docs/eu_data_pilots.md"

# Status vocabulary (traffic light). Every Status cell must use one of these.
STATUS_FILLS = {
    "MATCHES": "C6EFCE",      # green  — reproduces the Danish numbers (or already EU-sourced)
    "CLOSE MATCH": "E2EFDA",  # light green — small, understood, quantified differences
    "COARSER": "FFEB9C",      # yellow — same concept but less detail than the Danish data
    "CONSTRUCTED": "F8CBAD",  # orange — no direct source; built transparently from public controls
    "GAP": "FFC7CE",          # red    — no EU-wide source; needs construction method / decision
    "KEPT": "D9D9D9",         # grey   — not a Danish-statistics dependency; stays as-is
}

STATUS_MEANINGS = {
    "MATCHES": "The EU source reproduces the Danish numbers (well under 1% difference), "
               "or the input is already EU-sourced in the model.",
    "CLOSE MATCH": "Small differences remain, but they are quantified and understood "
                   "(e.g. a known concept difference of a few percent).",
    "COARSER": "The same concept exists EU-wide, but at less detail than the Danish data "
               "(e.g. fewer industries). Aggregation works; extra detail needs splitting keys.",
    "CONSTRUCTED": "No EU source publishes this directly. It is built transparently from public "
                   "control totals, with every assumption and residual disclosed (Sweden 2020 pilot).",
    "GAP": "No EU-wide source exists. Needs a construction method, an external input, "
           "or a decision from colleagues.",
    "KEPT": "Not a Statistics Denmark dependency — kept as-is (classification/concordance file).",
}

SUMMARY_HEADERS = [
    "#",
    "Danish input",
    "What it contains",
    "EU source",
    "Status",
    "Key evidence",
    "What is still missing",
]

# One row per input the model consumes (12 to replace + metadata.xlsx kept
# + the financial-accounts module that is already EU-sourced).
SUMMARY_ROWS = [
    (
        "io_long_format.xlsx",
        "The full input-output table: who buys what from whom, production and "
        "imports for 57 industries, plus wages, surplus and taxes",
        "FIGARO inter-country IO tables (naio_10_fcp_s3 / _u3 / _ii3)",
        "COARSER",
        "DK 2020: totals match to 0.1% or better (output 4,055 bn DKK, wages, "
        "public consumption, investment all exact); imports differ −22% due to "
        "re-export treatment",
        "FIGARO has 64 industries vs GREU's 57 custom ones — a different split, not "
        "simply more detail: GREU separates organic/conventional farming, 5 waste "
        "types and 8 transport industries that FIGARO/NACE lump into one code each, "
        "so those need an external splitting key even though FIGARO has more "
        "categories overall. Also to be decided: re-export handling, and the "
        "real-estate boundary — NACE 'L' bundles housing rental with other "
        "real-estate services, while GREU keeps them in two industries (68203, "
        "71000); the current mapping sends all of L to 68203, overstating housing "
        "output by ~94 bn DKK and understating business services by ~96 bn DKK in "
        "the DK 2020 check",
    ),
    (
        "io_energy_long_format.xlsx",
        "The energy part of the IO table in money terms: what each industry "
        "and household pays for each energy product",
        "No direct source exists. Constructed from Eurostat physical energy "
        "accounts (env_ac_pefasu) calibrated to national supply-use tables "
        "(naio_10_cp15/cp16)",
        "CONSTRUCTED",
        "SE 2020 package built and twice independently reviewed: all accounting "
        "identities close to machine precision, but 0 of the monetary cells are "
        "directly observed anywhere in public EU data",
        "Direct observation is impossible EU-wide; the method discloses its residuals "
        "(119 bn SEK use side, 103 bn SEK supply side, shown to be almost entirely "
        "non-energy money). Colleague sign-off on accepting these residuals",
    ),
    (
        "energy_and_emissions.xlsx",
        "Physical energy use (PJ), emissions and full price/tax decomposition "
        "by industry, energy product and purpose",
        "Physical: env_ac_pefasu (energy accounts). Emissions: env_ac_ainah_r2. "
        "Purpose split: JRC-IDEES-2023 + EU ETS registry. Prices/taxes: "
        "constructed (see price rows in Detail)",
        "COARSER",
        "DK 2020: physical energy −0.611% (2,237.8 vs 2,251.6 PJ); household "
        "totals +0.04%; fossil CO2 −0.007%",
        "The purpose split (heating / process / transport / ETS) and the price/tax "
        "layer are not published EU-wide and must be constructed; industry detail "
        "is coarser in 7 of 28 industry groups",
    ),
    (
        "non_energy_emissions.xlsx",
        "Process (non-energy) emissions by industry, incl. F-gases",
        "env_ac_ainah_r2 (air emissions accounts, total emissions by industry)",
        "COARSER",
        "DK 2020 (combined with energy emissions, the comparable boundary): "
        "fossil CO2 −0.007%, F-gases exact; CH4 +8.6% and N2O +3.2% source gaps "
        "remain (mostly agriculture)",
        "Eurostat publishes only total emissions; the energy vs non-energy split "
        "must be derived (total minus fuel-based estimate). CH4/N2O gap to "
        "investigate",
    ),
    (
        "emissions_bridge_items.xlsx",
        "Adjustment items bridging resident-principle emissions to national "
        "territory (border trade, international transport, LULUCF)",
        "env_ac_aibrid_r2 (air emissions bridging items — the exact same concept, "
        "one dataset covers all three rows)",
        "MATCHES",
        "DK 2020: net residence adjustment matches to within 0.05% per gas; all "
        "27 member states publish every needed cell — the first pilot with zero "
        "coverage gaps",
        "The Danish split into two rows follows a national definition (quantified, "
        "internal reclassification only); LULUCF level differs by inventory "
        "vintage (+17.7%), same concept",
    ),
    (
        "employed.xlsx",
        "Employment and hours worked by industry, split employees / self-employed",
        "nama_10_a64_e (national-accounts employment by 64 industries)",
        "MATCHES",
        "DK 2020: hours −0.000% nationally, exact in 24 of 28 industry groups — and "
        "hours are the only per-industry content the model uses. Persons +3.52% "
        "(known concept difference)",
        "Six countries (DE, FR, BE, BG, LT, EE) publish hours at coarser industry "
        "level; the Danish person concept needs a colleague answer (affects one "
        "national scalar only)",
    ),
    (
        "fixed_assets.xlsx",
        "Capital stock by industry and 7 asset types",
        "nama_10_nfa_st (fixed asset stocks by industry and asset)",
        "COARSER",
        "Not yet piloted. Eurostat side is 21 industry groups (vs 64/57) and asset "
        "detail varies by country",
        "Industry detail; for countries without stock data a capital-stock "
        "estimation (PIM) from investment series may be needed",
    ),
    (
        "io_invest_long_format.xlsx",
        "Investment split: how much of each industry's investment is buildings, "
        "vehicles or machinery, and which industries supply those goods",
        "Two partial sources, no single one: FIGARO's investment column by "
        "product (supply side) and nama_10_a64_p5 / nama_10_nfa_st by asset "
        "and industry (use side)",
        "GAP",
        "Rescoped 2026-08-07: the model uses two separate breakdowns, not the "
        "full matrix previously assumed. New 2026-08-17: nama_10_a64_p5 "
        "publishes investment by asset type at near-full industry detail for "
        "13/27 countries incl. DK and SE — the use margin is nearly direct "
        "data there",
        "Supply side is mainly a classification job (construction → buildings, "
        "etc.). Use side: direct for 13 countries; the other 14 must be split "
        "from 21 industry groups using Denmark as a starting pattern — a "
        "method decision for colleagues",
    ),
    (
        "ets.xlsx",
        "EU ETS by industry: verified emissions, free and bought allowances, "
        "implied carbon cost",
        "EU Union Registry public bulk data (EUTL) + EEA ETS data viewer",
        "CLOSE MATCH",
        "DK 2020: verified emissions +0.007%, free allocation +0.002%, 'bought' "
        "+0.002% — all reproduce almost exactly",
        "'Bought' is a derived shortfall proxy, not observed purchases; the carbon "
        "cost needs an external EUA price; the installation→industry mapping relies "
        "on a secondary source covering 97.6% of DK emissions",
    ),
    (
        "government_finances.xlsx",
        "Government expenditure and revenue by transaction type (wages, "
        "investment, subsidies, ...)",
        "gov_10a_main + gov_10a_taxag (government main aggregates + tax detail)",
        "MATCHES",
        "Pilot done (DK 2020): number-exact — every mappable row reconciles to "
        "the third decimal except interest revenue (+0.62%); the expected MAKRO "
        "difference did not materialize. 14/27 countries publish every needed item",
        "Small enumerable leftovers, each with a named candidate source: "
        "domestic/abroad transfer splits, dividends/rent detail, pension-yield "
        "tax as separate series, EU-paid farm subsidies",
    ),
    (
        "institutional_financial_accounts.xlsx",
        "Net financial positions by sector (households, firms, government, "
        "abroad) and instrument group",
        "nasa_10_f_bs (financial balance sheets) + nasa_10_nf_tr (interest/"
        "dividend flows) — the model's financial-accounts module already "
        "pulls the balance sheets live from Eurostat",
        "MATCHES",
        "Pilot done (DK 2020, 2026-08-18) with a twist: the model never reads "
        "this spreadsheet — the live Eurostat module already supplies the two "
        "numbers the model uses, and the pilot verified that route: equity "
        "positions match exactly for government and rest-of-world, interest "
        "and dividend flows match exactly, and all 27 member states publish "
        "every needed cell (second input ever with zero EU-wide gaps)",
        "The Danish data applies a pension adjustment (household pension "
        "savings counted as owned by households directly rather than via "
        "pension funds) worth 2,703.8 bn DKK of equity and +837.3 bn DKK of "
        "household net wealth; the live module does not do this yet — "
        "colleagues must decide whether to replicate it (the needed Eurostat "
        "detail exists EU-wide) or adopt the unadjusted definition. "
        "Housekeeping: the module is hardcoded to Denmark/2019-2020 and keeps "
        "no download provenance",
    ),
    (
        "EU_GR_data.gdx",
        "Marginal energy and CO2 tax rates (currently imported from the Danish "
        "GreenREFORM model as a stopgap)",
        "No direct EU counterpart. Constructed public-core GDX (average tax per "
        "PJ as explicit average=marginal assumption)",
        "CONSTRUCTED",
        "SE 2020: complete compatible GDX built; energy rate = allocated average "
        "tax/PJ, separate CO2 marginal rate recorded as unavailable (zero)",
        "A legal excise/ETS rate engine is needed before these can be treated as "
        "true policy rates rather than calibration rates",
    ),
    (
        "metadata.xlsx",
        "Classifications and concordances (GREU industries ↔ NACE, energy "
        "products ↔ Eurostat codes, consumption groups ↔ COICOP, ...)",
        "Kept — becomes the master EU concordance file; not a data dependency",
        "KEPT",
        "The existing maps already express the Danish structure in EU vocabulary — "
        "a big head start",
        "Four energy-product map fixes pending owner review (P18 gasoil, heat-pump "
        "ambient energy, a spelling, P10); the real-estate L↔68203 split is "
        "value-inconsistent and needs a note or split key",
    ),
    (
        "Financial-accounts module (data/Modules/financial_accounts)",
        "Live Eurostat pull feeding the model's financial accounts — listed for "
        "completeness",
        "nasa_10_f_bs via the Eurostat API (already implemented)",
        "MATCHES",
        "Already EU-sourced and running in the model — and since 2026-08-18 "
        "verified against the Danish numbers: its instrument definitions are "
        "the Danish-consistent ones (a competing definition in the reference "
        "code is not)",
        "Implement or reject the Danish pension adjustment (decision for "
        "colleagues); generalize the hardcoded country/years; add download "
        "provenance",
    ),
]

DETAIL_HEADERS = [
    "Danish input",
    "Variable / datapoint",
    "What it is",
    "EU source (exact dataset code)",
    "Status",
    "Match evidence",
    "Verified in",
    "Notes / next step",
]

# One row per variable/component with a distinct status.
DETAIL_ROWS = [
    # ------------------------------------------------------------- io_long_format
    (
        "io_long_format.xlsx",
        "Main IO flows (production + imports, 57 industries)",
        "Who buys what from whom, in basic prices",
        "FIGARO naio_10_fcp_ii3 (+ _s3/_u3 supply-use)",
        "COARSER",
        "DK 2020: total output 4,055.4 bn DKK exact; GDP +0.1%; 7 of 28 industry "
        "groups are coarser (34 of 57 GREU industries affected)",
        "figaro_dk2020_reconciliation.xlsx",
        "GREU industries that split one NACE code (organic/conventional farming, "
        "5 waste industries, 2 energy industries) need country splitting keys. "
        "Note the 64-vs-57 counts are a different split, not simply more detail: "
        "FIGARO's extra categories sit where GREU aggregates anyway (harmless), "
        "while in these 7 groupings GREU is the finer one and FIGARO gives only a "
        "combined total",
    ),
    (
        "io_long_format.xlsx",
        "Imports and re-exports",
        "Import rows of the IO table",
        "FIGARO (same tables, inter-country logic)",
        "COARSER",
        "DK 2020: imports −22% because Danish re-exports (239.8 bn DKK) have no "
        "FIGARO counterpart; −2.3% once re-exports are excluded",
        "figaro_dk2020_reconciliation.xlsx",
        "Decide re-export handling; import composition also differs (foreign "
        "margins as service rows vs embedded in goods prices)",
    ),
    (
        "io_long_format.xlsx",
        "Primary inputs (wages, gross surplus, other taxes)",
        "Value-added components per industry",
        "FIGARO valuation layers + nama_10_a64",
        "CLOSE MATCH",
        "DK 2020: wages (D1) 1,210.4 and gross surplus 824.0 bn DKK exact; "
        "real-estate boundary shifts ±45 bn between two industry groups",
        "figaro_dk2020_reconciliation.xlsx",
        "The L↔68203 real-estate boundary needs a split key (hit by 3 pilots). "
        "It is a mapping error, not a data gap: NACE 'L' bundles housing rental "
        "with other real-estate services, but GREU keeps these in two industries "
        "(68203 housing, 71000 business services) and the concordance currently "
        "routes all of L to 68203 — worth ~94 bn DKK too much housing output and "
        "~96 bn DKK too little business services for DK 2020",
    ),
    (
        "io_long_format.xlsx",
        "5-way product-tax split (5 named Danish taxes)",
        "tax_products split into 5 named taxes + VAT row",
        "None — FIGARO publishes one combined row (D21X31)",
        "GAP",
        "DK 2020: matches as a sum only (305.3 vs 302.2 bn DKK)",
        "figaro_dk2020_reconciliation.xlsx",
        "Needs an explicit tax engine or an agreed simplification of the model "
        "interface",
    ),
    (
        "io_long_format.xlsx",
        "Household consumption detail (12 GREU groups)",
        "Consumption split into the model's 12 groups",
        "nama_10_co3_p3 (consumption by COICOP purpose)",
        "COARSER",
        "Not yet piloted. Eurostat publishes 2–3 digit COICOP; the Danish map "
        "uses 4-digit",
        "—",
        "3-digit is expected to approximate the 12 groups; check group by group",
    ),
    # ----------------------------------------------------- io_energy_long_format
    (
        "io_energy_long_format.xlsx",
        "Energy IO table in money terms",
        "Energy spending per product × user, all value components",
        "Constructed: env_ac_pefasu physical backbone calibrated to "
        "naio_10_cp15/cp16 supply-use controls",
        "CONSTRUCTED",
        "SE 2020: 610.6 bn SEK purchaser value; every accounting identity closes "
        "to machine precision; 0 monetary cells directly observed",
        "energy_money_se2020_public_core_reconciliation.xlsx",
        "Disclosed residuals: 118.8 bn SEK (use) / 102.6 bn SEK (supply), shown by "
        "two follow-up pilots to be ≥85–98.5% non-energy money. Colleague "
        "acceptance recommended",
    ),
    # -------------------------------------------------------- energy_and_emissions
    (
        "energy_and_emissions.xlsx",
        "Physical energy use (PJ) by industry and product",
        "The physical energy backbone of the model",
        "env_ac_pefasu (physical energy flow accounts)",
        "CLOSE MATCH",
        "DK 2020: 2,237.8 vs 2,251.6 PJ (−0.611%) on the comparable boundary; "
        "supply and use balance on both sides",
        "eurostat_energy_emissions_dk2020_reconciliation.xlsx",
        "7 of 28 industry groups coarser; 4 product-map fixes pending owner review; "
        "no separate transmission-loss cell",
    ),
    (
        "energy_and_emissions.xlsx",
        "Household energy by purpose (heating / transport / appliances)",
        "The three household end uses",
        "env_ac_pefasu (HH_HEAT / HH_TRA / HH_OTH)",
        "CLOSE MATCH",
        "DK 2020: household total +0.039%; allocation differs (heating −6.9 PJ, "
        "appliances +6.7 PJ)",
        "eurostat_energy_emissions_dk2020_reconciliation.xlsx",
        "JRC-IDEES can serve as optional finer split key (residential subtotal "
        "−0.24%)",
    ),
    (
        "energy_and_emissions.xlsx",
        "Industry purpose split (heating, normal/special process, transport)",
        "How each industry uses its energy",
        "JRC-IDEES-2023 (analytical database, EU-27, 2000–2023)",
        "GAP",
        "DK 2020: combined process envelope +3.05% works, but exact GREU "
        "categories are constructed (heating proxy −81%, special process far "
        "too broad)",
        "jrc_idees_dk2020_purpose_reconciliation.xlsx",
        "Needs an owner-approved IDEES-process→GREU-purpose concordance with PEFA "
        "as the balancing control",
    ),
    (
        "energy_and_emissions.xlsx",
        "in_ETS purpose flag (energy used in ETS-regulated activity)",
        "Which energy use falls under the EU ETS",
        "EUTL Union Registry + installation→industry bridge",
        "GAP",
        "EUTL proves membership and emissions but publishes no fuel, PJ or "
        "energy-use field; best public industry bridge covers 97.6% of DK "
        "emissions but is secondary",
        "eutl_dk2020_reconciliation.xlsx",
        "Needs a maintained installation→industry concordance plus explicit "
        "fuel/emission-factor modelling",
    ),
    (
        "energy_and_emissions.xlsx",
        "Energy-related emissions by industry",
        "CO2 etc. from fuel combustion",
        "env_ac_ainah_r2 (total emissions, energy + process combined)",
        "COARSER",
        "DK 2020 on the combined boundary: fossil CO2 −0.007%, biogenic CO2 "
        "+0.0005%",
        "eurostat_energy_emissions_dk2020_reconciliation.xlsx",
        "Eurostat does not split energy vs process emissions; split must be "
        "derived (see non_energy_emissions row)",
    ),
    (
        "energy_and_emissions.xlsx",
        "Basic values of energy flows",
        "Energy value before margins and taxes",
        "naio_10_cp15 (national supply-use, broad energy products)",
        "CONSTRUCTED",
        "SE 2020: calibrated to SUT controls; identities close exactly. Known "
        "hole: Sweden publishes no breakdown at all for coal and crude oil "
        "(916.8 PJ flagged, valued at zero)",
        "energy_money_se2020_public_core_reconciliation.xlsx",
        "Controls are broad product families, not the model's fine products; "
        "26/27 countries have the tables (Bulgaria absent)",
    ),
    (
        "energy_and_emissions.xlsx",
        "Trade margins (wholesale / retail / motor-vehicle)",
        "Three separate margin layers per energy cell",
        "naio_10_cp15 publishes only one combined trade+transport margin (OTTM)",
        "CONSTRUCTED",
        "SE 2020: combined margin carried in the wholesale column only, retail "
        "and motor margins set to zero (documented compatibility encoding)",
        "energy_money_se2020_public_core_reconciliation.xlsx",
        "No EU source splits the three margins; this is a permanent simplification "
        "unless colleagues change the model interface",
    ),
    (
        "energy_and_emissions.xlsx",
        "Five energy taxes (energy / CO2 / SO2 / NOx / PSO)",
        "Named Danish tax layers per energy cell",
        "naio_10_cp15 net product-tax wedge + env_ac_taxind2 (payer totals) + "
        "TAXUD excise tables (PDF)",
        "CONSTRUCTED",
        "SE 2020: the whole non-VAT wedge carried as 'energy tax', other four "
        "zero; 10.4 bn SEK concept difference vs the environmental-tax accounts "
        "kept visible",
        "energy_money_se2020_public_core_reconciliation.xlsx",
        "Tax accounts identify payers, not products/rates; a legal excise/ETS "
        "engine is needed for named taxes and true marginal rates",
    ),
    (
        "energy_and_emissions.xlsx",
        "VAT per energy cell",
        "VAT on household energy purchases",
        "TAXUD official VAT rate tables + SUT tax-wedge cap",
        "CONSTRUCTED",
        "SE 2020: legal-rate estimate 29.77 vs calibrated 28.39 bn SEK — the "
        "1.375 bn difference is kept as a visible audit column",
        "energy_money_se2020_public_core_reconciliation.xlsx",
        "Statutory rate ≠ observed revenue: taxable base and business VAT "
        "recovery are modelled by an explicit incidence rule",
    ),
    (
        "energy_and_emissions.xlsx",
        "Energy prices (basic → purchaser decomposition)",
        "Observed market prices to anchor the valuation",
        "nrg_pc_202_c–205_c (electricity/gas price components), EC Weekly Oil "
        "Bulletin",
        "GAP",
        "DK 2020: direct price controls cover only 15.0% of PJ and 24.7% of "
        "purchaser value (electricity + road fuels); gas components missing for "
        "3 of 27 countries",
        "energy_money_dk2020_feasibility_gap.xlsx",
        "Prices lack industry/purpose detail; they serve as weights and "
        "validation, not as direct cell values",
    ),
    (
        "energy_and_emissions.xlsx",
        "Purchaser values",
        "Total price paid = basic + margins + taxes + VAT",
        "Derived identity (sum of the components above)",
        "CONSTRUCTED",
        "SE 2020: component identity closes to 7×10⁻¹⁵ bn SEK",
        "energy_money_se2020_public_core_reconciliation.xlsx",
        "Sound by construction once the components exist; quality is inherited "
        "from the component rows above",
    ),
    # ------------------------------------------------------- non_energy_emissions
    (
        "non_energy_emissions.xlsx",
        "Process emissions by industry (incl. F-gases)",
        "Emissions not from fuel combustion (cement, agriculture, F-gases)",
        "env_ac_ainah_r2 (covers total emissions incl. F-gases)",
        "COARSER",
        "DK 2020 combined boundary: F-gases exact (+0.001 kt); CH4 +8.6% and "
        "N2O +3.2% source discrepancies (mostly agriculture, +811 kt CO2e)",
        "eurostat_energy_emissions_dk2020_reconciliation.xlsx",
        "Derive non-energy = total − energy-related estimate; investigate the "
        "CH4/N2O vintage/adjustment gap",
    ),
    # ---------------------------------------------------- emissions_bridge_items
    (
        "emissions_bridge_items.xlsx",
        "Residence-to-territory bridge items",
        "Border trade, international transport and LULUCF adjustments",
        "env_ac_aibrid_r2 (air emissions bridging items, incl. LULUCF block)",
        "MATCHES",
        "DK 2020: net residence adjustment within 0.05% per gas; LULUCF exact "
        "concept match (inventory-vintage level difference +17.7%); complete "
        "EU-27 coverage",
        "emissions_bridge_dk2020_reconciliation.xlsx",
        "Build note: derive border trade from net land transport (the Danish "
        "two-row split is a national definition; only border trade reaches the "
        "model)",
    ),
    # ---------------------------------------------------------------- employed
    (
        "employed.xlsx",
        "Hours worked (employees / self-employed × industry)",
        "The only per-industry content the model actually uses (wage imputation "
        "ratio)",
        "nama_10_a64_e (THS_HW, SAL_DC/SELF_DC)",
        "MATCHES",
        "DK 2020: national total −0.000%; exact in 24 of 28 industry groups",
        "employment_dk2020_reconciliation.xlsx",
        "DE, FR, BE, BG, LT, EE publish hours at coarser level — distribute over "
        "person shares (mild: only a ratio enters the model)",
    ),
    (
        "employed.xlsx",
        "Persons employed",
        "Head counts (model uses only one national total)",
        "nama_10_a64_e (THS_PER)",
        "CLOSE MATCH",
        "DK 2020: uniform +3.52% concept gap (Danish column uses a non-standard "
        "person concept); hours match proves both sides are the same accounts",
        "employment_dk2020_reconciliation.xlsx",
        "Colleague question: which person concept does the Danish column use? "
        "Affects one national scalar only",
    ),
    # -------------------------------------------------------------- fixed_assets
    (
        "fixed_assets.xlsx",
        "Capital stock by industry × 7 asset types",
        "Machinery, buildings, transport equipment etc. per industry",
        "nama_10_nfa_st (fixed asset stocks)",
        "COARSER",
        "Not yet piloted. Eurostat is 21 industry groups vs 64/57; asset detail "
        "varies by country",
        "—",
        "May need capital-stock estimation (PIM) from investment series for "
        "countries with missing stock data",
    ),
    # ------------------------------------------------------ io_invest_long_format
    (
        "io_invest_long_format.xlsx",
        "Supply margin: investment goods by supplying industry × asset type",
        "Which industries produce the buildings, vehicles and machinery that "
        "get invested",
        "FIGARO P51G column (by supplying product); split into three asset "
        "types via a concordance",
        "COARSER",
        "FIGARO P51G total verified against the Danish investment columns: "
        "516.1 bn DKK, ≤0.1% (DK 2020)",
        "figaro_dk2020_reconciliation.xlsx",
        "Mostly a classification job: construction products → buildings, "
        "CPA C29-C30 → transport, machinery/ICT/IP → other. Only ambiguous "
        "products need estimating",
    ),
    (
        "io_invest_long_format.xlsx",
        "Use margin: investment by investing industry × asset type",
        "How much of each industry's investment is buildings, vehicles or "
        "machinery",
        "nama_10_a64_p5 (asset × industry, near-full detail for 13/27 "
        "countries); nama_10_nfa_st (21 groups) as fallback",
        "GAP",
        "Probed 2026-08-17: nama_10_a64_p5 carries investment by asset at "
        "(near-)64-industry detail for 13/27 countries incl. DK and SE — for "
        "those the use margin is nearly direct data. The other 14 publish only "
        "~7–23 industry codes and still need disaggregation",
        "—",
        "For the 14 coarse countries: Denmark as starting pattern; identifiable "
        "per industry group only where enough years exist, so large groups such "
        "as manufacturing stay assumption-driven. Recommended first step is a "
        "Denmark back-test against its known answer",
    ),
    # ----------------------------------------------------------------------- ets
    (
        "ets.xlsx",
        "Verified ETS emissions by industry",
        "Emissions regulated under the EU ETS",
        "Union Registry public bulk CSV (installation level)",
        "MATCHES",
        "DK 2020: 11,039.4 vs 11,038.6 kt CO2e (+0.0067%; difference is a later "
        "aviation revision)",
        "eutl_dk2020_reconciliation.xlsx",
        "Direct field; industry attribution relies on the bridge row below",
    ),
    (
        "ets.xlsx",
        "Free allowance allocation",
        "Allowances handed out for free",
        "Union Registry public bulk CSV",
        "MATCHES",
        "DK 2020: +0.0023%",
        "eutl_dk2020_reconciliation.xlsx",
        "Direct field",
    ),
    (
        "ets.xlsx",
        "Bought allowances",
        "Allowances installations had to buy",
        "Derived: max(verified emissions − free allocation, 0) per installation",
        "CLOSE MATCH",
        "DK 2020: +0.0020% — reproduces the Danish number, but as a shortfall "
        "proxy, not observed purchases (banking/transfers invisible)",
        "eutl_dk2020_reconciliation.xlsx",
        "Same derivation the Danish input evidently used; document it as a proxy",
    ),
    (
        "ets.xlsx",
        "Implied ETS cost / carbon tax",
        "Money value of the allowance shortfall",
        "No price field in the Registry; needs an external EUA price",
        "GAP",
        "GREU's number implies a uniform 231.83 DKK/t on the shortfall; applying "
        "the same rate to current data reproduces it to 5 decimals",
        "eutl_dk2020_reconciliation.xlsx",
        "Choose and document an official EUA price series (e.g. auction results)",
    ),
    (
        "ets.xlsx",
        "Installation → industry mapping",
        "Assigning each ETS installation to a model industry",
        "Registry publishes ETS activity codes, not NACE; best public bridge is "
        "the secondary EUETS.INFO/Zenodo package",
        "COARSER",
        "Bridge covers 94.9% of DK 2020 records and 97.6% of emissions; the "
        "broad 'combustion of fuels' code alone spans 336 records",
        "eutl_dk2020_reconciliation.xlsx",
        "Decide: maintain a reviewed concordance, or redesign ETS inputs at "
        "regulatory-activity level",
    ),
    # -------------------------------------------------------- government_finances
    (
        "government_finances.xlsx",
        "Government expenditure/revenue by ESA transaction",
        "Wages, investment, subsidies, transfers etc. of general government",
        "gov_10a_main + gov_10a_taxag (main aggregates + tax detail)",
        "MATCHES",
        "Pilot done (DK 2020): number-exact to the third decimal for every "
        "mappable row except interest revenue (+0.62%); the MAKRO caveat did "
        "not materialize. 14/27 countries complete; gaps are plausibly-zero "
        "items plus patchy counterpart memo items",
        "government_finances_dk2020_reconciliation.xlsx",
        "Leftovers with named candidates (nasa_10_nf_tr or fixed shares): "
        "domestic/abroad transfer splits, D421/D422/D45 detail, PAL "
        "pension-yield tax, EU-paid CAP subsidies",
    ),
    # -------------------------------------- institutional_financial_accounts
    (
        "institutional_financial_accounts.xlsx",
        "Net financial positions by sector × instrument",
        "Who owns/owes what: households, firms, government, rest of world",
        "nasa_10_f_bs (financial balance sheets)",
        "MATCHES",
        "DK 2020: equity (instrument F5) matches exactly for government "
        "(+0.007 bn DKK) and rest-of-world (−0.020); firms and households "
        "differ only by the documented Danish pension adjustment, which the "
        "data pins down mirror-exactly at 2,703.8 bn DKK (≈ the insurance/"
        "pension subsector's entire portfolio, published EU-wide as sector "
        "S128_S129). Small debt-position differences remain for government "
        "(+23.6 bn) and rest-of-world (+46.8 bn) — 0.3–1.3%, most likely "
        "data-vintage. All 27 member states publish every needed cell",
        "financial_accounts_dk2020_reconciliation.xlsx",
        "Key finding: the model never reads this spreadsheet — it already "
        "loads these positions from the live Eurostat module, now verified. "
        "The pension adjustment is not net-neutral (+837.3 bn DKK household "
        "wealth) and is unimplemented in the module: decision 18. A competing "
        "equity definition in the colleague reference code (F51) does NOT "
        "reproduce the Danish numbers; the module's (F5) does",
    ),
    (
        "institutional_financial_accounts.xlsx",
        "Interest and dividend flows",
        "Property-income flows between sectors (D41, D42)",
        "nasa_10_nf_tr (non-financial transactions by sector)",
        "MATCHES",
        "DK 2020: interest (D41) and dividends (D42) match exactly for "
        "government and rest-of-world — including the gross received/paid "
        "sides — and exactly in sum for firms+households (the pension "
        "adjustment shifts 51.1 bn interest and 22.4 bn dividends between "
        "the two). Revaluations have no source in these datasets",
        "financial_accounts_dk2020_reconciliation.xlsx",
        "The model generates these flows from calibrated rates rather than "
        "loading them, so this is a calibration opportunity, not a gap. "
        "Bonus: the same dataset closed two government-pilot leftovers "
        "exactly (rent D45, the D42 dividend bundle) and offers a close "
        "candidate for EU-paid CAP subsidies (−3.2%); the dividend "
        "D421/D422 sub-split and the PAL pension-yield tax remain "
        "unavailable",
    ),
    # -------------------------------------------------------------- EU_GR_data
    (
        "EU_GR_data.gdx",
        "Marginal energy tax rates (tEAFG_REmarg)",
        "Marginal tax per PJ by product × user",
        "Constructed: allocated average energy tax / PJ",
        "CONSTRUCTED",
        "SE 2020: complete compatible GDX built with an explicit "
        "average=marginal assumption",
        "eu_core/SE/energy_money_manifest.json",
        "A legal excise-rate engine is needed to turn these into true policy "
        "rates",
    ),
    (
        "EU_GR_data.gdx",
        "Marginal CO2 tax rates (tCO2_REmarg)",
        "Marginal CO2 tax per unit by product × user",
        "None — no defensible EU-wide source at this grain",
        "GAP",
        "SE 2020: symbol delivered complete but zero, explicitly documented as "
        "unavailable",
        "eu_core/SE/energy_money_manifest.json",
        "Same excise/ETS engine as above would close this",
    ),
    # ---------------------------------------------------------------- metadata
    (
        "metadata.xlsx",
        "Classifications and concordances",
        "GREU↔NACE industries, energy products↔Eurostat codes, consumption↔COICOP, "
        "sectors/flows↔ESA",
        "Kept as the master EU concordance file",
        "KEPT",
        "The maps already express the model in EU vocabulary — this is what makes "
        "the whole replacement feasible",
        "—",
        "Pending: 4 energy-product map fixes (owner review), the L↔68203 "
        "real-estate split, extensions for EUTL activities and asset bridges, and "
        "a NACE Rev. 2.1 check",
    ),
]

# ---------------------------------------------------------------- Progress
# One-line scoreboard per consumed input: has it been piloted, when, and
# what came out. Transcribed from docs/eu_data_mapping.md (mapping table +
# status) and docs/eu_data_pilots.md.

PROGRESS_STATE_FILLS = {
    "PILOT DONE": "C6EFCE",     # green
    "BUILT": "E2EFDA",          # light green — constructed package accepted
    "PROBED": "FFEB9C",         # yellow — sources probed, no full pilot yet
    "PARTLY": "FFEB9C",         # yellow — checked inside another pilot
    "NOT PILOTED": "FFC7CE",    # red
    "KEPT": "D9D9D9",           # grey
}

PROGRESS_HEADERS = [
    "#",
    "Danish input",
    "Verdict",
    "Pilot state",
    "Date",
    "Headline result",
    "Evidence",
]

PROGRESS_ROWS = [
    ("io_long_format.xlsx", "COARSER", "PILOT DONE", "2026-07-29",
     "Totals match to 0.1%; re-exports and the 7 finer-than-NACE industry "
     "groups remain open",
     "figaro_dk2020_reconciliation.xlsx"),
    ("io_energy_long_format.xlsx", "CONSTRUCTED", "BUILT", "2026-07-30/31",
     "Sweden 2020 package accepted; 0 monetary cells directly observed; "
     "residuals disclosed and shown to be non-energy money",
     "energy_money_se2020_public_core_reconciliation.xlsx"),
    ("energy_and_emissions.xlsx", "COARSER", "PILOT DONE", "2026-07-30",
     "Physical energy −0.611%; emissions −0.007% (CO2); purpose and "
     "price/tax layers must be constructed",
     "eurostat_energy_emissions_dk2020_reconciliation.xlsx"),
    ("non_energy_emissions.xlsx", "COARSER", "PARTLY", "2026-07-30",
     "Combined emissions boundary verified inside the PEFA pilot (F-gases "
     "exact); the energy/non-energy split derivation is untested",
     "eurostat_energy_emissions_dk2020_reconciliation.xlsx"),
    ("emissions_bridge_items.xlsx", "MATCHES", "PILOT DONE", "2026-08-17",
     "Net residence adjustment ≤0.05% per gas; first input with zero EU-27 "
     "coverage gaps",
     "emissions_bridge_dk2020_reconciliation.xlsx"),
    ("employed.xlsx", "MATCHES", "PILOT DONE", "2026-07-31",
     "Hours (the only per-industry content used) essentially exact; persons "
     "+3.52% concept question",
     "employment_dk2020_reconciliation.xlsx"),
    ("fixed_assets.xlsx", "COARSER", "NOT PILOTED", "—",
     "NEXT IN LINE: nama_10_nfa_st at 21 industry groups; doubles as "
     "groundwork for the investment split",
     "—"),
    ("io_invest_long_format.xlsx", "GAP", "PROBED", "2026-08-17",
     "Rescoped to two margins (2026-08-07); use margin is near-direct data "
     "for 13/27 countries via nama_10_a64_p5",
     "probe_nama_10_a64_p5_asset_detail.py"),
    ("ets.xlsx", "CLOSE MATCH", "PILOT DONE", "2026-07-30",
     "Emissions/allocations reproduce to +0.007%; industry bridge and EUA "
     "price remain decisions",
     "eutl_dk2020_reconciliation.xlsx"),
    ("government_finances.xlsx", "MATCHES", "PILOT DONE", "2026-08-17",
     "Number-exact except interest revenue +0.62%; splits/PAL structural "
     "leftovers have named candidates",
     "government_finances_dk2020_reconciliation.xlsx"),
    ("institutional_financial_accounts.xlsx", "MATCHES", "PILOT DONE",
     "2026-08-18",
     "Model never reads the Excel — the live Eurostat module is now "
     "verified; pension adjustment quantified (decision 18); zero EU-27 "
     "coverage gaps",
     "financial_accounts_dk2020_reconciliation.xlsx"),
    ("Household consumption detail (12 groups)", "COARSER", "NOT PILOTED",
     "—",
     "nama_10_co3_p3 at 2–3 digit COICOP vs the Danish 4-digit map; check "
     "group by group",
     "—"),
    ("EU_GR_data.gdx", "CONSTRUCTED", "BUILT", "2026-07-30/31",
     "Sweden compatible GDX with explicit average=marginal assumption; a "
     "legal excise/ETS engine is still needed for true policy rates",
     "eu_core/SE/energy_money_manifest.json"),
    ("metadata.xlsx", "KEPT", "KEPT", "—",
     "Master concordance file, kept; 4 energy-product fixes pending owner "
     "review",
     "—"),
]

# --------------------------------------------------------------- Decisions
# One row per open decision for colleagues. Transcribed from the tagged
# list in docs/eu_data_mapping.md ('Decisions needed from colleagues');
# days-open is computed at generation time.

DECISIONS_HEADERS = [
    "ID",
    "Decision needed",
    "Owner",
    "Raised",
    "Days open",
    "What it blocks",
]

DECISIONS_ROWS = [
    ("6",
     "Which GREU industry splits finer than NACE A64 survive in the EU "
     "version (organic/conventional farming, five waste industries, "
     "electricity subdivisions)? Country keys or aggregated level?",
     "Model owners", "2026-07-28",
     "The final industry dimension of every converted input — packages "
     "built before this may need re-cutting"),
    ("7",
     "Re-export handling, and the NACE-L ↔ 68203 real-estate split (hit by "
     "three pilots)",
     "metadata.xlsx concordance owner", "2026-07-29",
     "Exact cluster-level reconciliation in the FIGARO, PEFA and "
     "employment pilots"),
    ("8",
     "Review the four energy-product concordance fixes the PEFA pilot "
     "exposed (P18 diesel, ambient heat, spelling, P10)",
     "metadata.xlsx concordance owner", "2026-07-30",
     "Removing the pilots' ad-hoc adjustment layer from every PEFA-based "
     "build"),
    ("10",
     "Agree the JRC-IDEES process-code concordance and the rules for "
     "heating / process_normal / process_special",
     "Energy/purpose work-stream colleagues", "2026-07-30",
     "Structural gap 2 (the purpose dimension)"),
    ("11",
     "ETS: maintain a reviewed installation→industry concordance, or "
     "redesign at regulatory-activity level? Plus the fuel/emission-factor "
     "method for in_ETS",
     "Model owners + emissions colleagues", "2026-07-30",
     "The ets.xlsx industry bridge and the in_ETS part of gap 2"),
    ("13",
     "Accept the evidence-backed non-energy residuals (CPA_C16 ≥98.5%, "
     "CPA_E37-E39 ≥85%) as permanent disclosed features of the public-core "
     "method (recommended)",
     "Energy-money work-stream colleagues", "2026-07-31",
     "Formally closing the Sweden monetary-residual work stream"),
    ("16",
     "Which person concept does the Danish employed column use? (uniform "
     "+3.52% vs national accounts while hours match exactly)",
     "employed.xlsx author (MAKRO/DST side)", "2026-07-31",
     "Only the nEmployed(t) scalar — low stakes"),
    ("Invest",
     "Investment split method: Denmark-as-prior, time-invariant shares, "
     "Julia as a new toolchain dependency",
     "Model owners / management", "2026-08-07",
     "Starting the gap-3 estimator and the Denmark back-test"),
    ("18",
     "Pension-asset reallocation: replicate the Danish adjustment from "
     "S128_S129 subsector balance sheets (published EU-wide), or accept "
     "the unadjusted Eurostat definition? It is worth 2,703.8 bn DKK of "
     "equity and +837.3 bn DKK of household net wealth, and the live "
     "module does not implement it",
     "Model owners", "2026-08-18",
     "Closing the institutional_financial_accounts row; household-wealth "
     "levels in calibration"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
THIN_BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def _autofit_row_heights(ws, widths, first_data_row=2):
    """Rough row-height estimate so wrapped text stays visible."""
    for row in ws.iter_rows(min_row=first_data_row):
        lines = 1
        for cell, width in zip(row, widths):
            if cell.value:
                text = str(cell.value)
                lines = max(lines, -(-len(text) // max(width - 2, 10)))
        row[0].parent.row_dimensions[row[0].row].height = min(15 * lines, 150)


def _write_table(ws, headers, rows, widths, status_col_idx):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in rows:
        ws.append(row)
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_TOP
            cell.border = THIN_BORDER
        status_cell = row[status_col_idx - 1]
        status = status_cell.value
        if status not in STATUS_FILLS:
            raise ValueError(f"Illegal status {status!r} in sheet {ws.title}")
        status_cell.fill = PatternFill("solid", fgColor=STATUS_FILLS[status])
        status_cell.font = Font(bold=True)
        status_cell.alignment = Alignment(vertical="top", horizontal="center",
                                          wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    _autofit_row_heights(ws, widths)


def _write_plain_table(ws, headers, rows, widths, fill_col_idx=None,
                       fill_map=None):
    """Table writer for sheets whose highlight column uses its own
    vocabulary (Progress / Decisions) instead of the Status legend."""
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for row in rows:
        ws.append(row)
    for col_idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_TOP
            cell.border = THIN_BORDER
        if fill_col_idx is not None:
            cell = row[fill_col_idx - 1]
            key = cell.value
            if key not in fill_map:
                raise ValueError(f"Illegal state {key!r} in sheet {ws.title}")
            cell.fill = PatternFill("solid", fgColor=fill_map[key])
            cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical="top", horizontal="center",
                                       wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    _autofit_row_heights(ws, widths)


def build_workbook() -> Workbook:
    wb = Workbook()

    # ---------------------------------------------------------------- README
    ws = wb.active
    ws.title = "README"
    today = dt.date.today().isoformat()
    intro = [
        ("EU data overview — what we need, where it comes from, and how well it matches",),
        (),
        (f"Created: {today}. Generated by "
         "data/preprocessing/scripts/build_eu_data_overview_xlsx.py — re-run the "
         "script after new pilots; do not edit this file by hand.",),
        (),
        ("Purpose: the GREU model currently runs on Danish (Statistics Denmark) "
         "input data. To make the model usable by any EU member state, every "
         "Danish input must be replaced with data available EU-wide. This "
         "workbook lists every such input, the EU source found for it, and how "
         "well that source reproduces the Danish numbers in the completed pilot "
         "reconciliations (Denmark 2020 and Sweden 2020).",),
        (),
        ("Sheets: 'Progress' is the scoreboard — one line per input: piloted "
         "or not, when, and the headline result. 'Decisions' lists every "
         "question currently waiting on colleagues, with owner, age and what "
         "it blocks — the sheet to bring to a meeting. 'Summary' has one row "
         "per input file — the one-page view. 'Detail' breaks each input "
         "into its individual variables/datapoints, because pieces of one "
         "file can have very different statuses.",),
        (),
        ("Where it stands (see Progress): 9 of the 13 inputs needing "
         "replacement have pilot or build evidence, 1 is probed and rescoped "
         "(the investment split), and 3 are untouched (fixed assets — next "
         "in line, the non-energy emissions split, and the household "
         "consumption detail). No input has been found infeasible. The "
         "binding constraint is now the Decisions sheet: several open items "
         "block more than one pilot.",),
        (),
        (f"All content is transcribed from {SOURCE_DOC} (the technical audit "
         "log) and the pilot reconciliation workbooks named in the 'Verified in' "
         "column, which live in data/preprocessing/data/. Every number is "
         "reproducible from the scripts in data/preprocessing/scripts/.",),
        (),
        ("Status legend:",),
    ]
    # Prose cells are left unwrapped so the text visually overflows across the
    # empty neighbouring columns; only the legend meanings wrap.
    for line in intro:
        ws.append(line)
    ws["A1"].font = Font(bold=True, size=14)
    for status, meaning in STATUS_MEANINGS.items():
        ws.append((status, meaning))
        row = ws.max_row
        cell = ws.cell(row=row, column=1)
        cell.fill = PatternFill("solid", fgColor=STATUS_FILLS[status])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = WRAP_TOP
        ws.row_dimensions[row].height = 30
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 100

    # -------------------------------------------------------------- Progress
    ws_prog = wb.create_sheet("Progress")
    prog_rows = [(i + 1, *row) for i, row in enumerate(PROGRESS_ROWS)]
    _write_plain_table(
        ws_prog,
        PROGRESS_HEADERS,
        prog_rows,
        widths=[4, 34, 14, 14, 12, 62, 44],
        fill_col_idx=4,
        fill_map=PROGRESS_STATE_FILLS,
    )

    # ------------------------------------------------------------- Decisions
    ws_dec = wb.create_sheet("Decisions")
    today_date = dt.date.today()
    dec_rows = []
    for dec_id, question, owner, raised, blocks in DECISIONS_ROWS:
        days_open = (today_date - dt.date.fromisoformat(raised)).days
        dec_rows.append((dec_id, question, owner, raised, days_open, blocks))
    _write_plain_table(
        ws_dec,
        DECISIONS_HEADERS,
        dec_rows,
        widths=[8, 62, 26, 12, 10, 52],
    )

    # --------------------------------------------------------------- Summary
    ws_sum = wb.create_sheet("Summary")
    summary_rows = [(i + 1, *row) for i, row in enumerate(SUMMARY_ROWS)]
    _write_table(
        ws_sum,
        SUMMARY_HEADERS,
        summary_rows,
        widths=[4, 30, 42, 38, 14, 48, 48],
        status_col_idx=5,
    )

    # ---------------------------------------------------------------- Detail
    ws_det = wb.create_sheet("Detail")
    _write_table(
        ws_det,
        DETAIL_HEADERS,
        DETAIL_ROWS,
        widths=[26, 34, 34, 36, 14, 50, 40, 46],
        status_col_idx=5,
    )

    return wb


def main() -> None:
    wb = build_workbook()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")
    print(f"Summary rows: {len(SUMMARY_ROWS)}; Detail rows: {len(DETAIL_ROWS)}; "
          f"Progress rows: {len(PROGRESS_ROWS)}; "
          f"Decisions rows: {len(DECISIONS_ROWS)}")


if __name__ == "__main__":
    main()
