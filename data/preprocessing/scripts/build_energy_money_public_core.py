"""Build and audit a calibrated monetary-energy public-core package.

The default is the approved Sweden 2020 pilot. No Danish values or allocation
shares are read. ``metadata.xlsx`` is used only as the GREU schema/classification
bridge; connected components prevent an A64 category from being split.

Run:
    python data/preprocessing/scripts/build_energy_money_public_core.py
"""

from __future__ import annotations

import argparse
from collections import defaultdict
from datetime import datetime, timezone
import hashlib
import json
from pathlib import Path
import re
import shutil
import tempfile

import gamspy as gp
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DATA = Path(__file__).resolve().parents[1] / "data"
REPO = Path(__file__).resolve().parents[3]
RETRIEVAL_DATE = "2026-07-30"
POLICY_VERSION = "public_core_v1.0"
MONEY_FIELDS = [
    "basic", "ws_marg", "ret_marg", "mvs_marg", "ener_tax", "co2_tax",
    "so2_tax", "nox_tax", "pso_tax", "vat", "purch",
]
EMISSION_FIELDS = ["ch4", "co2_bio", "co2_xbio", "n2o", "co2_eq"]
ENERGY_COLUMNS = [
    "year", "bal", "flow", "indu", "purp", "product",
    *EMISSION_FIELDS, "pj", *MONEY_FIELDS,
]
IO_COLUMNS = ["year", "row_l1", "row_l2", "col_l1", "col_l2", "value"]
CONCEPT_PRODUCTS = [
    *[f"P{i:02d}" for i in range(8, 28)],
    "N03", "N04", "N05", "N07", "R28", "R29",
]
CPA_BY_PEFA = {
    "P08": "CPA_B05", "P09": "CPA_B05", "P11": "CPA_B05",
    "P12": "CPA_B06", "P13": "CPA_B06",
    "P10": "CPA_D", "P26": "CPA_D", "P27": "CPA_D",
    "P23": "CPA_C16", "R28": "CPA_E37-E39", "R29": "CPA_E37-E39",
    **{f"P{i:02d}": "CPA_C19" for i in range(14, 26) if i != 23},
}
CPA_PRODUCER = {
    "CPA_B05": "0600a",
    "CPA_B06": "0600a",
    "CPA_C16": "16000",
    "CPA_C19": "19000",
    "CPA_D": "35002",
    "CPA_E37-E39": "37000",
}
VAT_RATE = 0.25


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def jsonstat_to_long(path: Path) -> tuple[pd.DataFrame, dict[str, dict[str, str]]]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    dims = payload["id"]
    categories: list[list[str]] = []
    labels: dict[str, dict[str, str]] = {}
    for dim in dims:
        category = payload["dimension"][dim]["category"]
        categories.append(
            [
                key
                for key, _ in sorted(
                    category["index"].items(), key=lambda item: item[1]
                )
            ]
        )
        labels[dim] = category.get("label", {})
    rows: list[dict] = []
    for flat_index, value in payload["value"].items():
        positions = np.unravel_index(int(flat_index), payload["size"])
        row = {
            dim: categories[index][positions[index]]
            for index, dim in enumerate(dims)
        }
        row["value"] = float(value)
        row["obs_status"] = payload.get("status", {}).get(str(flat_index), "")
        rows.append(row)
    return pd.DataFrame(rows), labels


def norm_greu(value) -> str:
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text.zfill(5) if text.isdigit() else text


def norm_a64(value) -> str:
    text = str(value).strip()
    text = {
        "C31_32": "C31_C32",
        "E37-39": "E37-E39",
        "J59_60": "J59_J60",
        "J62_63": "J62_J63",
        "M69_70": "M69_M70",
        "M74_75": "M74_M75",
        "N80-82": "N80-N82",
        "Q87_88": "Q87_Q88",
        "R90-92": "R90-R92",
    }.get(text, text)
    return text


def connected_components(
    mapping: pd.DataFrame, left: str, right: str
) -> tuple[pd.DataFrame, dict[str, str], dict[str, str]]:
    parent: dict[tuple[str, str], tuple[str, str]] = {}

    def find(node: tuple[str, str]) -> tuple[str, str]:
        parent.setdefault(node, node)
        if parent[node] != node:
            parent[node] = find(parent[node])
        return parent[node]

    def union(a: tuple[str, str], b: tuple[str, str]) -> None:
        parent[find(a)] = find(b)

    for row in mapping[[left, right]].drop_duplicates().itertuples(index=False):
        union(("left", str(row[0])), ("right", str(row[1])))
    groups: dict[tuple[str, str], list[tuple[str, str]]] = defaultdict(list)
    for node in parent:
        groups[find(node)].append(node)
    rows = []
    for members in groups.values():
        left_values = sorted(value for side, value in members if side == "left")
        right_values = sorted(value for side, value in members if side == "right")
        representative = left_values[0]
        if "49509" in left_values:
            representative = "49509"
        rows.append(
            {
                "cluster": "+".join(right_values),
                "representative": representative,
                "left_members": "|".join(left_values),
                "right_members": "|".join(right_values),
                "n_left": len(left_values),
                "n_right": len(right_values),
            }
        )
    result = pd.DataFrame(rows).sort_values("cluster").reset_index(drop=True)
    left_map = {
        member: row.representative
        for row in result.itertuples()
        for member in row.left_members.split("|")
    }
    right_map = {
        member: row.representative
        for row in result.itertuples()
        for member in row.right_members.split("|")
    }
    return result, left_map, right_map


def source_value(
    frame: pd.DataFrame,
    **filters: str,
) -> float:
    selected = frame
    for column, value in filters.items():
        selected = selected[selected[column] == value]
    return float(selected["value"].sum())


def price_controls(
    raw: Path,
    country: str,
    year: int,
    exchange_rate: float,
) -> tuple[pd.DataFrame, dict[tuple[str, str], float]]:
    rows: list[dict] = []
    controls: dict[tuple[str, str], float] = {}
    datasets = [
        ("nrg_pc_202_c", "household", "P13"),
        ("nrg_pc_203_c", "non_household", "P13"),
        ("nrg_pc_204_c", "household", "P26"),
        ("nrg_pc_205_c", "non_household", "P26"),
    ]
    for dataset, user, product in datasets:
        frame, labels = jsonstat_to_long(
            raw / f"{dataset}_{country}_{year}.json"
        )
        total_band = next(
            (
                code
                for code in frame["nrg_cons"].unique()
                if str(code).startswith("TOT")
            ),
            None,
        )
        selected = frame[
            (frame["currency"] == "NAC")
            & (frame["nrg_cons"] == total_band)
            & frame["nrg_prc"].isin(
                ["NRG_SUP", "NETC", "TAX_FEE_LEV_CHRG", "VAT"]
            )
        ]
        component_sum = float(selected["value"].sum())
        if component_sum > 0:
            controls[(product, user)] = component_sum
        for record in selected.itertuples():
            rows.append(
                {
                    "source": dataset,
                    "product_family": product,
                    "user_group": user,
                    "component": record.nrg_prc,
                    "component_label": labels["nrg_prc"].get(record.nrg_prc, ""),
                    "band": total_band,
                    "value_source_units": record.value,
                    "unit": "national-currency price component",
                    "role": "initial allocation weight; SUT remains hard control",
                    "status": record.obs_status,
                }
            )

    oil_path = raw / "Weekly_Oil_Bulletin_Prices_History_2026-07-30.xlsx"
    oil_sheet = "Prices with taxes"
    oil = pd.read_excel(oil_path, sheet_name=oil_sheet)
    dates = pd.to_datetime(oil.iloc[:, 0], errors="coerce", format="mixed")
    oil = oil.loc[dates.dt.year == year]
    oil_products = {
        "euro95": ("P14", 32.0),
        "diesel": ("P17", 35.8),
        "heating_oil": ("P18", 36.0),
        "fuel_oil_1": ("P19", 40.0),
    }
    for token, (product, gj_per_unit) in oil_products.items():
        candidates = [
            column
            for column in oil.columns
            if str(column).startswith(f"{country}_price_with_tax_{token}")
        ]
        if not candidates:
            continue
        values = pd.to_numeric(oil[candidates[0]], errors="coerce").dropna()
        if values.empty:
            continue
        mean_eur_per_unit = float(values.mean())
        sek_per_gj = mean_eur_per_unit * exchange_rate / gj_per_unit
        controls[(product, "all")] = sek_per_gj
        rows.append(
            {
                "source": "Commission Weekly Oil Bulletin",
                "product_family": product,
                "user_group": "all",
                "component": "purchaser_price_with_tax",
                "component_label": token,
                "band": "weekly observations in 2020",
                "value_source_units": mean_eur_per_unit,
                "unit": "EUR/1000 litres or EUR/tonne",
                "role": (
                    "initial allocation weight; converted with documented "
                    f"{gj_per_unit:g} GJ/unit assumption; SUT hard control"
                ),
                "status": f"{len(values)} weekly observations",
            }
        )
    return pd.DataFrame(rows), controls


def build(
    country: str,
    year: int,
    *,
    output_root: Path,
    audit_path: Path,
    force: bool,
    temporary: bool = False,
) -> dict:
    raw = DATA / "eu_core_raw" / country / str(year)
    package = output_root / country
    policy_path = output_root / f"{POLICY_VERSION}_policy.json"
    required_raw = [
        f"env_ac_pefasu_{country}_{year}.json",
        f"env_ac_ainah_r2_{country}_{year}.json",
        f"env_ac_taxind2_{country}_{year}.json",
        f"naio_10_cp15_{country}_{year}.json",
        f"naio_10_cp16_{country}_{year}.json",
        f"ert_bil_eur_a_SEK_{year}.csv",
        "Weekly_Oil_Bulletin_Prices_History_2026-07-30.xlsx",
        "manifest.json",
    ]
    missing = [name for name in required_raw if not (raw / name).is_file()]
    if missing:
        raise FileNotFoundError(f"missing raw inputs: {missing}")

    outputs = [
        package / "energy_and_emissions.xlsx",
        package / "io_energy_long_format.xlsx",
        package / "EU_GR_data.gdx",
        package / "energy_money_manifest.json",
        package / "README.md",
        policy_path,
        audit_path,
    ]
    if not force:
        existing = [path for path in outputs if path.exists()]
        if existing:
            raise FileExistsError(
                "generated outputs already exist; use --force: "
                + ", ".join(str(path) for path in existing)
            )
    package.mkdir(parents=True, exist_ok=True)
    output_root.mkdir(parents=True, exist_ok=True)

    policy = {
        "policy_version": POLICY_VERSION,
        "country": country,
        "reference_year": year,
        "approved_method": "transparent calibrated public-core engine",
        "industry_rule": (
            "connected components of public NACE A64 to GREU schema map; one "
            "representative GREU code labels each whole cluster; never split A64"
        ),
        "purpose_rule": (
            "industry=unspecified; PEFA HH_HEAT/HH_TRA/HH_OTH retained as "
            "heating/transport/appliances"
        ),
        "physical_rule": (
            "PEFA TJ is the account control; converted to PJ; reporting-detail "
            "and supply/use residuals remain explicit"
        ),
        "valuation_rule": (
            "national SUT purchaser/basic/margin/net-product-tax controls; "
            "public price families form initial weights; block calibration "
            "closes available SUT controls"
        ),
        "reporting_detail_redirect_rule": (
            "if PEFA has no USE row at all for any NACE code in a GREU "
            "industry cluster, that cluster's naio_10_cp16 CPA control is "
            "pooled with indu=res (the same bucket already holding its "
            "PEFA reporting-detail physical residual) instead of being "
            "left as unmatched residual against a zero-weight industry; "
            "industries PEFA does detail (e.g. financial/business "
            "services) are never redirected"
        ),
        "margin_rule": (
            "combined SUT OTTM encoded wholly in ws_marg/EAV; ret_marg and "
            "mvs_marg are zero compatibility fields"
        ),
        "duty_rule": (
            "aggregate non-VAT SUT product-tax wedge encoded in ener_tax; "
            "co2_tax/so2_tax/nox_tax/pso_tax zero unless directly supported"
        ),
        "vat_rule": (
            "25% statutory Sweden rate; households non-recoverable, businesses "
            "recoverable, exports zero; estimate capped by the SUT product-tax "
            "wedge and the difference retained as a calibration residual"
        ),
        "purchaser_rule": "derived exactly as component sum",
        "marginal_rule": (
            "allocated average ener_tax per PJ is assumed marginal in "
            "tEAFG_REmarg; tCO2_REmarg is zero because no defensible separate "
            "CO2 rate is observed at product-user grain"
        ),
        "forbidden_inputs": [
            "Danish monetary allocation shares",
            "Danish energy-purpose shares",
            "Danish marginal-tax GDX values",
        ],
        "retrieval_date": RETRIEVAL_DATE,
    }
    policy_path.write_text(
        json.dumps(policy, indent=2, sort_keys=True) + "\n", encoding="utf-8"
    )

    pefa, pefa_labels = jsonstat_to_long(
        raw / f"env_ac_pefasu_{country}_{year}.json"
    )
    aea, _ = jsonstat_to_long(
        raw / f"env_ac_ainah_r2_{country}_{year}.json"
    )
    tax, tax_labels = jsonstat_to_long(
        raw / f"env_ac_taxind2_{country}_{year}.json"
    )
    cp15, _ = jsonstat_to_long(raw / f"naio_10_cp15_{country}_{year}.json")
    cp16, _ = jsonstat_to_long(raw / f"naio_10_cp16_{country}_{year}.json")
    exchange = pd.read_csv(raw / f"ert_bil_eur_a_SEK_{year}.csv")
    exchange_rate = float(
        exchange.loc[exchange["TIME_PERIOD"] == year, "OBS_VALUE"].iloc[0]
    )
    prices, price_map = price_controls(raw, country, year, exchange_rate)

    metadata = DATA / "metadata.xlsx"
    industry_map = pd.read_excel(
        metadata, sheet_name="industries_naceA64_map"
    )
    industry_map["indu_greu"] = industry_map["indu_greu"].map(norm_greu)
    industry_map["indu_naceA64"] = industry_map["indu_naceA64"].map(norm_a64)
    industry_clusters, _, nace_to_indu = connected_components(
        industry_map, "indu_greu", "indu_naceA64"
    )
    a64_codes = set(nace_to_indu)

    product_map = pd.read_excel(
        metadata, sheet_name="energy_products_pefa_map"
    )
    product_map["product_greu"] = product_map["product_greu"].astype(str)
    product_map["product_greu"] = product_map["product_greu"].replace(
        {"sem_refin_oil": "semi_refin_oil"}
    )
    product_map = pd.concat(
        [
            product_map[["product_greu", "product_pefa"]],
            pd.DataFrame(
                [
                    {"product_greu": "diesel_transp", "product_pefa": "P18"},
                ]
            ),
        ],
        ignore_index=True,
    ).drop_duplicates()
    product_clusters, product_to_rep, pefa_to_product = connected_components(
        product_map, "product_greu", "product_pefa"
    )
    product_text = (
        pd.read_excel(metadata, sheet_name="energy_products")
        .assign(product_greu=lambda x: x["product_greu"].astype(str))
        .set_index("product_greu")["product_greu_txt"]
        .to_dict()
    )
    product_text["semi_refin_oil"] = "Semi-refined oil"
    for code in CONCEPT_PRODUCTS:
        if code not in pefa_to_product:
            name = f"pefa_{code}_unmapped"
            pefa_to_product[code] = name
            product_text[name] = name

    pefa["nace_norm"] = pefa["nace_r2"].map(norm_a64)
    pefa = pefa[
        (pefa["unit"] == "TJ") & pefa["prod_nrg"].isin(CONCEPT_PRODUCTS)
    ].copy()

    # Some countries' PEFA USE side is reported only at NACE *section* level
    # (e.g. the whole of manufacturing "C") for industries the concordance
    # expects at *division* level (C10-C33, ...). The physical account
    # already captures that money-relevant energy correctly as an explicit
    # `indu=res` reporting-detail residual (below), but naio_10_cp16 still
    # publishes purchaser-value controls at the finer division level GREU
    # expects. Left unmodified, every one of those finer industries' SUT
    # money has no matching physical row and becomes unmatched residual,
    # even though its physical counterpart is not actually missing — it is
    # sitting, correctly, in the `res` bucket. A GREU cluster is flagged
    # here only if *none* of its NACE members ever report any USE row for
    # our energy products, so genuine section-level detail (e.g. financial
    # services K64/K65/K66) is left untouched.
    use_reported_codes = set(
        pefa.loc[pefa["stk_flow"] == "USE", "nace_norm"].unique()
    )
    industry_clusters["use_detail_available"] = industry_clusters[
        "right_members"
    ].map(
        lambda members: any(
            code in use_reported_codes for code in members.split("|")
        )
    )

    def selected_pefa(stk_flow: str, nace: str) -> pd.DataFrame:
        return pefa[
            (pefa["stk_flow"] == stk_flow) & (pefa["nace_norm"] == nace)
        ]

    physical_internal: list[dict] = []

    def add_physical(
        bal: str,
        flow: str,
        indu: str,
        purp: str,
        code: str,
        tj: float,
        source_status: str,
        source_nace: str,
    ) -> None:
        if abs(tj) <= 1e-10:
            return
        physical_internal.append(
            {
                "year": year,
                "bal": bal,
                "flow": flow,
                "indu": indu,
                "purp": purp,
                "product": pefa_to_product[code],
                "pefa_product": code,
                "cpa": CPA_BY_PEFA.get(code),
                "pj": tj / 1000.0,
                "physical_status": source_status,
                "source_nace": source_nace,
            }
        )

    for nace in sorted(a64_codes):
        for record in selected_pefa("SUP", nace).itertuples():
            add_physical(
                "sup", "production", nace_to_indu[nace], "unspecified",
                record.prod_nrg, record.value, "direct_PEFA_A64",
                record.nace_r2,
            )
        for record in selected_pefa("USE", nace).itertuples():
            add_physical(
                "use", "cons_inter", nace_to_indu[nace], "unspecified",
                record.prod_nrg, record.value, "direct_PEFA_A64",
                record.nace_r2,
            )

    for record in selected_pefa("SUP", "ROW_ACT").itertuples():
        add_physical(
            "sup", "import", "", "unspecified", record.prod_nrg,
            record.value, "direct_PEFA_rest_of_world", record.nace_r2,
        )
    for record in selected_pefa("SUP", "CH_INV_PA").itertuples():
        add_physical(
            "sup", "invent_change", "", "unspecified", record.prod_nrg,
            record.value, "direct_PEFA_inventory", record.nace_r2,
        )
    for record in selected_pefa("SUP", "ENV").itertuples():
        add_physical(
            "sup", "other_supply", "env", "unspecified", record.prod_nrg,
            record.value, "direct_PEFA_environment", record.nace_r2,
        )

    total_industry = selected_pefa("USE", "TOTAL").groupby("prod_nrg")[
        "value"
    ].sum()
    detailed_industry = (
        pefa[
            (pefa["stk_flow"] == "USE") & pefa["nace_norm"].isin(a64_codes)
        ]
        .groupby("prod_nrg")["value"]
        .sum()
    )
    for code in CONCEPT_PRODUCTS:
        residual = total_industry.get(code, 0.0) - detailed_industry.get(code, 0.0)
        add_physical(
            "use", "cons_inter", "res", "unspecified", code, residual,
            "explicit_PEFA_reporting_detail_residual", "TOTAL-minus-A64",
        )

    household_purposes = {
        "HH_HEAT": ("cHouEne", "heating"),
        "HH_TRA": ("cCarEne", "transport"),
        "HH_OTH": ("cHouEne", "appliances"),
    }
    for nace, (indu, purpose) in household_purposes.items():
        for record in selected_pefa("USE", nace).itertuples():
            add_physical(
                "use", "cons_hh", indu, purpose, record.prod_nrg,
                record.value, "direct_PEFA_household_purpose", record.nace_r2,
            )
    for nace, flow, status in [
        ("ROW_ACT", "export", "direct_PEFA_rest_of_world"),
        ("CH_INV_PA", "invent_change", "direct_PEFA_inventory"),
        ("SD_SU", "other_use", "direct_PEFA_statistical_difference"),
    ]:
        for record in selected_pefa("USE", nace).itertuples():
            add_physical(
                "use", flow, "res" if flow == "other_use" else "",
                "unspecified", record.prod_nrg, record.value, status,
                record.nace_r2,
            )

    physical = pd.DataFrame(physical_internal)
    group_key = [
        "year", "bal", "flow", "indu", "purp", "product", "pefa_product",
        "cpa", "physical_status", "source_nace",
    ]
    physical = physical.groupby(
        group_key, dropna=False, as_index=False
    )["pj"].sum()

    balances = (
        physical.groupby(["pefa_product", "bal"])["pj"]
        .sum()
        .unstack(fill_value=0.0)
        .reset_index()
    )
    balances["residual_sup_minus_use_before"] = (
        balances.get("sup", 0.0) - balances.get("use", 0.0)
    )
    for row in balances.itertuples():
        residual = row.residual_sup_minus_use_before
        if abs(residual) <= 1e-9:
            continue
        add_physical(
            "use" if residual > 0 else "sup",
            "balance_residual",
            "res",
            "unspecified",
            row.pefa_product,
            abs(residual) * 1000.0,
            "explicit_rounding_or_scope_balance_residual",
            "constructed_from_PEFA_identity",
        )
    if len(physical_internal) != len(physical):
        physical = pd.DataFrame(physical_internal).groupby(
            group_key, dropna=False, as_index=False
        )["pj"].sum()

    cp15["prd_norm"] = cp15["prd_amo"].replace(
        {"CPA_E37-39": "CPA_E37-E39"}
    )
    cp15["ind_norm"] = cp15["ind_impv"].map(norm_a64)
    cp16["prd_norm"] = cp16["prd_ava"].replace(
        {"CPA_E37-39": "CPA_E37-E39"}
    )
    cp16["ind_norm"] = cp16["ind_use"].map(norm_a64)
    cp15 = cp15[cp15["unit"] == "MIO_NAC"].copy()
    cp16 = cp16[cp16["unit"] == "MIO_NAC"].copy()
    cpas = sorted(set(CPA_BY_PEFA.values()))

    # `source_value()` cannot distinguish "Eurostat reports a genuine zero"
    # from "Eurostat has no row at all for this CPA code" (both collapse to
    # 0.0 once filtered). A CPA with zero rows anywhere in the dataset means
    # every basic/purchaser value derived for it is unobserved, not
    # calibrated, and must be flagged explicitly rather than silently
    # written as a clean zero.
    cp15_cpa_has_data = {
        cpa: bool((cp15["prd_norm"] == cpa).any()) for cpa in cpas
    }
    cp16_cpa_has_data = {
        cpa: bool((cp16["prd_norm"] == cpa).any()) for cpa in cpas
    }

    valuation_rows = []
    for cpa in cpas:
        values = {
            item: source_value(cp15, prd_norm=cpa, ind_norm=item) / 1000.0
            for item in ["TS_BP", "TS_PP", "OTTM", "D21X31", "TOTAL", "P7"]
        }
        valuation_rows.append({"cpa": cpa, **values})
    valuation = pd.DataFrame(valuation_rows)
    valuation["identity_residual"] = (
        valuation["TS_PP"]
        - valuation["TS_BP"]
        - valuation["OTTM"]
        - valuation["D21X31"]
    )
    valuation["basic_supply_residual"] = (
        valuation["TS_BP"] - valuation["TOTAL"] - valuation["P7"]
    )
    valuation["cp15_has_source_rows"] = valuation["cpa"].map(cp15_cpa_has_data)
    valuation["cp16_has_source_rows"] = valuation["cpa"].map(cp16_cpa_has_data)
    valuation_index = valuation.set_index("cpa")

    def user_block(row: pd.Series) -> str:
        if row["flow"] == "cons_inter":
            return f"industry:{row['indu']}"
        if row["flow"] == "cons_hh":
            return "household"
        if row["flow"] == "export":
            return "export"
        if row["flow"] == "invent_change":
            return "inventory"
        return f"residual:{row['flow']}"

    uses = physical[physical["bal"] == "use"].copy()
    uses["user_block"] = uses.apply(user_block, axis=1)
    control_rows: list[dict] = []
    redirect_rows: list[dict] = []
    for cpa in cpas:
        for cluster in industry_clusters.itertuples():
            value = sum(
                source_value(cp16, prd_norm=cpa, ind_norm=nace) / 1000.0
                for nace in cluster.right_members.split("|")
            )
            if cluster.use_detail_available:
                target_block = f"industry:{cluster.representative}"
                source_label = "naio_10_cp16_A64_cluster"
            else:
                target_block = "industry:res"
                source_label = (
                    "naio_10_cp16_A64_cluster_redirected_no_PEFA_USE_detail"
                )
                if abs(value) > 1e-10:
                    redirect_rows.append(
                        {
                            "cpa": cpa,
                            "from_industry": cluster.representative,
                            "nace_members": cluster.right_members,
                            "redirected_purch_control_bn": value,
                            "reason": (
                                "PEFA reports no USE row for any of this "
                                "cluster's NACE codes; its physical energy "
                                "is already inside the explicit indu=res "
                                "reporting-detail residual, so its SUT "
                                "money is pooled with that same bucket "
                                "instead of becoming unmatched residual"
                            ),
                        }
                    )
            control_rows.append(
                {
                    "cpa": cpa,
                    "user_block": target_block,
                    "purch_control": value,
                    "control_source": source_label,
                }
            )
        for block, source_codes in [
            ("household", ["P3_S14", "P3_S15"]),
            ("export", ["P6"]),
            ("inventory", ["P52"]),
        ]:
            value = sum(
                source_value(cp16, prd_norm=cpa, ind_norm=code) / 1000.0
                for code in source_codes
            )
            control_rows.append(
                {
                    "cpa": cpa,
                    "user_block": block,
                    "purch_control": value,
                    "control_source": f"naio_10_cp16_{'+'.join(source_codes)}",
                }
            )
    controls = pd.DataFrame(control_rows)
    controls = controls.groupby(
        ["cpa", "user_block"], as_index=False
    ).agg(
        purch_control=("purch_control", "sum"),
        control_source=("control_source", lambda s: "; ".join(sorted(set(s)))),
    )
    redirects = pd.DataFrame(
        redirect_rows,
        columns=[
            "cpa", "from_industry", "nace_members",
            "redirected_purch_control_bn", "reason",
        ],
    )
    included = controls.groupby("cpa")["purch_control"].sum()
    controls = pd.concat(
        [
            controls,
            pd.DataFrame(
                [
                    {
                        "cpa": cpa,
                        "user_block": "residual:other_final_use",
                        "purch_control": (
                            source_value(cp16, prd_norm=cpa, ind_norm="TU")
                            / 1000.0
                            - included.get(cpa, 0.0)
                        ),
                        "control_source": "naio_10_cp16_TU_minus_included",
                    }
                    for cpa in cpas
                ]
            ),
        ],
        ignore_index=True,
    )
    controls["cp16_has_source_rows"] = controls["cpa"].map(cp16_cpa_has_data)

    def initial_price_weight(row: pd.Series) -> tuple[float, str]:
        code = row["pefa_product"]
        user = "household" if row["flow"] == "cons_hh" else "non_household"
        price = price_map.get((code, user), price_map.get((code, "all")))
        if price is None or not np.isfinite(price) or price <= 0:
            return max(row["pj"], 0.0), "PEFA_PJ_fallback"
        return max(row["pj"], 0.0) * float(price), "public_price_times_PEFA_PJ"

    uses[["allocation_weight", "weight_status"]] = uses.apply(
        lambda row: pd.Series(initial_price_weight(row)), axis=1
    )
    uses = uses.merge(controls, on=["cpa", "user_block"], how="left")
    # cp16_has_source_rows is per-CPA audit metadata for the `controls`/
    # `valuation` sheets only; drop the incidental merge copy so it can't
    # go stale (NaN) on residual rows appended below via pd.concat.
    uses = uses.drop(columns="cp16_has_source_rows")
    uses["purch_control"] = uses["purch_control"].fillna(0.0)
    uses["block_weight_sum"] = uses.groupby(
        ["cpa", "user_block"], dropna=False
    )["allocation_weight"].transform("sum")
    uses["purch"] = np.where(
        uses["block_weight_sum"] > 0,
        uses["purch_control"]
        * uses["allocation_weight"]
        / uses["block_weight_sum"],
        0.0,
    )

    allocated_blocks = uses.groupby(["cpa", "user_block"])[
        "purch"
    ].sum()
    residual_rows: list[dict] = []
    for control in controls.itertuples():
        allocated = allocated_blocks.get((control.cpa, control.user_block), 0.0)
        residual = control.purch_control - allocated
        if abs(residual) <= 1e-10:
            continue
        if control.user_block.startswith("industry:"):
            flow = "cons_inter"
            indu = control.user_block.split(":", 1)[1]
        elif control.user_block == "household":
            flow, indu = "cons_hh", "cHouEne"
        elif control.user_block == "export":
            flow, indu = "export", ""
        elif control.user_block == "inventory":
            flow, indu = "invent_change", ""
        else:
            flow, indu = "other_use", "res"
        residual_rows.append(
            {
                "year": year,
                "bal": "use",
                "flow": flow,
                "indu": indu,
                "purp": "unspecified",
                "product": f"monetary_residual_{control.cpa}",
                "pefa_product": f"RES_{control.cpa}",
                "cpa": control.cpa,
                "pj": 0.0,
                "physical_status": "no_direct_physical_quantity",
                "source_nace": "",
                "user_block": control.user_block,
                "allocation_weight": 0.0,
                "weight_status": "explicit_unallocated_SUT_residual",
                "purch_control": control.purch_control,
                "control_source": control.control_source,
                "block_weight_sum": 0.0,
                "purch": residual,
            }
        )
        product_text[f"monetary_residual_{control.cpa}"] = (
            f"Explicit monetary residual {control.cpa}"
        )
    uses = pd.concat([uses, pd.DataFrame(residual_rows)], ignore_index=True)

    uses = uses.merge(
        valuation[["cpa", "TS_BP", "TS_PP", "OTTM", "D21X31"]],
        on="cpa",
        how="left",
    )
    uses["margin_wedge"] = np.where(
        uses["TS_PP"].abs() > 1e-12,
        uses["purch"] * uses["OTTM"] / uses["TS_PP"],
        0.0,
    )
    uses["tax_wedge"] = np.where(
        uses["TS_PP"].abs() > 1e-12,
        uses["purch"] * uses["D21X31"] / uses["TS_PP"],
        0.0,
    )
    uses["basic"] = uses["purch"] - uses["margin_wedge"] - uses["tax_wedge"]
    uses["vat_legal_estimate"] = np.where(
        uses["flow"] == "cons_hh",
        uses["purch"] * VAT_RATE / (1.0 + VAT_RATE),
        0.0,
    )
    uses["vat"] = np.where(
        uses["tax_wedge"] >= 0,
        np.minimum(uses["vat_legal_estimate"], uses["tax_wedge"]),
        0.0,
    )
    uses["ener_tax"] = uses["tax_wedge"] - uses["vat"]
    uses["ws_marg"] = uses["margin_wedge"]
    for field in ["ret_marg", "mvs_marg", "co2_tax", "so2_tax", "nox_tax", "pso_tax"]:
        uses[field] = 0.0
    for field in EMISSION_FIELDS:
        uses[field] = 0.0

    supplies = physical[physical["bal"] == "sup"].copy()
    supplies["basic"] = 0.0
    supply_residual_rows: list[dict] = []
    for cpa in cpas:
        cpa_supply = supplies[supplies["cpa"] == cpa]
        for supply_flow, control_name in [
            ("production", "TOTAL"),
            ("import", "P7"),
        ]:
            block = cpa_supply[cpa_supply["flow"] == supply_flow]
            control = float(valuation_index.loc[cpa, control_name])
            if supply_flow == "production":
                producer_controls = []
                for cluster in industry_clusters.itertuples():
                    value = sum(
                        source_value(
                            cp15, prd_norm=cpa, ind_norm=nace
                        )
                        / 1000.0
                        for nace in cluster.right_members.split("|")
                    )
                    producer_controls.append((cluster.representative, value))
            else:
                producer_controls = [("", control)]
            for indu, producer_control in producer_controls:
                mask = (
                    (supplies["cpa"] == cpa)
                    & (supplies["flow"] == supply_flow)
                    & (supplies["indu"] == indu)
                )
                weight = supplies.loc[mask, "pj"].clip(lower=0).sum()
                if weight > 0:
                    supplies.loc[mask, "basic"] = (
                        producer_control
                        * supplies.loc[mask, "pj"].clip(lower=0)
                        / weight
                    )
                elif abs(producer_control) > 1e-10:
                    supply_residual_rows.append(
                        {
                            "year": year,
                            "bal": "sup",
                            "flow": supply_flow,
                            "indu": indu,
                            "purp": "unspecified",
                            "product": f"monetary_residual_{cpa}",
                            "pefa_product": f"RES_{cpa}",
                            "cpa": cpa,
                            "pj": 0.0,
                            "physical_status": "no_direct_physical_quantity",
                            "source_nace": "",
                            "basic": producer_control,
                        }
                    )
    supplies = pd.concat(
        [supplies, pd.DataFrame(supply_residual_rows)], ignore_index=True
    )
    for field in [*EMISSION_FIELDS, *[x for x in MONEY_FIELDS if x != "basic"]]:
        supplies[field] = 0.0
    supplies["purch"] = 0.0

    runtime_internal = pd.concat([supplies, uses], ignore_index=True)
    runtime_internal["component_sum"] = runtime_internal[
        ["basic", "ws_marg", "ret_marg", "mvs_marg", "ener_tax", "co2_tax",
         "so2_tax", "nox_tax", "pso_tax", "vat"]
    ].sum(axis=1)
    runtime_internal["identity_residual"] = np.where(
        runtime_internal["bal"] == "use",
        runtime_internal["purch"] - runtime_internal["component_sum"],
        0.0,
    )
    runtime = runtime_internal[ENERGY_COLUMNS].copy()
    runtime = runtime.groupby(
        ["year", "bal", "flow", "indu", "purp", "product"],
        dropna=False,
        as_index=False,
    )[[*EMISSION_FIELDS, "pj", *MONEY_FIELDS]].sum()
    runtime = runtime.sort_values(
        ["year", "bal", "flow", "indu", "purp", "product"],
        kind="stable",
    ).reset_index(drop=True)

    io_rows: list[dict] = []
    supply_basic = runtime_internal[
        (runtime_internal["bal"] == "sup")
        & runtime_internal["flow"].isin(["production", "import"])
    ]
    producer_shares: dict[tuple[str, str], list[tuple[str, float]]] = {}
    for cpa in cpas:
        domestic = supply_basic[
            (supply_basic["cpa"] == cpa)
            & (supply_basic["flow"] == "production")
        ].groupby("indu")["basic"].sum()
        if domestic.sum() > 0:
            producer_shares[(cpa, "production")] = [
                (str(indu), float(value / domestic.sum()))
                for indu, value in domestic.items()
                if abs(value) > 1e-12
            ]
        imports = supply_basic[
            (supply_basic["cpa"] == cpa)
            & (supply_basic["flow"] == "import")
        ]["basic"].sum()
        producer_shares[(cpa, "import")] = [
            (CPA_PRODUCER[cpa], 1.0)
        ] if abs(imports) > 1e-12 else []

    def io_column(row: pd.Series) -> tuple[str, str]:
        if row["flow"] == "cons_inter":
            return "cons_inter", row["indu"]
        if row["flow"] == "cons_hh":
            return "cons_hh", row["indu"]
        if row["flow"] == "export":
            return "export", ""
        if row["flow"] == "invent_change":
            return "invent_change", ""
        return "cons_inter", "res"

    for row in runtime_internal[runtime_internal["bal"] == "use"].iterrows():
        record = row[1]
        if not record.get("cpa") or pd.isna(record.get("cpa")):
            continue
        col_l1, col_l2 = io_column(record)
        cpa = record["cpa"]
        bp = float(valuation_index.loc[cpa, "TS_BP"])
        domestic = float(valuation_index.loc[cpa, "TOTAL"])
        imported = float(valuation_index.loc[cpa, "P7"])
        domestic_share = domestic / bp if abs(bp) > 1e-12 else 0.0
        import_share = imported / bp if abs(bp) > 1e-12 else 0.0
        for supply_flow, share in [
            ("production", domestic_share),
            ("import", import_share),
        ]:
            for producer, producer_share in producer_shares.get(
                (cpa, supply_flow), []
            ):
                value = record["basic"] * share * producer_share
                if abs(value) > 1e-12:
                    io_rows.append(
                        {
                            "year": year,
                            "row_l1": supply_flow,
                            "row_l2": producer,
                            "col_l1": col_l1,
                            "col_l2": col_l2,
                            "value": value,
                        }
                    )
        for row_l1, row_l2, field in [
            ("production", "46000", "ws_marg"),
            ("prim_input", "tax_products", "ener_tax"),
            ("prim_input", "tax_vat", "vat"),
        ]:
            value = float(record[field])
            if abs(value) > 1e-12:
                io_rows.append(
                    {
                        "year": year,
                        "row_l1": row_l1,
                        "row_l2": row_l2,
                        "col_l1": col_l1,
                        "col_l2": col_l2,
                        "value": value,
                    }
                )
    io = pd.DataFrame(io_rows, columns=IO_COLUMNS)
    io = (
        io.groupby(IO_COLUMNS[:-1], dropna=False, as_index=False)["value"]
        .sum()
        .sort_values(IO_COLUMNS[:-1], kind="stable")
        .reset_index(drop=True)
    )

    # GDX: average allocated aggregate duty per PJ is transparently assumed
    # marginal. Product names match read_data.py's schema mapping.
    marginal_source = runtime_internal[
        (runtime_internal["bal"] == "use")
        & (runtime_internal["pj"] > 0)
    ].copy()
    marginal_source["r"] = marginal_source["indu"].replace(
        {"": "res"}
    )
    marginal_source.loc[
        marginal_source["flow"] == "export", "r"
    ] = "xOth"
    marginal_source.loc[
        marginal_source["flow"] == "invent_change", "r"
    ] = "invt"
    marginal_source["energy19"] = marginal_source["product"].map(
        lambda item: product_text.get(item, item)
    )
    marginal_source["level"] = (
        marginal_source["ener_tax"] / marginal_source["pj"] * 1000.0
    )
    marginal_eafg = (
        marginal_source.groupby(
            ["year", "energy19", "purp", "r"], as_index=False
        )
        .apply(
            lambda group: pd.Series(
                {
                    "level": (
                        group["ener_tax"].sum() / group["pj"].sum() * 1000.0
                    )
                }
            ),
            include_groups=False,
        )
        .reset_index()
        .rename(columns={"year": "t", "purp": "purpose"})
    )
    if "level_0" in marginal_eafg:
        marginal_eafg = marginal_eafg.drop(columns="level_0")
    if "index" in marginal_eafg:
        marginal_eafg = marginal_eafg.drop(columns="index")
    marginal_eafg["t"] = marginal_eafg["t"].astype(str)
    marginal_co2 = marginal_eafg[
        ["t", "energy19", "purpose", "r"]
    ].copy()
    marginal_co2["emm_eq"] = "co2ubio"
    marginal_co2["level"] = 0.0
    gdx_path = package / "EU_GR_data.gdx"
    container = gp.Container()
    t_set = gp.Set(container, "t", records=sorted(marginal_eafg["t"].unique()))
    e_set = gp.Set(
        container, "energy19",
        records=sorted(marginal_eafg["energy19"].unique()),
    )
    p_set = gp.Set(
        container, "purpose",
        records=sorted(marginal_eafg["purpose"].unique()),
    )
    r_set = gp.Set(container, "r", records=sorted(marginal_eafg["r"].unique()))
    em_set = gp.Set(container, "emm_eq", records=["co2ubio"])
    gp.Variable(
        container,
        "tEAFG_REmarg",
        domain=[t_set, e_set, p_set, r_set],
        records=marginal_eafg,
    )
    gp.Variable(
        container,
        "tCO2_REmarg",
        domain=[t_set, e_set, p_set, r_set, em_set],
        records=marginal_co2,
    )
    container.write(str(gdx_path))

    metadata_rows = pd.DataFrame(
        [
            {"field": "country", "value": country},
            {"field": "reference_year", "value": year},
            {"field": "policy_version", "value": POLICY_VERSION},
            {"field": "retrieval_date", "value": RETRIEVAL_DATE},
            {
                "field": "hard_truth",
                "value": (
                    "Monetary product-user-purpose cells are calibrated/modelled, "
                    "not directly observed; residuals remain explicit."
                ),
            },
            {
                "field": "currency",
                "value": "bn SEK; physical PJ; emissions kt (zeros in core)",
            },
            {
                "field": "emissions_status",
                "value": (
                    "Air-account totals are audit controls only; not allocated "
                    "to products because total AEA includes process emissions."
                ),
            },
        ]
    )
    energy_path = package / "energy_and_emissions.xlsx"
    io_path = package / "io_energy_long_format.xlsx"
    with pd.ExcelWriter(energy_path, engine="openpyxl") as writer:
        runtime.to_excel(writer, sheet_name="ems_energy", index=False)
        metadata_rows.to_excel(writer, sheet_name="metadata", index=False)
    with pd.ExcelWriter(io_path, engine="openpyxl") as writer:
        io.to_excel(writer, sheet_name="io", index=False)
        metadata_rows.to_excel(writer, sheet_name="metadata", index=False)

    component_identity = runtime_internal[
        runtime_internal["bal"] == "use"
    ][
        [
            "year", "flow", "indu", "purp", "product", "pj", "basic",
            "ws_marg", "ener_tax", "vat", "purch", "component_sum",
            "identity_residual", "physical_status",
        ]
    ].copy()
    final_balance = (
        runtime.groupby(["product", "bal"])["pj"]
        .sum()
        .unstack(fill_value=0.0)
        .reset_index()
    )
    final_balance["residual_sup_minus_use"] = (
        final_balance.get("sup", 0.0) - final_balance.get("use", 0.0)
    )
    sut_comparison = runtime_internal.groupby(["cpa", "bal"])[
        ["basic", "ws_marg", "ener_tax", "vat", "purch"]
    ].sum().reset_index()
    use_sut = sut_comparison[sut_comparison["bal"] == "use"].merge(
        valuation, on="cpa", how="left"
    )
    use_sut["basic_residual"] = use_sut["basic"] - use_sut["TS_BP"]
    use_sut["margin_residual"] = use_sut["ws_marg"] - use_sut["OTTM"]
    use_sut["tax_residual"] = (
        use_sut["ener_tax"] + use_sut["vat"] - use_sut["D21X31"]
    )
    use_sut["purch_residual"] = use_sut["purch"] - use_sut["TS_PP"]

    tax["nace_norm"] = tax["nace_r2"].map(norm_a64)
    tax_nrg = tax[
        (tax["tax"] == "NRG")
        & (tax["unit"] == "MIO_EUR")
        & tax["nace_norm"].isin(a64_codes)
    ]
    tax_controls = []
    for cluster in industry_clusters.itertuples():
        source_total = (
            tax_nrg[
                tax_nrg["nace_norm"].isin(cluster.right_members.split("|"))
            ]["value"].sum()
            * exchange_rate
            / 1000.0
        )
        allocated = runtime_internal[
            (runtime_internal["bal"] == "use")
            & (runtime_internal["indu"] == cluster.representative)
        ]["ener_tax"].sum()
        tax_controls.append(
            {
                "representative": cluster.representative,
                "nace_cluster": cluster.right_members,
                "env_ac_taxind2_NRG_bn_SEK": source_total,
                "allocated_SUT_nonVAT_wedge_bn_SEK": allocated,
                "residual_allocated_minus_tax_account": allocated - source_total,
                "interpretation": (
                    "secondary payer control; not forced because SUT wedge "
                    "includes broader net product taxes"
                ),
            }
        )
    tax_comparison = pd.DataFrame(tax_controls)

    aea_summary = (
        aea[aea["nace_r2"].isin(["TOTAL", "HH"])]
        .groupby(["airpol", "unit"], as_index=False)["value"]
        .sum()
    )
    aea_summary["runtime_allocation"] = 0.0
    aea_summary["reason"] = (
        "total air account includes energy and process emissions; no defensible "
        "product allocation in this coarse monetary core"
    )
    calibration = uses[
        [
            "cpa", "user_block", "product", "pj", "allocation_weight",
            "weight_status", "purch_control", "block_weight_sum", "purch",
            "vat_legal_estimate", "vat", "ener_tax",
        ]
    ].copy()
    ras_iterations = pd.DataFrame(
        [
            {
                "stage": "initial public price/PEFA weights",
                "iteration": 0,
                "hard_control": "none",
                "max_abs_control_residual_bn_SEK": np.nan,
                "note": "starting weights only",
            },
            {
                "stage": "block calibration",
                "iteration": 1,
                "hard_control": "naio_10_cp16 CPA×user purchaser controls",
                "max_abs_control_residual_bn_SEK": float(
                    use_sut["purch_residual"].abs().max()
                ),
                "note": (
                    "one-dimensional block scaling closed controls; two-way RAS "
                    "was not required. Tax-account mismatch remains explicit."
                ),
            },
        ]
    )
    anomalies = []
    supply_residual_total = float(
        runtime_internal.loc[
            (runtime_internal["bal"] == "sup")
            & runtime_internal["product"].astype(str).str.startswith(
                "monetary_residual_"
            ),
            "basic",
        ].abs().sum()
    )
    if supply_residual_total > 1e-9:
        anomalies.append(
            {
                "severity": "INFO",
                "item": "supply_side_monetary_residual",
                "value": supply_residual_total,
                "unit": "bn SEK",
                "description": (
                    "producer-side (basic value) unmatched SUT control, "
                    "mainly cp15 CPA totals for industries producing "
                    "little/no physical quantity of the CPA's mapped "
                    "energy product (e.g. the waste/sewerage industry's "
                    "own broad CPA_E37-E39 output); not counted in "
                    "explicit_monetary_residual_bn_SEK because that field "
                    "reads `purch`, which is always 0 on the supply side "
                    "— see explicit_supply_side_monetary_residual_bn_SEK"
                ),
            }
        )
    if not redirects.empty:
        for cpa, group in redirects.groupby("cpa"):
            anomalies.append(
                {
                    "severity": "INFO",
                    "item": cpa,
                    "value": float(group["redirected_purch_control_bn"].sum()),
                    "unit": "bn SEK",
                    "description": (
                        f"{len(group)} industry control(s) redirected to "
                        "indu=res because PEFA reports no USE row at "
                        "GREU's expected NACE detail; see "
                        "reporting_detail_redirects sheet"
                    ),
                }
            )
    for row in final_balance[
        final_balance["residual_sup_minus_use"].abs() > 1e-9
    ].itertuples():
        anomalies.append(
            {
                "severity": "ERROR",
                "item": row.product,
                "value": row.residual_sup_minus_use,
                "unit": "PJ",
                "description": "physical supply/use does not close",
            }
        )
    for row in valuation[
        valuation["identity_residual"].abs() > 1e-8
    ].itertuples():
        anomalies.append(
            {
                "severity": "ERROR",
                "item": row.cpa,
                "value": row.identity_residual,
                "unit": "bn SEK",
                "description": "published SUT valuation identity residual",
            }
        )
    for row in runtime_internal[
        runtime_internal["basic"] < -1e-10
    ].itertuples():
        anomalies.append(
            {
                "severity": "CHECK",
                "item": f"{row.flow}/{row.indu}/{row.product}",
                "value": row.basic,
                "unit": "bn SEK",
                "description": "negative calibrated basic value retained",
            }
        )
    anomalies.append(
        {
            "severity": "INFO",
            "item": "emissions",
            "value": 0.0,
            "unit": "runtime rows",
            "description": (
                "AEA source inspected but not allocated; process/energy split "
                "is unavailable"
            ),
        }
    )
    for cpa in cpas:
        if not cp16_cpa_has_data.get(cpa, True):
            affected_pj = float(uses.loc[uses["cpa"] == cpa, "pj"].sum())
            anomalies.append(
                {
                    "severity": "ERROR",
                    "item": cpa,
                    "value": affected_pj,
                    "unit": "PJ",
                    "description": (
                        "naio_10_cp16 publishes no rows at all for this CPA "
                        "in this country/year (not a published zero); every "
                        "use-side purchaser value for it is unobserved, not "
                        "calibrated, and is written as 0.0 rather than "
                        "invented — see valuation_controls/user_controls "
                        "'cp16_has_source_rows'"
                    ),
                }
            )
        if not cp15_cpa_has_data.get(cpa, True):
            affected_pj_sup = float(
                supplies.loc[supplies["cpa"] == cpa, "pj"].sum()
            )
            anomalies.append(
                {
                    "severity": "ERROR",
                    "item": cpa,
                    "value": affected_pj_sup,
                    "unit": "PJ",
                    "description": (
                        "naio_10_cp15 publishes no rows at all for this CPA "
                        "in this country/year (not a published zero); every "
                        "supply-side basic value for it is unobserved, not "
                        "calibrated, and is written as 0.0 rather than "
                        "invented — see valuation_controls "
                        "'cp15_has_source_rows'"
                    ),
                }
            )
    anomalies_df = pd.DataFrame(anomalies)

    raw_manifest = json.loads((raw / "manifest.json").read_text(encoding="utf-8"))
    source_register = pd.DataFrame(raw_manifest["raw_files"])
    compatibility = pd.DataFrame(
        [
            {"check": "energy first sheet", "result": "PASS", "detail": "ems_energy"},
            {"check": "io first sheet", "result": "PASS", "detail": "io"},
            {
                "check": "energy unique keys",
                "result": "PASS" if not runtime.duplicated(
                    ["year", "bal", "flow", "indu", "purp", "product"]
                ).any() else "FAIL",
                "detail": len(runtime),
            },
            {
                "check": "io unique keys",
                "result": "PASS" if not io.duplicated(IO_COLUMNS[:-1]).any() else "FAIL",
                "detail": len(io),
            },
            {
                "check": "component identity",
                "result": "PASS" if component_identity["identity_residual"].abs().max() < 1e-9 else "FAIL",
                "detail": component_identity["identity_residual"].abs().max(),
            },
            {
                "check": "physical product balance",
                "result": "PASS" if final_balance["residual_sup_minus_use"].abs().max() < 1e-9 else "FAIL",
                "detail": final_balance["residual_sup_minus_use"].abs().max(),
            },
            {
                "check": "direct complete monetary cells",
                "result": "0",
                "detail": "all monetary cells calibrated/modelled",
            },
            {
                "check": "Danish production values/shares",
                "result": "PASS",
                "detail": "none read; metadata.xlsx classifications only",
            },
        ]
    )

    audit_sheets = {
        "metadata": metadata_rows,
        "source_register": source_register,
        "policy": pd.DataFrame(
            [{"rule": key, "value": json.dumps(value) if isinstance(value, list) else value}
             for key, value in policy.items()]
        ),
        "industry_clusters": industry_clusters,
        "product_clusters": product_clusters,
        "reporting_detail_redirects": redirects,
        "physical_rows": physical,
        "physical_balance": final_balance,
        "valuation_controls": valuation,
        "user_controls": controls,
        "allocation_weights": calibration,
        "ras_iterations": ras_iterations,
        "component_identity": component_identity,
        "sut_comparison": use_sut,
        "tax_comparison": tax_comparison,
        "price_controls": prices,
        "air_account_control": aea_summary,
        "marginal_rates": marginal_eafg,
        "compatibility": compatibility,
        "anomalies": anomalies_df,
    }
    with pd.ExcelWriter(audit_path, engine="openpyxl") as writer:
        for sheet, frame in audit_sheets.items():
            frame.to_excel(writer, sheet_name=sheet[:31], index=False)
    style_workbook(audit_path)

    manifest = {
        "country": country,
        "reference_year": year,
        "policy_version": POLICY_VERSION,
        "retrieval_date": RETRIEVAL_DATE,
        "currency": "bn SEK",
        "hard_truth": (
            "No monetary cell is directly observed at product×user×purpose; "
            "all monetary allocations are calibrated and residuals are explicit."
        ),
        "danish_inputs": "none; metadata.xlsx classifications only",
        "scripts": {
            "downloader": {
                "path": "data/preprocessing/scripts/download_energy_money_public_core.py",
                "sha256": file_sha256(
                    DATA.parent / "scripts" / "download_energy_money_public_core.py"
                ),
            },
            "builder": {
                "path": "data/preprocessing/scripts/build_energy_money_public_core.py",
                "sha256": file_sha256(Path(__file__).resolve()),
            },
        },
        "policy": {
            "path": (
                str(policy_path.relative_to(REPO))
                if policy_path.is_relative_to(REPO)
                else str(policy_path)
            ),
            "sha256": file_sha256(policy_path),
        },
        "artifacts": {
            path.name: {
                "sha256": file_sha256(path),
                **(
                    {"rows": len(runtime)}
                    if path == energy_path
                    else {"rows": len(io)}
                    if path == io_path
                    else {}
                ),
            }
            for path in [energy_path, io_path, gdx_path]
        },
        "raw_manifest": str((raw / "manifest.json").relative_to(REPO)),
        "raw_manifest_sha256": file_sha256(raw / "manifest.json"),
        "audit_workbook": (
            str(audit_path.relative_to(REPO))
            if audit_path.is_relative_to(REPO)
            else str(audit_path)
        ),
        "audit_sha256": file_sha256(audit_path),
        "quantitative_results": {
            "physical_supply_PJ": float(
                runtime.loc[runtime["bal"] == "sup", "pj"].sum()
            ),
            "physical_use_PJ": float(
                runtime.loc[runtime["bal"] == "use", "pj"].sum()
            ),
            "max_product_balance_residual_PJ": float(
                final_balance["residual_sup_minus_use"].abs().max()
            ),
            "purchaser_value_bn_SEK": float(
                runtime.loc[runtime["bal"] == "use", "purch"].sum()
            ),
            "max_component_identity_residual_bn_SEK": float(
                component_identity["identity_residual"].abs().max()
            ),
            "max_SUT_purchaser_residual_bn_SEK": float(
                use_sut["purch_residual"].abs().max()
            ),
            "explicit_physical_residual_PJ": float(
                runtime_internal.loc[
                    runtime_internal["physical_status"].astype(str).str.contains(
                        "residual"
                    ),
                    "pj",
                ].abs().sum()
            ),
            "explicit_monetary_residual_bn_SEK": float(
                runtime_internal.loc[
                    runtime_internal["product"].astype(str).str.startswith(
                        "monetary_residual_"
                    ),
                    "purch",
                ].abs().sum()
            ),
            "explicit_supply_side_monetary_residual_bn_SEK": float(
                runtime_internal.loc[
                    (runtime_internal["bal"] == "sup")
                    & runtime_internal["product"].astype(str).str.startswith(
                        "monetary_residual_"
                    ),
                    "basic",
                ].abs().sum()
            ),
            "explicit_supply_side_monetary_residual_note": (
                "use-side residuals are measured on `purch`; supply-side "
                "residuals live in `basic` because supply rows never carry "
                "a purchaser price. This field makes that separate, "
                "otherwise invisible, producer-side unmatched-control "
                "total explicit rather than silently absent from the "
                "headline use-side residual figure."
            ),
            "reporting_detail_redirected_to_res_bn_SEK": float(
                redirects["redirected_purch_control_bn"].sum()
                if not redirects.empty
                else 0.0
            ),
            "reporting_detail_redirected_note": (
                "SUT money for industry clusters where PEFA has no USE row "
                "at GREU's expected NACE detail is pooled with the "
                "indu=res reporting-detail residual (same physical bucket) "
                "instead of becoming unmatched residual; see "
                "reporting_detail_redirects audit sheet. Not a new "
                "allocation rule, only a relabelling of which explicit "
                "bucket already-disclosed money/quantity sit in."
            ),
            "direct_monetary_cells": 0,
            "modelled_monetary_cells": int(
                (
                    runtime.loc[runtime["bal"] == "use", "purch"].abs()
                    > 1e-12
                ).sum()
            ),
            "vat_legal_estimate_bn_SEK": float(
                uses["vat_legal_estimate"].sum()
            ),
            "vat_calibrated_bn_SEK": float(uses["vat"].sum()),
            "vat_control_residual_bn_SEK": float(
                (uses["vat_legal_estimate"] - uses["vat"]).sum()
            ),
            "tax_account_control_bn_SEK": float(
                tax_comparison["env_ac_taxind2_NRG_bn_SEK"].sum()
            ),
            "allocated_nonVAT_wedge_bn_SEK": float(
                tax_comparison["allocated_SUT_nonVAT_wedge_bn_SEK"].sum()
            ),
            "tax_control_residual_bn_SEK": float(
                tax_comparison[
                    "residual_allocated_minus_tax_account"
                ].sum()
            ),
            "negative_basic_rows": int(
                (runtime_internal["basic"] < -1e-10).sum()
            ),
            "negative_basic_total_bn_SEK": float(
                runtime_internal.loc[
                    runtime_internal["basic"] < -1e-10, "basic"
                ].sum()
            ),
            "unobserved_no_source_breakdown_use_PJ": float(
                uses.loc[
                    uses["cpa"].map(
                        lambda cpa: not cp16_cpa_has_data.get(cpa, True)
                    ),
                    "pj",
                ].sum()
            ),
            "unobserved_no_source_breakdown_supply_PJ": float(
                supplies.loc[
                    supplies["cpa"].map(
                        lambda cpa: not cp15_cpa_has_data.get(cpa, True)
                    ),
                    "pj",
                ].sum()
            ),
            "unobserved_no_source_breakdown_cpas": sorted(
                cpa
                for cpa in cpas
                if not cp16_cpa_has_data.get(cpa, True)
                or not cp15_cpa_has_data.get(cpa, True)
            ),
        },
    }
    manifest_path = package / "energy_money_manifest.json"
    manifest_path.write_text(
        json.dumps(manifest, indent=2, sort_keys=True) + "\n",
        encoding="utf-8",
    )
    package_readme = f"""# Sweden 2020 monetary-energy public core

Generated by `data/preprocessing/scripts/build_energy_money_public_core.py`
under policy `{POLICY_VERSION}` from public European sources retrieved
{RETRIEVAL_DATE}. This directory is a runtime compatibility package, not a full
Sweden GREU model dataset.

The physical account is controlled to Eurostat PEFA. Monetary cells are
calibrated to Sweden's national SUT and public price/tax controls; none is
directly observed at product × user × purpose. Connected NACE–GREU clusters
prevent splitting A64 categories with Danish weights. Explicit residual rows
and the audit workbook preserve unresolved reporting and allocation gaps.

**Known gap — no SUT breakdown for {", ".join(manifest["quantitative_results"]["unobserved_no_source_breakdown_cpas"]) or "none"}:**
Eurostat's `naio_10_cp15`/`naio_10_cp16` publish zero rows at all (not a
published zero) for these CPA codes for Sweden 2020, covering
{manifest["quantitative_results"]["unobserved_no_source_breakdown_use_PJ"]:.3f} PJ
of physical energy (present on both the supply and use side of the
accounting identity). Every monetary value for that energy is therefore
unobserved, not calibrated — it is written as 0.0 because there is no
control to scale to, and is flagged in the `anomalies` and
`valuation_controls` sheets of the audit workbook rather than passing
silently as a clean zero.

**Unmatched-SUT-residual narrowing (2026-07-31):** Sweden's PEFA reports
USE-side energy consumption for manufacturing (NACE section C), agriculture
(A), water/waste (E36/E37-E39) and trade/transport (G45-47/H49-53) only at
the whole-section level, not at the finer NACE divisions (C16, C19, E37-E39,
...) GREU's concordance expects. That physical energy is not missing — it is
already the explicit `indu=res` reporting-detail residual — but before this
fix, `naio_10_cp16`'s division-level purchaser-value controls for those same
industries had no matching physical row and were booked as pure unmatched
residual. The builder now pools that money with the same `res` bucket that
already holds its physical counterpart (see `reporting_detail_redirects` and
the `policy.reporting_detail_redirect_rule` sheets/fields), redirecting
**{manifest["quantitative_results"]["reporting_detail_redirected_to_res_bn_SEK"]:.3f} bn SEK**
that is no longer unmatched. This is a relabelling of an already-disclosed
gap, not a new allocation assumption: industries PEFA does detail (e.g.
financial/business services) are never redirected, and their SUT money for
CPA codes with no matching PEFA product for them remains a genuine residual
because it reflects real non-energy spending within a too-broad CPA (e.g.
office furniture within CPA_C16 wood products), not a reporting gap.

Runtime artifacts:
- `energy_and_emissions.xlsx`
- `io_energy_long_format.xlsx`
- `EU_GR_data.gdx`

Provenance and hashes are in `energy_money_manifest.json`. Detailed controls,
weights, residuals and anomalies are in
`data/preprocessing/data/{audit_path.name}`.
"""
    (package / "README.md").write_text(package_readme, encoding="utf-8")
    return {
        "manifest": manifest,
        "runtime": runtime,
        "io": io,
        "marginal_eafg": marginal_eafg,
        "marginal_co2": marginal_co2,
        "package": package,
        "audit": audit_path,
    }


def style_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = fill
            cell.font = font
        for column in sheet.columns:
            letter = get_column_letter(column[0].column)
            width = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[letter].width = min(max(width + 2, 10), 60)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000"
    workbook.save(path)


def write_raw_readme(raw: Path, country: str, year: int) -> None:
    manifest = json.loads((raw / "manifest.json").read_text(encoding="utf-8"))
    coverage = manifest["eu27_coverage"]
    lines = [
        f"# Public-core raw sources — {country}, {year}",
        "",
        "Official source deliveries preserved byte-for-byte by",
        "`data/preprocessing/scripts/download_energy_money_public_core.py`.",
        f"Retrieval date: **{RETRIEVAL_DATE}**. SHA-256 hashes, exact URLs,",
        "filters, units, source status and reuse terms are in `manifest.json`.",
        "",
        "Eurostat inputs: `env_ac_pefasu`, `env_ac_ainah_r2`,",
        "`env_ac_taxind2`, `naio_10_cp15`, `naio_10_cp16`,",
        "`nrg_pc_202_c`–`205_c`, FIGARO `naio_10_fcp_s3/u3/ii3`, and",
        "`ert_bil_eur_a`. Other Commission sources are the Weekly Oil Bulletin",
        "and TAXUD VAT/excise references.",
        "",
        "EU-27 observation coverage measured by the downloader:",
    ]
    for dataset, result in sorted(coverage.items()):
        lines.append(
            f"- `{dataset}`: {result['countries_with_observations']}/27; "
            f"missing {result['missing'] or 'none'}."
        )
    lines.extend(
        [
            "",
            "The 2020 VAT reference is exact. The 2021-07-01 excise table is",
            "the nearest stable TAXUD reference and is not treated as exact 2020",
            "evidence. Raw files are controls, not direct product×user×purpose",
            "monetary observations.",
            "",
        ]
    )
    (raw / "README.md").write_text("\n".join(lines), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--country", default="SE")
    parser.add_argument("--year", type=int, default=2020)
    parser.add_argument("--force", action="store_true")
    parser.add_argument("--check-determinism", action="store_true")
    args = parser.parse_args()
    country = args.country.upper()
    output_root = DATA / "eu_core"
    audit = DATA / f"energy_money_{country.lower()}{args.year}_public_core_reconciliation.xlsx"
    raw = DATA / "eu_core_raw" / country / str(args.year)
    write_raw_readme(raw, country, args.year)
    result = build(
        country,
        args.year,
        output_root=output_root,
        audit_path=audit,
        force=args.force,
    )
    if args.check_determinism:
        with tempfile.TemporaryDirectory() as temporary:
            temp = Path(temporary)
            second = build(
                country,
                args.year,
                output_root=temp / "eu_core",
                audit_path=temp / "audit.xlsx",
                force=False,
                temporary=True,
            )
            pd.testing.assert_frame_equal(result["runtime"], second["runtime"])
            pd.testing.assert_frame_equal(result["io"], second["io"])
            pd.testing.assert_frame_equal(
                result["marginal_eafg"], second["marginal_eafg"]
            )
            pd.testing.assert_frame_equal(
                result["marginal_co2"], second["marginal_co2"]
            )
    print(json.dumps(result["manifest"]["quantitative_results"], indent=2))


if __name__ == "__main__":
    main()
