"""Pilot JRC-IDEES-2023 against GREU's Denmark 2020 purpose dimension.

The script keeps observed values separate from mapping assumptions.  IDEES is
an analytical energy-balance decomposition, whereas GREU is a residence-based
physical energy account; the workbook therefore reports both the full GREU
industry total and a closer boundary excluding own-account transport.

Run
---
python data/preprocessing/scripts/reconcile_jrc_idees_dk_2020.py
"""

from __future__ import annotations

import hashlib
import pathlib
import zipfile

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

DATA = pathlib.Path(__file__).resolve().parents[1] / "data"
RAW = DATA / "jrc_idees_2023_raw"
EXTRACTED = RAW / "extracted"
INDUSTRY = EXTRACTED / "JRC-IDEES-2023_Industry_DK.xlsx"
RESIDENTIAL = EXTRACTED / "JRC-IDEES-2023_Residential_DK.xlsx"
ARCHIVE = RAW / "JRC-IDEES-2023_DK.zip"
OUT = DATA / "jrc_idees_dk2020_purpose_reconciliation.xlsx"

YEAR = 2020
KTOE_TO_PJ = 0.041868
RETRIEVAL_DATE = "2026-07-30"
DIRECT_URL = (
    "https://jeodpp.jrc.ec.europa.eu/ftp/jrc-opendata/JRC-IDEES/"
    "JRC-IDEES-2023_v1/JRC-IDEES-2023_DK.zip"
)
CATALOGUE_URL = (
    "https://data.jrc.ec.europa.eu/dataset/"
    "1f0b480c-6d21-4d95-897d-20c7ca33df6f"
)
REPORT_URL = "https://publications.jrc.ec.europa.eu/repository/handle/JRC144707"
IDEES_PAGE = (
    "https://joint-research-centre.ec.europa.eu/scientific-tools-and-databases/"
    "potencia-policy-oriented-tool-energy-and-climate-change-impact-assessment/"
    "jrc-idees_en"
)
EUTL_URL = (
    "https://climate.ec.europa.eu/eu-action/carbon-markets/"
    "eu-emissions-trading-system-eu-ets/union-registry_en"
)

SECTORS = ["ISI", "NFM", "CHI", "NMM", "PPA", "FBT", "TRE", "MAE", "TEL", "WWP", "OIS"]
SECTOR_NAMES = {
    "ISI": "Iron and steel",
    "NFM": "Non-ferrous metals",
    "CHI": "Chemical industry",
    "NMM": "Non-metallic mineral products",
    "PPA": "Pulp, paper and printing",
    "FBT": "Food, beverages and tobacco",
    "TRE": "Transport equipment",
    "MAE": "Machinery equipment",
    "TEL": "Textiles and leather",
    "WWP": "Wood and wood products",
    "OIS": "Other industrial sectors",
}

# Connected comparison groups dictated by metadata.xlsx. The broad final group
# is necessary because several IDEES sectors map through the same GREU codes.
SECTOR_GROUPS = {
    "Food, beverages and tobacco": {
        "idees": ["FBT"],
        "greu": ["10010", "10020", "10030", "10040", "10120"],
        "mapping": "direct at NACE C10-C12 aggregate",
    },
    "Chemicals and pharmaceuticals": {
        "idees": ["CHI"],
        "greu": ["20000", "21000"],
        "mapping": "direct at NACE C20-C21 aggregate",
    },
    "Non-metallic minerals": {
        "idees": ["NMM"],
        "greu": ["23001", "23002"],
        "mapping": "direct sector; GREU has two sub-industries",
    },
    "Wood products": {
        "idees": ["WWP"],
        "greu": ["16000"],
        "mapping": "direct at NACE C16 aggregate",
    },
    "Other mapped industry cluster": {
        "idees": ["ISI", "NFM", "PPA", "TRE", "MAE", "TEL", "OIS"],
        "greu": ["0600a", "13150", "25000", "41430"],
        "mapping": (
            "constructed connected cluster: metadata combines C13-C15/C26-C32 "
            "in 13150 and C17/C18/C22/C24/C25/C33 in 25000; IDEES OIS also "
            "combines mining, construction, C22 and C31-C32"
        ),
    },
}


def sha256(path: pathlib.Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def year_column(sheet, year: int) -> int:
    for cell in sheet[1]:
        if cell.value == year:
            return cell.column
    raise ValueError(f"{year} absent from sheet {sheet.title}")


def code_column(sheet) -> int:
    for cell in sheet[1]:
        if cell.value == "Code":
            return cell.column
    raise ValueError(f"Code column absent from sheet {sheet.title}")


def coded_rows(workbook, sheet_name: str) -> pd.DataFrame:
    sheet = workbook[sheet_name]
    ycol = year_column(sheet, YEAR)
    ccol = code_column(sheet)
    rows = []
    for row in range(2, sheet.max_row + 1):
        code = sheet.cell(row, ccol).value
        if not code:
            continue
        rows.append(
            {
                "sheet": sheet_name,
                "source_row": row,
                "label": sheet.cell(row, 1).value,
                "value_ktoe": float(sheet.cell(row, ycol).value or 0),
                "source_code": str(code),
            }
        )
    return pd.DataFrame(rows)


def source_row(sheet, label: str) -> float:
    ycol = year_column(sheet, YEAR)
    matches = [
        float(sheet.cell(row, ycol).value or 0)
        for row in range(1, sheet.max_row + 1)
        if sheet.cell(row, 1).value == label
    ]
    if not matches:
        raise ValueError(f"{sheet.title}: no row labelled {label!r}")
    # Summary sheets repeat some labels later as dimensionless market shares;
    # the first occurrence is the physical level under the sheet's main unit.
    return matches[0]


def source_rows_sum(sheet, label: str) -> float:
    """Sum repeated carrier rows across non-overlapping end-use blocks."""
    ycol = year_column(sheet, YEAR)
    matches = [
        float(sheet.cell(row, ycol).value or 0)
        for row in range(1, sheet.max_row + 1)
        if sheet.cell(row, 1).value == label
    ]
    if not matches:
        raise ValueError(f"{sheet.title}: no rows labelled {label!r}")
    return sum(matches)


def compare(metric: str, greu: float, idees: float, unit: str, note: str = "") -> dict:
    difference = idees - greu if pd.notna(greu) and pd.notna(idees) else np.nan
    return {
        "metric": metric,
        "unit": unit,
        "greu_value": greu,
        "idees_value": idees,
        "difference_idees_minus_greu": difference,
        "difference_pct_of_greu": (
            difference / greu * 100 if pd.notna(difference) and abs(greu) > 1e-12 else np.nan
        ),
        "note": note,
    }


def style_workbook(path: pathlib.Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column in sheet.columns:
            letter = get_column_letter(column[0].column)
            max_length = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 60)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.0000"
    workbook.save(path)


def main() -> None:
    required = [ARCHIVE, INDUSTRY, RESIDENTIAL, DATA / "energy_and_emissions.xlsx"]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError(f"missing inputs; run downloader first: {missing}")
    with zipfile.ZipFile(ARCHIVE) as archive:
        assert archive.testzip() is None

    industry_wb = load_workbook(INDUSTRY, read_only=True, data_only=True)
    residential_wb = load_workbook(RESIDENTIAL, read_only=True, data_only=True)
    industry_rows = pd.concat(
        [coded_rows(industry_wb, f"{sector}_fec") for sector in SECTORS],
        ignore_index=True,
    )
    parts = industry_rows["source_code"].str.split(".")
    industry_rows["measure"] = parts.str[0]
    industry_rows["unit"] = parts.str[1]
    industry_rows["country"] = parts.str[2]
    industry_rows["idees_sector"] = parts.str[3]
    industry_rows["subsector"] = parts.str[4]
    industry_rows["process"] = parts.str[5]
    industry_rows["end_use_technology"] = parts.str[6]
    industry_rows["fuel"] = parts.str[7]
    industry_rows["value_pj"] = industry_rows["value_ktoe"] * KTOE_TO_PJ

    fec = industry_rows[industry_rows["measure"] == "FEC"].copy()
    assert set(fec["country"]) == {"DK"} and set(fec["unit"]) == {"ktoe"}

    sector_totals = (
        fec[
            (fec["process"] == "TOTAL")
            & (fec["end_use_technology"] == "TOTAL")
            & (fec["fuel"] == "TOTAL")
        ]
        .groupby("idees_sector", as_index=True)["value_ktoe"]
        .sum()
        .reindex(SECTORS, fill_value=0)
    )

    summary = industry_wb["Ind_Summary_fec"]
    summary_total = source_row(summary, "All industrial sectors")
    summary_buckets = {
        "Lighting": source_row(summary, "Lighting"),
        "Air compressors": source_row(summary, "Air compressors"),
        "Motor drives": source_row(summary, "Motor drives"),
        "Fans and pumps": source_row(summary, "Fans and pumps"),
        "Low-enthalpy heat": source_row(summary, "Low-enthalpy heat"),
        "Steam processes": source_row(summary, "Steam processes"),
        "Other processes": source_row(summary, "Other processes"),
    }
    assert abs(sum(summary_buckets.values()) - summary_total) < 1e-7
    assert abs(sector_totals.sum() - summary_total) < 1e-7

    # Candidate mapping. Heating is the closest source concept. "Special" is
    # explicitly an assumption: process energy in metallurgical/mineralogical
    # sectors, excluding cross-cutting non-process uses. Normal is the residual.
    low_enthalpy = fec[
        (fec["process"] == "LOW_ENTH")
        & (fec["end_use_technology"] == "TOTAL")
        & (fec["fuel"] == "TOTAL")
    ].groupby("idees_sector")["value_ktoe"].sum()
    cross_cutting = fec[
        fec["process"].isin(["LIGHT", "AIRCOMP", "MOTOR", "FANS"])
    ].groupby("idees_sector")["value_ktoe"].sum()
    special_sectors = ["ISI", "NFM", "NMM"]
    special_proxy = sum(
        sector_totals.get(sector, 0)
        - low_enthalpy.get(sector, 0)
        - cross_cutting.get(sector, 0)
        for sector in special_sectors
    )
    heating_proxy = low_enthalpy.sum()
    normal_proxy = summary_total - heating_proxy - special_proxy
    assert abs(heating_proxy + normal_proxy + special_proxy - summary_total) < 1e-7

    energy = pd.read_excel(DATA / "energy_and_emissions.xlsx", sheet_name="ems_energy")
    energy["indu"] = energy["indu"].astype(str)
    use = energy[
        (energy["year"] == YEAR)
        & (energy["bal"] == "use")
        & (energy["flow"] == "cons_inter")
    ].copy()
    mapped_industries = sorted(
        {industry for group in SECTOR_GROUPS.values() for industry in group["greu"]}
    )
    greu_industry = use[use["indu"].isin(mapped_industries)].copy()
    greu_purpose = greu_industry.groupby("purp")["pj"].sum()
    greu_all = greu_industry["pj"].sum()
    greu_without_transport = greu_all - greu_purpose.get("transport", 0)
    idees_total_pj = summary_total * KTOE_TO_PJ

    totals = pd.DataFrame(
        [
            compare(
                "All mapped GREU industrial use vs IDEES industry FEC",
                greu_all,
                idees_total_pj,
                "PJ",
                "Not like-for-like: GREU includes own-account transport by resident industries.",
            ),
            compare(
                "Closer boundary: GREU excluding transport purpose vs IDEES industry FEC",
                greu_without_transport,
                idees_total_pj,
                "PJ",
                (
                    "Still not exact: GREU is residence-based and has finer industries; IDEES follows "
                    "territorial Eurostat energy-balance industry sectors and includes estimated end uses."
                ),
            ),
        ]
    )

    purpose_idees = {
        "heating": heating_proxy * KTOE_TO_PJ,
        "process_normal": normal_proxy * KTOE_TO_PJ,
        "process_special": special_proxy * KTOE_TO_PJ,
        "transport": np.nan,
        "in_ETS": np.nan,
    }
    purpose_rows = []
    purpose_notes = {
        "heating": (
            "Direct conceptual proxy: IDEES non-process low-enthalpy heat. It is narrower "
            "than GREU room/water heating and is an estimated end-use decomposition."
        ),
        "process_normal": (
            "Constructed residual after heating and the special-process proxy; includes "
            "cross-cutting uses and all process energy outside ISI/NFM/NMM."
        ),
        "process_special": (
            "Assumption-based: all process energy in iron/steel, non-ferrous metals and "
            "non-metallic minerals, excluding cross-cutting uses. GREU's definition is narrower."
        ),
        "transport": (
            "Not identified by the IDEES industry workbook. IDEES transport is mode-based "
            "and cannot allocate own-account transport back to user industries."
        ),
        "in_ETS": (
            "Not present in IDEES. Installation-level EUTL/Union Registry data and an "
            "installation-to-industry bridge are required."
        ),
    }
    mapping_types = {
        "heating": "direct conceptual proxy",
        "process_normal": "constructed / assumption",
        "process_special": "constructed / assumption",
        "transport": "not identifiable",
        "in_ETS": "not identifiable",
    }
    for purpose in ["heating", "transport", "process_normal", "process_special", "in_ETS"]:
        row = compare(
            purpose,
            greu_purpose.get(purpose, 0),
            purpose_idees[purpose],
            "PJ",
            purpose_notes[purpose],
        )
        row["mapping_type"] = mapping_types[purpose]
        purpose_rows.append(row)
    process_envelope_greu = sum(
        greu_purpose.get(purpose, 0)
        for purpose in ["process_normal", "process_special", "in_ETS"]
    )
    process_envelope_idees = purpose_idees["process_normal"] + purpose_idees["process_special"]
    envelope = compare(
        "combined process envelope (normal + special + in_ETS)",
        process_envelope_greu,
        process_envelope_idees,
        "PJ",
        (
            "Diagnostic only. Combining categories avoids pretending IDEES identifies ETS "
            "status or GREU's normal/special boundary."
        ),
    )
    envelope["mapping_type"] = "robust aggregate diagnostic"
    purpose_rows.append(envelope)
    purpose_comparison = pd.DataFrame(purpose_rows)

    source_enduses = pd.DataFrame(
        [
            {
                "source_end_use": name,
                "value_ktoe": value,
                "value_pj": value * KTOE_TO_PJ,
                "source_or_assumption": "source value",
            }
            for name, value in summary_buckets.items()
        ]
    )

    sector_rows = []
    for group_name, group in SECTOR_GROUPS.items():
        idees_pj = sector_totals.reindex(group["idees"], fill_value=0).sum() * KTOE_TO_PJ
        greu_pj = greu_industry[greu_industry["indu"].isin(group["greu"])]["pj"].sum()
        row = compare(group_name, greu_pj, idees_pj, "PJ", group["mapping"])
        row["idees_sectors"] = ", ".join(group["idees"])
        row["greu_industries"] = ", ".join(group["greu"])
        sector_rows.append(row)
    sector_comparison = pd.DataFrame(sector_rows)

    # Reconstruct carrier totals from the non-overlapping summary buckets.
    carrier_rows = {
        "coal_and_coke": source_rows_sum(summary, "Solids") + source_rows_sum(summary, "Coke"),
        "refinery_gas": source_rows_sum(summary, "Refinery gas"),
        "oil_products_including_bio_diesel": (
            source_rows_sum(summary, "LPG")
            + source_rows_sum(summary, "Diesel oil")
            + source_rows_sum(summary, "Fuel oil")
            + source_rows_sum(summary, "Other liquids")
        ),
        "natural_gas_and_biogas": source_rows_sum(summary, "Natural gas"),
        "derived_gases": source_rows_sum(summary, "Derived gases"),
        "biomass_and_waste": source_rows_sum(summary, "Biomass and waste"),
        "distributed_heat": source_rows_sum(summary, "Distributed steam"),
        "solar_and_geothermal": source_rows_sum(summary, "Solar and geothermal"),
        "ambient_heat": source_rows_sum(summary, "Ambient heat"),
        "electricity": (
            summary_buckets["Lighting"]
            + summary_buckets["Air compressors"]
            + summary_buckets["Motor drives"]
            + summary_buckets["Fans and pumps"]
            + source_rows_sum(summary, "Electricity")
        ),
    }
    assert abs(sum(carrier_rows.values()) - summary_total) < 1e-7
    fuel_product_map = {
        "coal_and_coke": ["coal"],
        "refinery_gas": ["refin_gas"],
        "oil_products_including_bio_diesel": [
            "diesel_transp",
            "liq_biofuel",
            "other_oil",
            "sem_refin_oil",
            "waste_oil",
        ],
        # The data workbook uses these two operational gas labels even though
        # metadata.xlsx's product list has the shorter label "natgas".
        "natural_gas_and_biogas": [
            "natgas_incl_biongas",
            "natgas_extraction",
            "biogas",
        ],
        "derived_gases": [],
        "biomass_and_waste": ["firewood", "straw", "waste", "wood_pellets", "wood_waste"],
        "distributed_heat": ["district_heat"],
        "solar_and_geothermal": ["renewable"],
        "ambient_heat": ["heat_pump"],
        "electricity": ["electricity"],
    }
    greu_fuel_boundary = greu_industry[greu_industry["purp"] != "transport"]
    fuel_rows = []
    mapped_greu_fuel = 0.0
    for carrier, products in fuel_product_map.items():
        greu_value = greu_fuel_boundary[greu_fuel_boundary["product"].isin(products)]["pj"].sum()
        mapped_greu_fuel += greu_value
        row = compare(
            carrier,
            greu_value,
            carrier_rows[carrier] * KTOE_TO_PJ,
            "PJ",
            (
                "Broad carrier bridge; IDEES fuel groups combine products differently from GREU."
                if products
                else "No direct GREU product counterpart assigned."
            ),
        )
        row["greu_products"] = ", ".join(products) if products else "none"
        fuel_rows.append(row)
    unallocated_greu = greu_without_transport - mapped_greu_fuel
    fuel_rows.append(
        compare(
            "GREU products not allocated by broad bridge",
            unallocated_greu,
            0.0,
            "PJ",
            "Diagnostic residual; no claim of source equivalence.",
        )
    )
    fuel_comparison = pd.DataFrame(fuel_rows)

    # Household context: IDEES residential excludes household transport.  PEFA
    # already provides total/heat/transport/other on the physical-account basis.
    res = residential_wb["RES_summary"]
    idees_res_total = source_row(res, "Energy consumption by fuel - Eurostat structure (ktoe)")
    idees_hh_heating = (
        source_row(res, "Space heating (without ambient heat)")
        + source_row(res, "Water heating")
        + source_row(res, "Ambient heat")
    )
    idees_hh_other = (
        source_row(res, "Cooling")
        + source_row(res, "Cooking")
        + source_row(res, "Specific electricity uses (appliances and lighting)")
    )
    assert abs(idees_hh_heating + idees_hh_other - idees_res_total) < 1e-7
    hh_greu = energy[
        (energy["year"] == YEAR)
        & (energy["bal"] == "use")
        & (energy["flow"] == "cons_hh")
    ].groupby("purp")["pj"].sum()
    pefa_path = DATA / "eurostat_energy_emissions_dk2020_reconciliation.xlsx"
    pefa_hh = pd.read_excel(pefa_path, sheet_name="hh_energy_purpose")
    pefa_by_metric = dict(zip(pefa_hh["metric"], pefa_hh["eurostat"]))
    pefa_heat = pefa_by_metric["heating / HH_HEAT"]
    pefa_transport = pefa_by_metric["transport / HH_TRA"]
    pefa_other = pefa_by_metric["appliances / HH_OTH"]
    pefa_total = pefa_by_metric["Household total / HH"]
    hh_rows = [
        {
            "component": "heating",
            "greu_pj": hh_greu.get("heating", 0),
            "pefa_pj": pefa_heat,
            "idees_pj": idees_hh_heating * KTOE_TO_PJ,
            "recommended_role": (
                "PEFA supplies the account total; IDEES can refine residential "
                "space/water-heating detail."
            ),
        },
        {
            "component": "transport",
            "greu_pj": hh_greu.get("transport", 0),
            "pefa_pj": pefa_transport,
            "idees_pj": np.nan,
            "recommended_role": (
                "Use PEFA HH_TRA. IDEES transport is mode-based and is outside "
                "the Residential workbook."
            ),
        },
        {
            "component": "other / appliances",
            "greu_pj": hh_greu.get("appliances", 0),
            "pefa_pj": pefa_other,
            "idees_pj": idees_hh_other * KTOE_TO_PJ,
            "recommended_role": (
                "PEFA supplies HH_OTH; IDEES separates cooking, cooling and "
                "specific electricity uses."
            ),
        },
        {
            "component": "residential subtotal excluding transport",
            "greu_pj": hh_greu.get("heating", 0) + hh_greu.get("appliances", 0),
            "pefa_pj": pefa_heat + pefa_other,
            "idees_pj": idees_res_total * KTOE_TO_PJ,
            "recommended_role": "Like-for-like residential boundary.",
        },
        {
            "component": "combined IDEES residential + PEFA transport",
            "greu_pj": hh_greu.sum(),
            "pefa_pj": pefa_total,
            "idees_pj": idees_res_total * KTOE_TO_PJ + pefa_transport,
            "recommended_role": (
                "Hybrid diagnostic only; retain PEFA as the control total to avoid "
                "mixing accounting boundaries silently."
            ),
        },
    ]
    household_context = pd.DataFrame(hh_rows)
    household_context["idees_minus_greu_pj"] = (
        household_context["idees_pj"] - household_context["greu_pj"]
    )
    household_context["idees_minus_greu_pct"] = (
        household_context["idees_minus_greu_pj"] / household_context["greu_pj"] * 100
    )

    purpose_mapping = pd.DataFrame(
        [
            {
                "greu_purpose": "heating",
                "idees_evidence": "Industry LOW_ENTH (non-process low-enthalpy heat)",
                "verdict": "direct conceptual proxy, not exact",
                "construction_needed": "Map LOW_ENTH; validate sector-by-sector.",
                "remaining_gap": "IDEES proxy is much narrower for DK 2020.",
            },
            {
                "greu_purpose": "transport",
                "idees_evidence": "Separate mode-based Transport workbook",
                "verdict": "cannot identify by user industry",
                "construction_needed": "Use PEFA/product evidence or a separate allocation key.",
                "remaining_gap": "No own-account transport allocation to GREU industries.",
            },
            {
                "greu_purpose": "process_normal",
                "idees_evidence": "Cross-cutting and process-specific end uses",
                "verdict": "assumption-based construction",
                "construction_needed": "Classify process codes and use a residual rule.",
                "remaining_gap": "GREU normal/special boundary is not an IDEES dimension.",
            },
            {
                "greu_purpose": "process_special",
                "idees_evidence": "Named metallurgical/mineralogical/process end uses",
                "verdict": "assumption-based construction",
                "construction_needed": "Owner-approved process-code concordance.",
                "remaining_gap": "Some IDEES end uses combine heating and non-heating applications.",
            },
            {
                "greu_purpose": "in_ETS",
                "idees_evidence": "No ETS-status field in downloaded IDEES files or documentation",
                "verdict": "not identifiable",
                "construction_needed": "Public EUTL installation data + installation-to-NACE bridge.",
                "remaining_gap": "Allocate installation coverage/fuels to GREU industries without double counting.",
            },
        ]
    )

    access_coverage = pd.DataFrame(
        [
            {
                "topic": "edition/vintage",
                "verified_fact": "JRC-IDEES-2023 v1; dataset issued 2025-11-17, documentation 2026",
                "source_url": CATALOGUE_URL,
            },
            {
                "topic": "access",
                "verified_fact": "Static ZIP; direct anonymous access; no login or query parameters",
                "source_url": DIRECT_URL,
            },
            {
                "topic": "licence",
                "verified_fact": "Creative Commons Attribution 4.0 International (CC BY 4.0)",
                "source_url": CATALOGUE_URL,
            },
            {
                "topic": "countries",
                "verified_fact": "All EU-27 Member States plus EU27 aggregate; one archive per country",
                "source_url": (
                    "https://jeodpp.jrc.ec.europa.eu/ftp/jrc-opendata/"
                    "JRC-IDEES/JRC-IDEES-2023_v1/"
                ),
            },
            {
                "topic": "years",
                "verified_fact": "Annual 2000-2023; Denmark 2020 present",
                "source_url": REPORT_URL,
            },
            {
                "topic": "industry classification",
                "verified_fact": (
                    "11 sectors, 21 subsectors, mapped to NACE Rev. 2; "
                    "6-11 processes per subsector"
                ),
                "source_url": REPORT_URL,
            },
            {
                "topic": "energy carriers",
                "verified_fact": (
                    "Solids/coke, refinery gas, LPG, diesel+liquid biofuels, fuel oil, "
                    "other liquids, gas+biogas, derived gases, biomass+waste, distributed "
                    "steam, solar/geothermal, ambient heat, electricity"
                ),
                "source_url": REPORT_URL,
            },
            {
                "topic": "end uses",
                "verified_fact": (
                    "Process-specific uses plus cross-cutting lighting, air compressors, "
                    "motor drives, fans/pumps and low-enthalpy heat; values are modelled "
                    "decompositions constrained to Eurostat energy balances"
                ),
                "source_url": REPORT_URL,
            },
            {
                "topic": "ETS",
                "verified_fact": (
                    "IDEES does not expose installation ETS status; EUTL publicly exposes "
                    "free allocation and verified emissions"
                ),
                "source_url": EUTL_URL,
            },
        ]
    )

    detail_loss = pd.DataFrame(
        [
            {
                "dimension": "accounting boundary",
                "idees": "territorial Eurostat energy-balance sectors",
                "greu": "SEEA residence-based energy account by producing/user industry",
                "diagnostic": (
                    f"IDEES {idees_total_pj:.3f} PJ vs GREU mapped industries "
                    f"{greu_all:.3f} PJ; vs GREU excluding transport {greu_without_transport:.3f} PJ."
                ),
            },
            {
                "dimension": "industry",
                "idees": "11 sectors / 21 subsectors; OIS is an aggregate",
                "greu": "57 industries with several splits below NACE A64",
                "diagnostic": "Five connected groups are safe for numeric comparison; one is very broad.",
            },
            {
                "dimension": "purpose",
                "idees": "named technical processes and cross-cutting end uses",
                "greu": "five mutually exclusive industrial purposes",
                "diagnostic": "Only heating has a close source concept; normal/special require rules.",
            },
            {
                "dimension": "ETS",
                "idees": "not identified",
                "greu": "mutually exclusive in_ETS purpose",
                "diagnostic": f"GREU mapped industries contain {greu_purpose.get('in_ETS', 0):.3f} PJ in_ETS.",
            },
            {
                "dimension": "transport",
                "idees": "separate mode-based transport sector",
                "greu": "transport purpose assigned to user industries",
                "diagnostic": (
                    f"GREU mapped industries contain {greu_purpose.get('transport', 0):.3f} PJ "
                    "that IDEES Industry cannot allocate."
                ),
            },
        ]
    )

    assumptions = pd.DataFrame(
        [
            {
                "id": "A1",
                "assumption": (
                    "LOW_ENTH is used as the closest proxy for GREU heating because JRC describes "
                    "it as a non-process cross-cutting use."
                ),
                "effect": "Creates a comparable number but not an exact semantic match.",
            },
            {
                "id": "A2",
                "assumption": (
                    "Special-process proxy equals process-related energy in ISI, NFM and NMM, "
                    "excluding LOW_ENTH and four electrical cross-cutting uses."
                ),
                "effect": "Over-broad relative to GREU's narrower special-process definition.",
            },
            {
                "id": "A3",
                "assumption": "Normal process is the residual after A1 and A2.",
                "effect": "Contains ETS-covered uses because IDEES has no ETS status.",
            },
            {
                "id": "A4",
                "assumption": "GREU transport purpose is excluded for the closer industry total.",
                "effect": "Reduces but does not eliminate residence/territory and classification differences.",
            },
            {
                "id": "A5",
                "assumption": "ktoe converted at exactly 0.041868 PJ per ktoe.",
                "effect": "Standard energy-unit conversion.",
            },
        ]
    )

    validation = pd.DataFrame(
        [
            {"check": "archive ZIP CRC", "value": "pass", "tolerance": "all members"},
            {"check": "archive SHA-256", "value": sha256(ARCHIVE), "tolerance": "recorded"},
            {"check": "comparison year present", "value": YEAR, "tolerance": "exact"},
            {
                "check": "industry sector totals equal IDEES summary",
                "value": sector_totals.sum() - summary_total,
                "tolerance": "<1e-7 ktoe",
            },
            {
                "check": "source end-use buckets equal IDEES summary",
                "value": sum(summary_buckets.values()) - summary_total,
                "tolerance": "<1e-7 ktoe",
            },
            {
                "check": "candidate purpose buckets equal IDEES summary",
                "value": heating_proxy + normal_proxy + special_proxy - summary_total,
                "tolerance": "<1e-7 ktoe",
            },
            {
                "check": "IDEES carriers equal industry summary",
                "value": sum(carrier_rows.values()) - summary_total,
                "tolerance": "<1e-7 ktoe",
            },
            {
                "check": "IDEES household end uses equal residential fuel total",
                "value": idees_hh_heating + idees_hh_other - idees_res_total,
                "tolerance": "<1e-7 ktoe",
            },
            {
                "check": "GREU mapped industry purpose sum",
                "value": greu_purpose.sum() - greu_all,
                "tolerance": "<1e-9 PJ",
            },
        ]
    )

    metadata = pd.DataFrame(
        [
            {"key": "created", "value": RETRIEVAL_DATE},
            {"key": "comparison_year", "value": YEAR},
            {"key": "country", "value": "Denmark (DK)"},
            {"key": "source_edition", "value": "JRC-IDEES-2023 v1"},
            {"key": "source_unit", "value": "ktoe; converted to PJ at 0.041868"},
            {"key": "raw_archive", "value": str(ARCHIVE.relative_to(DATA.parent.parent))},
            {"key": "raw_archive_sha256", "value": sha256(ARCHIVE)},
            {"key": "direct_download", "value": DIRECT_URL},
            {"key": "download_parameters", "value": "none; static DK country ZIP"},
            {"key": "catalogue", "value": CATALOGUE_URL},
            {"key": "technical_report", "value": REPORT_URL},
            {"key": "licence", "value": "CC BY 4.0"},
            {
                "key": "interpretation_warning",
                "value": (
                    "IDEES end uses below sector/fuel totals are analytical estimates, not "
                    "separately observed statistics; GREU and IDEES accounting boundaries differ."
                ),
            },
        ]
    )

    raw_sector_totals = pd.DataFrame(
        [
            {
                "idees_sector": sector,
                "sector_name": SECTOR_NAMES[sector],
                "value_ktoe": sector_totals[sector],
                "value_pj": sector_totals[sector] * KTOE_TO_PJ,
                "source_or_assumption": "source value",
            }
            for sector in SECTORS
        ]
    )

    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        metadata.to_excel(writer, sheet_name="metadata", index=False)
        access_coverage.to_excel(writer, sheet_name="access_coverage", index=False)
        purpose_mapping.to_excel(writer, sheet_name="purpose_mapping", index=False)
        totals.to_excel(writer, sheet_name="totals_reconciliation", index=False)
        purpose_comparison.to_excel(writer, sheet_name="purpose_comparison", index=False)
        source_enduses.to_excel(writer, sheet_name="idees_enduses", index=False)
        raw_sector_totals.to_excel(writer, sheet_name="idees_sector_totals", index=False)
        sector_comparison.to_excel(writer, sheet_name="sector_comparison", index=False)
        fuel_comparison.to_excel(writer, sheet_name="fuel_comparison", index=False)
        household_context.to_excel(writer, sheet_name="household_context", index=False)
        detail_loss.to_excel(writer, sheet_name="detail_loss", index=False)
        assumptions.to_excel(writer, sheet_name="assumptions", index=False)
        validation.to_excel(writer, sheet_name="validation", index=False)
        industry_rows.to_excel(writer, sheet_name="source_rows_2020", index=False)
    style_workbook(OUT)
    print(f"wrote {OUT}")
    print(totals.to_string(index=False))
    print(purpose_comparison.to_string(index=False))
    print(household_context.to_string(index=False))


if __name__ == "__main__":
    main()
