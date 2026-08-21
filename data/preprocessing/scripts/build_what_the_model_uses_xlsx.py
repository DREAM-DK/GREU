"""Build docs/What_the_model_uses.xlsx and .md — what GREU uses, and where it lives.

Three sheets / sections: a half-page Read me; an everyday overview; then
one row per data-fed model variable, grouped by submodule. Only variables
that ``@load`` from data.gdx are listed. Endogenous prices, j-terms and
behavioural parameters are out. The Markdown file is the same content
for an agent; do not hand-edit either file.

Re-run after the list of load-bearing inputs changes:

    python data/preprocessing/scripts/build_what_the_model_uses_xlsx.py
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

REPO_ROOT = Path(__file__).resolve().parents[3]
OUTPUT_PATH = REPO_ROOT / "docs" / "What_the_model_uses.xlsx"
MD_OUTPUT_PATH = REPO_ROOT / "docs" / "What_the_model_uses.md"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
AREA_FILLS = {
    "Energy": "D6EAF8",
    "Economy": "D5F5E3",
    "Labour": "FCF3CF",
    "Capital": "FADBD8",
    "Climate": "E8DAEF",
    "Government": "F5CBA7",
    "Note": "D5D8DC",
}
THIN_BORDER = Border(*[Side(style="thin", color="BFBFBF")] * 4)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")

HEADERS = [
    "Area",
    "What the number is",
    "Where it comes from today",
    "What the model does with it",
    "Unit",
    "EU equivalent",
    "Name in the model",
]

# Managers read columns A–F. The last column is for a modeller in the room.
ROWS = [
    (
        "Energy",
        "How much energy is used — fuel, electricity, heat — by who uses it and for what",
        "Danish energy accounts",
        "This is the physical energy the model runs on",
        "Petajoules (PJ)",
        "Eurostat physical energy accounts, in terajoules (divide by 1,000)",
        "qEpj",
    ),
    (
        "Energy",
        "The price of that energy before tax and shop mark-ups",
        "Danish energy accounts (producer prices)",
        "Turns physical energy into a money amount",
        "Money per PJ",
        "Not published at this detail; built from broader money totals",
        "pEpj_base",
    ),
    (
        "Energy",
        "Wholesale, retail and garage mark-ups on energy",
        "Danish energy accounts",
        "The gap between the producer price and what the buyer pays",
        "Money",
        "Not published separately; built from combined trade-margin totals",
        "vE_margins",
    ),
    (
        "Energy",
        "Energy taxes and VAT",
        "Danish energy accounts",
        "What government collects on energy use",
        "Money",
        "Not published per fuel and user; built from tax totals and rates",
        "vtE_duty, vtE_vat",
    ),
    (
        "Energy",
        "Emissions from burning fuels",
        "Danish energy accounts",
        "Links each fuel use to greenhouse gases",
        "Thousand tonnes",
        "Eurostat air-emissions accounts (energy and process combined)",
        "qEmmE_BU",
    ),
    (
        "Energy",
        "How the energy is used: heating, transport, or industrial process",
        "Danish energy accounts (a purpose tag on each row)",
        "Lets the model treat a factory boiler differently from a truck",
        "Same PJ, split by use",
        "Not published in these categories; optional split of known "
        "purchases (heating vs process vs transport). Different from qI_k_i "
        "(who invests in which capital type)",
        "es (purpose dimension of qEpj)",
    ),
    (
        "Energy",
        "Energy used in plants covered by the EU emissions trading system",
        "Danish energy accounts (an ETS tag on some rows)",
        "Separates regulated from unregulated energy use",
        "Petajoules (PJ)",
        "The ETS register shows who is in the system, not how much energy they used — still a gap",
        "qEpj with purpose in_ETS",
    ),
    (
        "Economy",
        "Who sells what to whom — firms, households, government, investment, exports",
        "The Danish input-output table",
        "The money skeleton of the economy",
        "Money",
        "FIGARO money tables (also Eurostat — money, not energy)",
        "vY_i_d",
    ),
    (
        "Economy",
        "What is imported, and who buys the imports",
        "The Danish input-output table (import rows)",
        "Separates domestic production from foreign supply",
        "Money",
        "FIGARO money tables",
        "vM_i_d",
    ),
    (
        "Economy",
        "What households buy, in 12 spending groups (food, housing, cars, …)",
        "The household columns of the input-output table",
        "How household demand is split in the model",
        "Money",
        "Eurostat household consumption by purpose, at coarser detail",
        "qD for household groups",
    ),
    (
        "Labour",
        "Wages, and hours worked by employees and the self-employed",
        "The input-output table plus the employment file",
        "Labour cost by industry; hours are used to include the self-employed",
        "Money (wages); thousand hours",
        "Eurostat employment by industry (hours and persons)",
        "qL, vWages_i",
    ),
    (
        "Labour",
        "How many people are in work, nationwide",
        "The employment file",
        "A single national headcount",
        "People",
        "Eurostat employment (persons)",
        "nL",
    ),
    (
        "Capital",
        "The stock of buildings, vehicles and machinery by industry",
        "Danish capital-stock accounts (seven asset types, collapsed to three)",
        "What each industry already owns",
        "Money",
        "Eurostat net capital stocks",
        "qK_k_i",
    ),
    (
        "Capital",
        "New investment in buildings, vehicles and machinery",
        "The investment part of the input-output table",
        "What each industry is adding this year",
        "Money",
        "Eurostat investment by asset type; where industry detail is only "
        "A21, construct with a Danish-prior balance (capital types, not "
        "heating/process/transport purposes)",
        "qI_k_i",
    ),
    (
        "Climate",
        "Emissions that do not come from burning fuel (industrial processes, F-gases)",
        "Danish process-emissions file",
        "The non-energy part of the emissions total",
        "Thousand tonnes",
        "Eurostat air-emissions accounts, minus the energy part",
        "qEmmxE",
    ),
    (
        "Climate",
        "Land-use emissions, and a bridge from “residents” to “on the territory”",
        "Danish emissions-bridge file",
        "Aligns the model with inventory reporting (forests, bunkers, border trade)",
        "Thousand tonnes",
        "Eurostat bridging items and the UNFCCC inventory",
        "qEmmLULUCF, qEmmBorderTrade",
    ),
    (
        "Climate",
        "Free allowances under the EU emissions trading system",
        "Danish ETS file",
        "How many permits industries received for free",
        "Thousand tonnes",
        "EU ETS register (who is in; industry labels need a bridge)",
        "qCO2_ETS_freeallowances",
    ),
    (
        "Government",
        "Taxes, spending and transfers — VAT, income tax, public consumption, subsidies",
        "Danish government-finance accounts",
        "The public budget the model is calibrated to",
        "Money",
        "Eurostat government accounts",
        "vtVAT, vG, vtCorp, and similar",
    ),
    (
        "Government",
        "Financial positions of households, firms, government and the rest of the world",
        "Already pulled live from Eurostat — the Danish spreadsheet is not used",
        "Who owes whom, in broad instrument groups",
        "Money",
        "Eurostat financial accounts (already in the model)",
        "vNetFinAssets, vNetDebtInstruments",
    ),
    (
        "Note",
        "The final shelf price of energy (what the buyer actually pays)",
        "Not a separate input — added up from producer price + mark-ups + taxes + VAT",
        "A check that the pieces fit; the model rebuilds it rather than reading it",
        "Money",
        "Same construction once the pieces exist",
        "not read (purch is stored, then ignored)",
    ),
]

WIDTHS = [12, 42, 36, 40, 18, 42, 28]

# ---------------------------------------------------------------------------
# Sheet 3 — By model module
# One row per variable that a module @loads from data.gdx. Modules that
# take no data of their own get a single grey row so the list is complete.
# Order follows model/base_model.gms.
# ---------------------------------------------------------------------------

NO_DATA = "No data of its own — uses variables loaded by other modules."

MODULE_HEADERS = [
    "Module",
    "Variable",
    "What it is",
    "Where to find it",
    "Considerations",
]

MODULE_FILLS = {
    "input_output": "D5F5E3",
    "labor_market": "FCF3CF",
    "factor_demand": "FADBD8",
    "factor_demand_energy": "F5CBA7",
    "households": "D5F5E3",
    "financial_accounts": "F5CBA7",
    "government": "F5CBA7",
    "exports": "D5F5E3",
    "production": "D6EAF8",
    "emissions": "E8DAEF",
    "energy_markets": "D6EAF8",
    "energy_and_emissions_taxes": "E8DAEF",
    "production_CET": "D6EAF8",
    "energy_technology": "D6EAF8",
}
NO_DATA_FILL = "D5D8DC"

# Where to find it: plain source + table code, or blank if not in Eurostat /
# FIGARO / the ETS register. "Only as a total" still fills the cell.
# Considerations: same object vs only a total vs how we construct it.
MODULE_ROWS = [
    # input_output
    (
        "input_output",
        "vY_i_d",
        "Domestic output by industry and who buys it (purchaser prices)",
        "FIGARO input-output table (naio_10_fcp_ii3)",
        "Same who-sells-to-whom, but fewer industries, and FIGARO is at "
        "producer prices not shelf prices.",
    ),
    (
        "input_output",
        "vY_i_d_base",
        "Same domestic deliveries, before product taxes",
        "FIGARO input-output table (naio_10_fcp_ii3)",
        "Closer match — FIGARO is already at producer prices.",
    ),
    (
        "input_output",
        "vtY_i_d",
        "Net product taxes on domestic output, by industry and buyer",
        "FIGARO input-output table (naio_10_fcp_ii3), product-tax row D21X31",
        "Only as a total: one combined product-tax row, not a tax on each "
        "sale, and not the Danish named taxes.",
    ),
    (
        "input_output",
        "vtY_i_Sub",
        "Production subsidies by industry",
        "FIGARO input-output table (naio_10_fcp_ii3), row D29X39",
        "Same idea as a production-subsidy total by industry; FIGARO "
        "publishes it netted with production taxes.",
    ),
    (
        "input_output",
        "vtY_i_Tax",
        "Production taxes by industry",
        "FIGARO input-output table (naio_10_fcp_ii3), row D29X39",
        "Same idea as a production-tax total by industry; FIGARO publishes "
        "it netted with subsidies.",
    ),
    (
        "input_output",
        "vM_i_d",
        "Imports by foreign industry and who buys them (purchaser prices)",
        "FIGARO input-output table (naio_10_fcp_ii3)",
        "Same who-buys-imports, but FIGARO treats re-exports differently "
        "and is at producer prices.",
    ),
    (
        "input_output",
        "vM_i_d_base",
        "Same imports, before product taxes",
        "FIGARO input-output table (naio_10_fcp_ii3)",
        "Closer match — producer prices.",
    ),
    (
        "input_output",
        "vtM_i_d",
        "Net product taxes on imports, by industry and buyer",
        "FIGARO input-output table (naio_10_fcp_ii3), product-tax row D21X31",
        "Only as a total: the same combined product-tax row covers imports "
        "too, not a tax on each import cell.",
    ),
    # labor_market
    (
        "labor_market",
        "nL",
        "Total employment",
        "Eurostat employment by industry, persons (nama_10_a64_e)",
        "Same object as a national headcount; industry detail is coarser in "
        "some countries.",
    ),
    (
        "labor_market",
        "vWages_i",
        "Compensation of employees by industry",
        "FIGARO / national accounts wages (naio_10_fcp_ii3, D1); hours from "
        "nama_10_a64_e",
        "Same object (compensation of employees by industry). Hours are used "
        "only to include the self-employed.",
    ),
    (
        "labor_market",
        "vW",
        "Compensation per employee",
        "",
        "Not a published Eurostat series. We divide total wages by "
        "employment (naio_10_fcp_ii3 D1 / nama_10_a64_e).",
    ),
    # factor_demand
    (
        "factor_demand",
        "qK_k_i",
        "Capital stock by type (buildings, vehicles, machinery) and industry",
        "Eurostat net capital stocks (nama_10_nfa_st)",
        "Same object at the totals the model uses; seven Danish asset types "
        "are collapsed to three.",
    ),
    (
        "factor_demand",
        "qI_k_i",
        "New investment by type and industry",
        "Eurostat investment by asset and industry (nama_10_a64_p5); FIGARO "
        "P51G for who supplies the goods",
        "Amounts by type × industry (Split B: types of the single K column). "
        "National totals match. Where Eurostat is only A21, construct with a "
        "Danish-prior balance. Not a supplier×investor matrix. Not the "
        "heating/process/transport purpose split.",
    ),
    (
        "factor_demand",
        "qD[i]",
        "Non-energy intermediate demand, by supplying industry",
        "FIGARO input-output table (naio_10_fcp_ii3)",
        "Same object: intermediate-use columns.",
    ),
    (
        "factor_demand",
        "qD[k]",
        "Investment demand, by capital type",
        "FIGARO input-output table (naio_10_fcp_ii3), investment column P51G",
        "Only as one investment column — we split it into buildings / "
        "vehicles / other.",
    ),
    (
        "factor_demand",
        "qD[invt]",
        "Inventory investment, total",
        "FIGARO input-output table (naio_10_fcp_ii3), inventories P52",
        "Same object: the inventories column.",
    ),
    (
        "factor_demand",
        "qInvt_i",
        "Inventory investment by industry",
        "FIGARO input-output table (naio_10_fcp_ii3), inventories P52",
        "Same object, by industry.",
    ),
    # factor_demand_energy
    (
        "factor_demand_energy",
        "qE_re_i",
        "Energy used as an input, by industry (money, prices = 1 in the data year)",
        "",
        "Not in Eurostat as money. We construct it from physical energy "
        "accounts (env_ac_pefasu) times prices.",
    ),
    (
        "factor_demand_energy",
        "qD[re]",
        "Total intermediate energy demand",
        "",
        "Not in Eurostat as money. Same construction, summed over industries.",
    ),
    (
        "factor_demand_energy",
        "qInvt_ene_i",
        "Energy held as inventory, by industry",
        "",
        "Physical joules are in env_ac_pefasu. We construct the money value.",
    ),
    (
        "factor_demand_energy",
        "qD[invt_ene]",
        "Energy inventory demand, total",
        "",
        "Not in Eurostat as money. Same construction, summed.",
    ),
    # pricing
    ("pricing", "—", NO_DATA, "", ""),
    # households
    (
        "households",
        "qD[c]",
        "Household consumption in 12 spending groups",
        "Eurostat household consumption by purpose (nama_10_co3_p3)",
        "Same idea, but Eurostat does not uniquely identify all 12 model "
        "groups; energy groups come from the energy accounts.",
    ),
    # financial_accounts
    (
        "financial_accounts",
        "vNetFinAssets",
        "Net financial assets by sector (households, firms, government, abroad)",
        "Eurostat financial accounts (nasa_10_f_bs)",
        "Already used in the model. Same object for the stocks the model "
        "reads.",
    ),
    (
        "financial_accounts",
        "vNetDebtInstruments",
        "Net debt instruments by sector",
        "Eurostat financial accounts (nasa_10_f_bs)",
        "Already used. Same object (debt instruments).",
    ),
    # government
    (
        "government",
        "qD[g]",
        "Government consumption",
        "FIGARO input-output table (naio_10_fcp_ii3), public consumption P3",
        "Same object: public consumption.",
    ),
    (
        "government",
        "vtIndirect",
        "Revenue from indirect taxes",
        "Eurostat government accounts (gov_10a_main)",
        "Same object as a total.",
    ),
    (
        "government",
        "vtDirect",
        "Revenue from direct taxes",
        "Eurostat government accounts (gov_10a_main, D5)",
        "Same object as a total (income and wealth taxes).",
    ),
    (
        "government",
        "vtCorp",
        "Corporation tax",
        "Eurostat government tax detail (gov_10a_taxag, D51)",
        "Same object as a total.",
    ),
    (
        "government",
        "vCont",
        "Social-security contributions",
        "Eurostat government accounts (gov_10a_main, D61)",
        "Same object as a total.",
    ),
    (
        "government",
        "vGovRevQuasi",
        "Withdrawals from public quasi-corporations",
        "Eurostat government accounts (gov_10a_main, D42)",
        "Same idea; some countries only publish it bundled with similar "
        "items.",
    ),
    (
        "government",
        "vGovRent",
        "Government rent received",
        "Eurostat government accounts (gov_10a_main, D45)",
        "Same idea; some countries only publish it bundled.",
    ),
    (
        "government",
        "vtGovDepr",
        "Consumption of public fixed capital (depreciation)",
        "Eurostat government accounts (gov_10a_main, P51c)",
        "Same object.",
    ),
    (
        "government",
        "vGovReceiveCorp",
        "Capital transfers from firms to government",
        "Eurostat government accounts (gov_10a_main, D9)",
        "Same object as a total.",
    ),
    (
        "government",
        "vGovReceiveCorpNonCap",
        "Other (current) transfers from firms to government",
        "Eurostat government accounts (gov_10a_main, D7)",
        "Same object as a total.",
    ),
    (
        "government",
        "vGovReceiveF",
        "Transfers from abroad to government",
        "Eurostat government accounts (gov_10a_main)",
        "Same idea; the domestic/abroad split is not published for every "
        "country.",
    ),
    (
        "government",
        "vtCap",
        "Capital taxes",
        "Eurostat government accounts (gov_10a_main, D91)",
        "Same object as a total.",
    ),
    (
        "government",
        "vGov2Corp",
        "Transfers from government to firms",
        "Eurostat government accounts (gov_10a_main)",
        "Same object as a total.",
    ),
    (
        "government",
        "vGovSub",
        "Subsidies to firms",
        "Eurostat government accounts (gov_10a_main, D3)",
        "Same object as a total.",
    ),
    (
        "government",
        "vHhTransfers",
        "Transfers to households",
        "Eurostat government accounts (gov_10a_main, D6)",
        "Same object as a total.",
    ),
    (
        "government",
        "vGov2Foreign",
        "Transfers from government to abroad",
        "Eurostat government accounts (gov_10a_main)",
        "Same idea; the domestic/abroad split is not published for every "
        "country.",
    ),
    (
        "government",
        "vGovNetAcquisitions",
        "Net purchases of non-produced assets (e.g. land)",
        "Eurostat government accounts (gov_10a_main, NP)",
        "Same object as a total.",
    ),
    # imports
    ("imports", "—", NO_DATA, "", ""),
    # exports
    (
        "exports",
        "qD[x]",
        "Exports",
        "FIGARO input-output table (naio_10_fcp_ii3), export column P6",
        "Same object; FIGARO has no re-export concept.",
    ),
    # production
    (
        "production",
        "qProd",
        "How much of each production factor an industry uses (labour, capital, intermediates, energy)",
        "",
        "Not one Eurostat table. We construct it from the IO table "
        "(naio_10_fcp_ii3), capital stocks (nama_10_nfa_st) and energy "
        "accounts (env_ac_pefasu).",
    ),
    (
        "production",
        "pProd",
        "Price of each production factor (set to 1 in the data year)",
        "",
        "Not in Eurostat. Set to 1 in the data year.",
    ),
    # ramsey_household
    ("ramsey_household", "—", NO_DATA, "", ""),
    # consumption_disaggregated
    ("consumption_disaggregated", "—", NO_DATA, "", ""),
    # emissions
    (
        "emissions",
        "qEmmE_BU",
        "Emissions from burning fuels, by gas, purpose, product and user",
        "Eurostat air-emissions accounts (env_ac_ainah_r2)",
        "Only as a total by industry and gas — energy and process are "
        "combined, and there is no fuel or purpose split.",
    ),
    (
        "emissions",
        "qEmmxE",
        "Process (non-energy) emissions, by gas and user",
        "Eurostat air-emissions accounts (env_ac_ainah_r2)",
        "F-gases are the same object; CH4 and N2O are a constructed "
        "energy/process split of the same table.",
    ),
    (
        "emissions",
        "qEmmLULUCF",
        "Land-use, land-use change and forestry emissions",
        "Eurostat bridging items / UNFCCC inventory (env_ac_aibrid_r2; "
        "env_air_gge)",
        "Same object (land-use emissions); vintage can differ.",
    ),
    (
        "emissions",
        "qEmmBorderTrade",
        "Residence adjustment: residents’ fuel bought abroad vs fuel sold on the territory",
        "Eurostat air-emissions bridging items (env_ac_aibrid_r2)",
        "Same object as a net residence adjustment; Eurostat splits by "
        "transport mode, not the Danish two-row split.",
    ),
    (
        "emissions",
        "qEmmBunkering",
        "International bunker emissions (ships and aircraft)",
        "",
        "We take it from the model's energy-related bunker rows. UNFCCC has "
        "a territorial counterpart, not this residence concept.",
    ),
    (
        "emissions",
        "sBioNatGas",
        "Biogenic share of natural-gas CO2",
        "",
        "Not in Eurostat. We derive it from bio vs fossil CO2 on natural gas "
        "in the energy accounts (env_ac_pefasu).",
    ),
    # energy_markets
    (
        "energy_markets",
        "qEpj",
        "Physical energy use by purpose, product and user",
        "Eurostat physical energy accounts (env_ac_pefasu)",
        "Same energy, in terajoules — divide by 1,000 for PJ. No purpose "
        "split.",
    ),
    (
        "energy_markets",
        "qEpj_own",
        "Energy used from own production (unpriced, not in the national accounts)",
        "Eurostat physical energy accounts (env_ac_pefasu)",
        "Same unpriced rows (e.g. own-use); no money counterpart.",
    ),
    (
        "energy_markets",
        "pEpj_base",
        "Producer price of energy, per PJ",
        "",
        "Not published at this detail. We construct it from energy "
        "quantities (env_ac_pefasu) and supply-use money totals "
        "(naio_10_cp15).",
    ),
    (
        "energy_markets",
        "pEpj_own",
        "Price attached to own-use energy",
        "",
        "Not in Eurostat. Same construction; often zero when the flow has no "
        "money.",
    ),
    (
        "energy_markets",
        "qY_CET",
        "Domestic production of each product (energy in PJ, everything else in money)",
        "Eurostat physical energy accounts (env_ac_pefasu); FIGARO "
        "input-output table (naio_10_fcp_ii3)",
        "Energy: same supply. Non-energy: same output.",
    ),
    (
        "energy_markets",
        "qM_CET",
        "Imports of each product (energy in PJ, everything else in money)",
        "Eurostat physical energy accounts (env_ac_pefasu); FIGARO "
        "input-output table (naio_10_fcp_ii3)",
        "Energy: same imports. Non-energy: same imports.",
    ),
    (
        "energy_markets",
        "pY_CET",
        "Price of domestic supply of each product",
        "",
        "Energy: we construct as money per PJ. Non-energy: set to 1 (FIGARO "
        "naio_10_fcp_ii3 is already money).",
    ),
    (
        "energy_markets",
        "pM_CET",
        "Price of imports of each product",
        "",
        "Not in Eurostat as a price. Same idea on the import side.",
    ),
    (
        "energy_markets",
        "vE_margins",
        "Wholesale, retail and garage mark-ups on energy",
        "Eurostat supply-use tables (naio_10_cp15)",
        "Combined trade margins only. We construct the three separate "
        "mark-ups.",
    ),
    # non_energy_markets
    ("non_energy_markets", "—", NO_DATA, "", ""),
    # production_CES_energydemand
    ("production_CES_energydemand", "—", NO_DATA, "", ""),
    # energy_and_emissions_taxes
    (
        "energy_and_emissions_taxes",
        "vtE_duty",
        "Energy-duty revenue by tax, purpose, product and user",
        "Eurostat environmental-tax accounts (env_ac_taxind2)",
        "Payer totals only. We construct the split across fuels and users.",
    ),
    (
        "energy_and_emissions_taxes",
        "vtE_vat",
        "VAT on energy, by purpose, product and user",
        "",
        "Not in Eurostat at this grain. We apply the statutory VAT rate to "
        "the taxable energy base.",
    ),
    (
        "energy_and_emissions_taxes",
        "tEmarg_duty",
        "Marginal energy-tax rate per PJ",
        "",
        "Not in Eurostat. We use average tax per PJ as a stand-in for a "
        "marginal rate.",
    ),
    (
        "energy_and_emissions_taxes",
        "tCO2_Emarg",
        "Marginal carbon-tax rate",
        "",
        "Not in Eurostat as a cell. ETS price is external; a separate "
        "national CO2 rate is constructed or set to zero.",
    ),
    (
        "energy_and_emissions_taxes",
        "qCO2_ETS_freeallowances",
        "Free ETS allowances by industry",
        "EU ETS Union Registry",
        "Same free-allowance totals; industry labels need a bridge from "
        "installation to industry.",
    ),
    # production_CET
    (
        "production_CET",
        "qY_CETgross",
        "Gross domestic production of each product, including own-use",
        "Eurostat physical energy accounts (env_ac_pefasu); FIGARO "
        "input-output table (naio_10_fcp_ii3)",
        "Energy: supply plus unpriced own-use. Non-energy: FIGARO output.",
    ),
    (
        "production_CET",
        "qY_CETown",
        "Own-use production (not sold)",
        "Eurostat physical energy accounts (env_ac_pefasu)",
        "Unpriced rows. Not in the money tables.",
    ),
    # consumption_disaggregated_energy
    ("consumption_disaggregated_energy", "—", NO_DATA, "", ""),
    # exports_energy
    ("exports_energy", "—", NO_DATA, "", ""),
    # energy_technology
    (
        "energy_technology",
        "sqTPotential",
        "How much of an energy service each technology can supply",
        "",
        "Danish Energy Agency technology catalogue / dummy data, not "
        "Eurostat.",
    ),
    (
        "energy_technology",
        "uTE",
        "Fuel input per unit of energy-service output, by technology",
        "",
        "Danish Energy Agency technology catalogue / dummy data, not "
        "Eurostat.",
    ),
    (
        "energy_technology",
        "vTI",
        "Investment cost of a technology, per PJ of output",
        "",
        "Danish Energy Agency technology catalogue / dummy data, not "
        "Eurostat.",
    ),
    (
        "energy_technology",
        "vTC",
        "Running capital cost of a technology, per PJ of output",
        "",
        "Danish Energy Agency technology catalogue / dummy data, not "
        "Eurostat.",
    ),
    (
        "energy_technology",
        "pTK",
        "User cost of capital in the technology model",
        "",
        "Not in Eurostat. Linked to the model's capital price; catalogue / "
        "dummy data.",
    ),
    (
        "energy_technology",
        "qES",
        "Quantity of each energy service demanded",
        "",
        "Not in Eurostat. Calibrated to the energy-market quantities (qEpj).",
    ),
]

MODULE_WIDTHS = [28, 22, 48, 52, 52]


def _autofit_row_heights(ws, widths, first_data_row=2):
    """Rough row-height estimate so wrapped text stays visible."""
    for row in ws.iter_rows(min_row=first_data_row):
        lines = 1
        for cell, width in zip(row, widths):
            if cell.value:
                text = str(cell.value)
                lines = max(lines, -(-len(text) // max(width - 2, 10)))
        row[0].parent.row_dimensions[row[0].row].height = min(15 * lines, 90)


def _readme_lines(today: str) -> list[str]:
    """Shared Read me prose for the Excel sheet and the Markdown twin."""
    return [
        "What the GREU model uses",
        "",
        f"Created: {today}. Generated by "
        "data/preprocessing/scripts/build_what_the_model_uses_xlsx.py — "
        "do not edit this file by hand.",
        "",
        "This is a one-page explainer for people who do not work in the "
        "model. It answers: what kind of number goes in, where it comes "
        "from, and what it is for.",
        "",
        "The model needs two kinds of numbers:",
        "  1. Money — who pays whom (sales, wages, taxes, investment).",
        "  2. Physical energy — how much fuel, electricity and heat is "
        "used, in petajoules (PJ).",
        "",
        "Units, in one paragraph. A petajoule is a large amount of energy. "
        "Eurostat publishes the same physical energy in terajoules (TJ), "
        "which are 1,000 times smaller — so we divide by 1,000. Same "
        "energy, different unit. FIGARO is also a Eurostat product, but "
        "it is a money table of the economy. It has no energy in joules.",
        "",
        "How to read the sheets.",
        "  • 'What the model uses' is the everyday overview. Columns from "
        "the left are in plain language. The last column is a name used "
        "inside the model — ignore it unless a modeller is in the room.",
        "  • 'By model module' is for reading the model files. One row per "
        "variable that takes data, grouped as the modules in "
        "model/modules/. Filter column A to stay in one module. "
        "'Where to find it' is the Eurostat / FIGARO / ETS table, with its "
        "code (e.g. env_ac_pefasu). It is blank when that variable is not "
        "in those sources. 'Considerations' says whether it is the same "
        "object, only a total, or something we construct.",
        "",
        "What this file is not. It is not a status report on replacing "
        "Danish data with EU data. That lives in docs/EU_data_overview.xlsx.",
    ]


def _write_readme(ws, today: str) -> None:
    for line in _readme_lines(today):
        ws.append((line,) if line else ())
    ws["A1"].font = Font(bold=True, size=16, color="1F4E79")
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in ws.iter_rows(min_row=3, max_col=1):
        for cell in row:
            cell.alignment = wrap
    # Let the prose sit in a wide first column so it reads as a short memo.
    ws.column_dimensions["A"].width = 110
    for r in range(3, ws.max_row + 1):
        ws.row_dimensions[r].height = 44
    ws.row_dimensions[1].height = 24
    ws.sheet_view.showGridLines = False


def _write_main(ws) -> None:
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 22

    for row in ROWS:
        ws.append(row)

    for col_idx, width in enumerate(WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    last_col_font = Font(italic=True, color="7A8798", size=9)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_TOP
            cell.border = THIN_BORDER
        area = row[0].value
        fill_hex = AREA_FILLS.get(area)
        if fill_hex:
            row[0].fill = PatternFill("solid", fgColor=fill_hex)
            row[0].font = Font(bold=True)
            row[0].alignment = Alignment(
                vertical="top", horizontal="center", wrap_text=True
            )
        row[-1].font = last_col_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{ws.max_row}"
    _autofit_row_heights(ws, WIDTHS)


def _write_module(ws) -> None:
    ws.append(MODULE_HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[1].height = 22

    for row in MODULE_ROWS:
        ws.append(row)

    for col_idx, width in enumerate(MODULE_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    var_font = Font(name="Consolas", size=10)
    grey_font = Font(italic=True, color="7A8798")
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP_TOP
            cell.border = THIN_BORDER
        module = row[0].value
        is_no_data = row[1].value == "—"
        fill_hex = NO_DATA_FILL if is_no_data else MODULE_FILLS.get(module)
        if fill_hex:
            row[0].fill = PatternFill("solid", fgColor=fill_hex)
            row[0].font = Font(bold=True)
            row[0].alignment = Alignment(
                vertical="top", horizontal="left", wrap_text=True
            )
        if is_no_data:
            for cell in row[1:]:
                cell.font = grey_font
        else:
            row[1].font = var_font
            if not row[3].value:
                row[3].font = grey_font

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(MODULE_HEADERS))}{ws.max_row}"
    )
    _autofit_row_heights(ws, MODULE_WIDTHS)


def _md_cell(value) -> str:
    text = "" if value is None else str(value)
    return text.replace("|", "\\|").replace("\n", " ")


def _md_table(headers, rows) -> str:
    head = "| " + " | ".join(_md_cell(h) for h in headers) + " |"
    sep = "| " + " | ".join("---" for _ in headers) + " |"
    body = [
        "| " + " | ".join(_md_cell(c) for c in row) + " |"
        for row in rows
    ]
    return "\n".join([head, sep, *body])


def build_markdown(today: str) -> str:
    lines = _readme_lines(today)
    title, rest = lines[0], lines[1:]
    parts = [
        f"# {title}",
        *[line if line else "" for line in rest],
        "",
        "## What the model uses",
        "",
        _md_table(HEADERS, ROWS),
        "",
        "## By model module",
        "",
        _md_table(MODULE_HEADERS, MODULE_ROWS),
        "",
    ]
    return "\n".join(parts)


def build_workbook() -> Workbook:
    wb = Workbook()
    ws_readme = wb.active
    ws_readme.title = "Read me"
    _write_readme(ws_readme, dt.date.today().isoformat())

    ws_main = wb.create_sheet("What the model uses")
    _write_main(ws_main)

    ws_mod = wb.create_sheet("By model module")
    _write_module(ws_mod)
    return wb


def main() -> None:
    today = dt.date.today().isoformat()
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)

    wb = build_workbook()
    wb.save(OUTPUT_PATH)
    print(f"Wrote {OUTPUT_PATH}")

    MD_OUTPUT_PATH.write_text(build_markdown(today), encoding="utf-8")
    print(f"Wrote {MD_OUTPUT_PATH}")

    n_data = sum(1 for row in MODULE_ROWS if row[1] != "—")
    n_empty = sum(1 for row in MODULE_ROWS if row[1] == "—")
    print(f"Overview rows: {len(ROWS)}; "
          f"module data rows: {n_data}; no-data modules: {n_empty}")


if __name__ == "__main__":
    main()
