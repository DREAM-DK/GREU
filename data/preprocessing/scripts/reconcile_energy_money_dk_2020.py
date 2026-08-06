"""Audit whether public EU data can build GREU's Denmark-2020 monetary energy layer.

This is deliberately a feasibility/gap construction, not a synthetic replacement.
No missing energy-product-by-user valuation cell is filled from Danish shares.

Inputs
------
- energy_and_emissions.xlsx and io_energy_long_format.xlsx (validation benchmarks)
- metadata.xlsx (classification concordances)
- official raw files downloaded by download_energy_money_sources_dk_2020.py
- prior PEFA reconciliation (physical-coverage benchmark)

Output
------
data/preprocessing/data/energy_money_dk2020_feasibility_gap.xlsx
"""

from __future__ import annotations

import json
import pathlib
from collections import OrderedDict

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DATA = pathlib.Path(__file__).resolve().parents[1] / "data"
RAW = DATA / "eurostat_energy_money_raw"
OUT = DATA / "energy_money_dk2020_feasibility_gap.xlsx"
YEAR = 2020
RETRIEVAL_DATE = "2026-07-30"
DKK_PER_EUR = 7.4542

MONEY_FIELDS = [
    "basic", "ws_marg", "ret_marg", "mvs_marg", "ener_tax", "co2_tax",
    "so2_tax", "nox_tax", "pso_tax", "vat", "purch",
]
MARGINS = ["ws_marg", "ret_marg", "mvs_marg"]
TAXES = ["ener_tax", "co2_tax", "so2_tax", "nox_tax", "pso_tax"]
ENERGY_CPA = OrderedDict(
    [
        ("CPA_B05", "Coal and lignite"),
        ("CPA_B06", "Crude petroleum and natural gas"),
        ("CPA_C19", "Coke and refined petroleum products"),
        ("CPA_D", "Electricity, gas, steam and air conditioning"),
    ]
)


def jsonstat_to_long(path: pathlib.Path) -> tuple[pd.DataFrame, dict]:
    payload = json.loads(path.read_text(encoding="utf-8"))
    dims = payload["id"]
    categories: list[list[str]] = []
    labels: dict[str, dict[str, str]] = {}
    for dim in dims:
        category = payload["dimension"][dim]["category"]
        ordered = [
            key for key, _ in sorted(category["index"].items(), key=lambda item: item[1])
        ]
        categories.append(ordered)
        labels[dim] = category.get("label", {})
    rows = []
    statuses = payload.get("status", {})
    for flat_index, value in payload["value"].items():
        positions = np.unravel_index(int(flat_index), payload["size"])
        row = {dim: categories[i][positions[i]] for i, dim in enumerate(dims)}
        row["value"] = value
        row["obs_status"] = statuses.get(str(flat_index), "")
        rows.append(row)
    return pd.DataFrame(rows), labels


def norm_code(value) -> str:
    if pd.isna(value):
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    if text.isdigit() and len(text) < 5:
        text = text.zfill(5)
    return text


def style_workbook(path: pathlib.Path) -> None:
    workbook = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column in sheet.columns:
            letter = get_column_letter(column[0].column)
            max_length = max(len(str(cell.value or "")) for cell in column)
            sheet.column_dimensions[letter].width = min(max(max_length + 2, 10), 58)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.000000"
    workbook.save(path)


def oil_annual_average(path: pathlib.Path) -> pd.DataFrame:
    rows = []
    for sheet, price_basis in [
        ("Prices with taxes", "purchaser_like_including_taxes"),
        ("Prices wo taxes", "consumer_price_excluding_duties_and_taxes"),
    ]:
        frame = pd.read_excel(path, sheet_name=sheet)
        dates = pd.to_datetime(frame.iloc[:, 0], errors="coerce", format="mixed")
        selected = frame.loc[dates.dt.year == YEAR].copy()
        for column in [col for col in frame.columns if str(col).startswith("DK_price_")]:
            values = pd.to_numeric(selected[column], errors="coerce").dropna()
            rows.append(
                {
                    "source_sheet": sheet,
                    "price_basis": price_basis,
                    "series": column,
                    "year": YEAR,
                    "weekly_observations": len(values),
                    "annual_mean_source_units": values.mean() if len(values) else np.nan,
                    "source_unit": (
                        "EUR per 1000 litres, except fuel oils EUR per tonne"
                    ),
                    "role": "Product/market price control only; no NACE or GREU purpose dimension",
                }
            )
    return pd.DataFrame(rows)


def oil_rate_snapshot(path: pathlib.Path) -> pd.DataFrame:
    rows = []
    for sheet in ["VAT", "Excise duties", "Other Indirect Taxes"]:
        raw = pd.read_excel(path, sheet_name=sheet, header=None)
        country = raw.iloc[:, 0].astype("string").str.replace("_", "", regex=False).ffill()
        dates = pd.to_datetime(raw.iloc[:, 1], errors="coerce", format="mixed")
        mask = (country == "DK") & dates.notna() & (dates <= pd.Timestamp("2020-12-31"))
        product_names = raw.iloc[2, 2:].astype("string").fillna("")
        units = raw.iloc[3, 2:].astype("string").fillna("")
        for offset in range(2, raw.shape[1]):
            product_mask = mask & raw.iloc[:, offset].notna()
            if not product_mask.any():
                continue
            selected_index = dates[product_mask].idxmax()
            value = raw.loc[selected_index, offset]
            rows.append(
                {
                    "source_sheet": sheet,
                    "country": "DK",
                    "effective_from": dates.loc[selected_index].date().isoformat(),
                    "situation_on": "2020-12-31",
                    "product": product_names.iloc[offset - 2],
                    "unit": "%" if sheet == "VAT" else units.iloc[offset - 2],
                    "value": value,
                    "role": "Legal/retail control; not a complete product×user tax matrix",
                }
            )
    return pd.DataFrame(rows)


def main() -> None:
    energy = pd.read_excel(DATA / "energy_and_emissions.xlsx", sheet_name="ems_energy")
    io = pd.read_excel(DATA / "io_energy_long_format.xlsx", sheet_name="io")
    industry_map = pd.read_excel(
        DATA / "metadata.xlsx", sheet_name="industries_naceA64_map"
    )
    energy = energy[energy["year"] == YEAR].copy()
    io = io[io["year"] == YEAR].copy()
    for field in ["pj", *MONEY_FIELDS]:
        energy[field] = pd.to_numeric(energy[field], errors="coerce").fillna(0.0)
    io["value"] = pd.to_numeric(io["value"], errors="coerce").fillna(0.0)

    # --------------------------- Target identities ---------------------------
    energy["component_sum"] = energy[
        ["basic", *MARGINS, *TAXES, "vat"]
    ].sum(axis=1)
    energy["purch_identity_residual"] = energy["purch"] - energy["component_sum"]
    # Purchaser values are a use-side concept here. Supply rows intentionally
    # carry basic values but no purchaser value, so they are outside this test.
    identity_cells = energy[energy["bal"] == "use"][
        [
            "year", "bal", "flow", "indu", "purp", "product", "pj", "basic",
            *MARGINS, *TAXES, "vat", "purch", "component_sum",
            "purch_identity_residual",
        ]
    ].copy()
    identity_cells["identity_status"] = np.where(
        identity_cells["purch_identity_residual"].abs() <= 1e-8, "PASS", "CHECK"
    )

    target_rows = []
    for field in ["pj", *MONEY_FIELDS]:
        for bal in ["sup", "use"]:
            values = energy.loc[energy["bal"] == bal, field]
            target_rows.append(
                {
                    "field": field,
                    "balance_side": bal,
                    "unit": "PJ" if field == "pj" else "bn DKK",
                    "target_total": values.sum(),
                    "nonzero_rows": int((values.abs() > 1e-12).sum()),
                    "total_rows": len(values),
                }
            )
    target_benchmarks = pd.DataFrame(target_rows)

    # Reconcile the two Danish target workbooks at column level.
    io_cols = io.copy()
    io_cols["col_l2_norm"] = io_cols["col_l2"].map(norm_code)
    io_cols["comparison_flow"] = np.where(
        io_cols["col_l1"].astype(str).str.startswith("cons_hh"),
        "cons_hh",
        io_cols["col_l1"].astype(str),
    )
    io_cols["comparison_user"] = io_cols["col_l2_norm"]
    io_cols.loc[
        io_cols["comparison_flow"].isin(
            ["export", "invent_change", "cons_publ", "invest_build", "invest_trans", "invest_other"]
        ),
        "comparison_user",
    ] = ""
    io_col_totals = (
        io_cols.groupby(["comparison_flow", "comparison_user"], dropna=False)["value"]
        .sum()
        .rename("io_total_bn_dkk")
        .reset_index()
    )

    use = energy[energy["bal"] == "use"].copy()
    use["comparison_flow"] = use["flow"].astype(str)
    use["comparison_user"] = use["indu"].map(norm_code)
    use.loc[use["flow"].isin(["export", "invent_change"]), "comparison_user"] = ""
    target_col_totals = (
        use.groupby(["comparison_flow", "comparison_user"], dropna=False)
        .agg(
            energy_purch_bn_dkk=("purch", "sum"),
            energy_components_bn_dkk=("component_sum", "sum"),
            energy_pj=("pj", "sum"),
        )
        .reset_index()
    )
    io_column_recon = io_col_totals.merge(
        target_col_totals, on=["comparison_flow", "comparison_user"], how="outer"
    ).fillna(0.0)
    io_column_recon["io_minus_energy_purch_bn_dkk"] = (
        io_column_recon["io_total_bn_dkk"] - io_column_recon["energy_purch_bn_dkk"]
    )
    io_column_recon["status"] = np.select(
        [
            io_column_recon["io_minus_energy_purch_bn_dkk"].abs() <= 1e-5,
            io_column_recon["io_minus_energy_purch_bn_dkk"].abs() <= 0.0021,
        ],
        ["PASS_EXACT", "PASS_SOURCE_ROUNDING"],
        default="CHECK_SCOPE_OR_CELL",
    )

    # Production-row reconciliation, including the explicit rerouting of margins.
    io_prod = (
        io[io["row_l1"] == "production"]
        .assign(row_l2_norm=lambda frame: frame["row_l2"].map(norm_code))
        .groupby("row_l2_norm")["value"]
        .sum()
        .rename("io_production_row_bn_dkk")
    )
    supply_prod = (
        energy[(energy["bal"] == "sup") & (energy["flow"] == "production")]
        .assign(indu_norm=lambda frame: frame["indu"].map(norm_code))
        .groupby("indu_norm")["basic"]
        .sum()
        .rename("energy_supply_basic_bn_dkk")
    )
    expected_prod = pd.concat([io_prod, supply_prod], axis=1).fillna(0.0)
    expected_prod["rerouted_margin_bn_dkk"] = 0.0
    expected_prod.loc["45000", "rerouted_margin_bn_dkk"] = use["mvs_marg"].sum()
    expected_prod.loc["46000", "rerouted_margin_bn_dkk"] = use["ws_marg"].sum()
    expected_prod.loc["47000", "rerouted_margin_bn_dkk"] = use["ret_marg"].sum()
    expected_prod["expected_energy_row_bn_dkk"] = (
        expected_prod["energy_supply_basic_bn_dkk"]
        + expected_prod["rerouted_margin_bn_dkk"]
    )
    expected_prod["io_minus_expected_bn_dkk"] = (
        expected_prod["io_production_row_bn_dkk"]
        - expected_prod["expected_energy_row_bn_dkk"]
    )
    expected_prod["status"] = np.select(
        [
            expected_prod["io_minus_expected_bn_dkk"].abs() <= 1e-5,
            expected_prod["io_minus_expected_bn_dkk"].abs() <= 0.0021,
        ],
        ["PASS_EXACT", "PASS_SOURCE_ROUNDING"],
        default="CHECK",
    )
    io_row_recon = expected_prod.reset_index(names="row_industry")

    io_totals = (
        io.groupby("row_l1")["value"].sum().rename("io_bn_dkk").reset_index()
    )
    target_io_expected = pd.DataFrame(
        [
            {
                "row_l1": "production",
                "expected_bn_dkk": energy.loc[
                    (energy["bal"] == "sup") & (energy["flow"] == "production"), "basic"
                ].sum()
                + use[MARGINS].sum().sum(),
                "derivation": "domestic basic supply + three use-side margins",
            },
            {
                "row_l1": "import",
                "expected_bn_dkk": energy.loc[
                    (energy["bal"] == "sup") & (energy["flow"] == "import"), "basic"
                ].sum(),
                "derivation": "import basic supply; industry split is proportional in Danish IO",
            },
            {
                "row_l1": "prim_input",
                "expected_bn_dkk": use[TAXES + ["vat"]].sum().sum(),
                "derivation": "five energy taxes + VAT",
            },
        ]
    )
    io_total_recon = io_totals.merge(target_io_expected, on="row_l1", how="outer")
    io_total_recon["io_minus_expected_bn_dkk"] = (
        io_total_recon["io_bn_dkk"] - io_total_recon["expected_bn_dkk"]
    )
    io_total_recon["status"] = np.select(
        [
            io_total_recon["io_minus_expected_bn_dkk"].abs() <= 1e-5,
            io_total_recon["io_minus_expected_bn_dkk"].abs() <= 0.0021,
        ],
        ["PASS_EXACT", "PASS_SOURCE_ROUNDING"],
        default="CHECK",
    )

    # --------------------------- Official controls ---------------------------
    tax, tax_labels = jsonstat_to_long(RAW / "env_ac_taxind2_DK_2020.json")
    tax["tax_label"] = tax["tax"].map(tax_labels["tax"])
    tax["payer_label"] = tax["nace_r2"].map(tax_labels["nace_r2"])
    tax_by_payer = tax[
        (tax["unit"] == "MIO_EUR")
        & tax["tax"].isin(["NRG", "NRG_CO2"])
        & ~tax["nace_r2"].isin(["TOTAL", "A", "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U"])
    ].copy()
    tax_by_payer["bn_dkk"] = tax_by_payer["value"] * DKK_PER_EUR / 1000.0
    tax_by_payer = tax_by_payer[
        ["tax", "tax_label", "nace_r2", "payer_label", "unit", "value", "bn_dkk", "obs_status"]
    ]

    price_frames = []
    price_names = {
        "nrg_pc_202_c": ("natural_gas", "household"),
        "nrg_pc_203_c": ("natural_gas", "non_household"),
        "nrg_pc_204_c": ("electricity", "household"),
        "nrg_pc_205_c": ("electricity", "non_household"),
    }
    for dataset, (product_family, user_class) in price_names.items():
        frame, labels = jsonstat_to_long(RAW / f"{dataset}_DK_2020.json")
        frame["dataset"] = dataset
        frame["product_family"] = product_family
        frame["user_class"] = user_class
        frame["component_label"] = frame["nrg_prc"].map(labels["nrg_prc"])
        frame["band_label"] = frame["nrg_cons"].map(labels["nrg_cons"])
        frame["unit"] = frame["unit"] if "unit" in frame else "KWH"
        price_frames.append(frame)
    price_components = pd.concat(price_frames, ignore_index=True)
    price_components = price_components[
        (price_components["currency"] == "EUR")
        & price_components["nrg_cons"].str.startswith("TOT")
    ].copy()
    price_components["dkk_per_gj"] = np.where(
        price_components["unit"] == "GJ_GCV",
        price_components["value"] * DKK_PER_EUR,
        price_components["value"] * DKK_PER_EUR * (1_000_000 / 3_600),
    )
    price_components = price_components[
        [
            "dataset", "product_family", "user_class", "nrg_prc",
            "component_label", "nrg_cons", "band_label", "currency", "unit",
            "value", "dkk_per_gj", "obs_status",
        ]
    ]

    cp15, cp15_labels = jsonstat_to_long(RAW / "naio_10_cp15_DK_2020.json")
    cp16, cp16_labels = jsonstat_to_long(RAW / "naio_10_cp16_DK_2020.json")
    cp15 = cp15[(cp15["unit"] == "MIO_EUR") & cp15["prd_amo"].isin(ENERGY_CPA)].copy()
    cp16 = cp16[(cp16["unit"] == "MIO_EUR") & cp16["prd_ava"].isin(ENERGY_CPA)].copy()
    supply_codes = ["TS_BP", "D21X31", "OTTM", "TS_PP", "P7"]
    sut_controls = cp15[cp15["ind_impv"].isin(supply_codes)].copy()
    sut_controls["product_label"] = sut_controls["prd_amo"].map(ENERGY_CPA)
    sut_controls["valuation_item_label"] = sut_controls["ind_impv"].map(
        cp15_labels["ind_impv"]
    )
    sut_controls["bn_dkk"] = sut_controls["value"] * DKK_PER_EUR / 1000.0
    sut_controls = sut_controls[
        [
            "prd_amo", "product_label", "ind_impv", "valuation_item_label",
            "unit", "value", "bn_dkk", "obs_status",
        ]
    ]

    sut_pivot = sut_controls.pivot_table(
        index=["prd_amo", "product_label"], columns="ind_impv", values="value", aggfunc="sum"
    ).reset_index()
    for column in supply_codes:
        if column not in sut_pivot:
            sut_pivot[column] = 0.0
    cp16_total = (
        cp16[cp16["ind_use"] == "TU"]
        .set_index("prd_ava")["value"]
        .rename("use_purchaser_mio_eur")
    )
    sut_identity = sut_pivot.set_index("prd_amo").join(cp16_total).reset_index()
    sut_identity["supply_valuation_residual_mio_eur"] = (
        sut_identity["TS_PP"]
        - sut_identity["TS_BP"]
        - sut_identity["D21X31"]
        - sut_identity["OTTM"]
    )
    sut_identity["supply_minus_use_mio_eur"] = (
        sut_identity["TS_PP"] - sut_identity["use_purchaser_mio_eur"]
    )
    sut_identity["status"] = np.where(
        (sut_identity["supply_valuation_residual_mio_eur"].abs() <= 0.2)
        & (sut_identity["supply_minus_use_mio_eur"].abs() <= 0.2),
        "PASS",
        "CHECK_ROUNDING_OR_SOURCE",
    )
    sut_identity["scope_warning"] = (
        "Broad CPA product, not PEFA energy product; controls totals only"
    )

    # Product-by-user purchaser controls at A64/final-demand level.
    a64 = set(industry_map["indu_naceA64"].astype(str).str.replace("_", "-", regex=False))
    final_uses = {"P3_S13", "P3_S14", "P3_S15", "P51G", "P52", "P53", "P6"}
    sut_use_controls = cp16[
        cp16["ind_use"].isin(a64 | final_uses)
    ].copy()
    sut_use_controls["product_label"] = sut_use_controls["prd_ava"].map(ENERGY_CPA)
    sut_use_controls["user_label"] = sut_use_controls["ind_use"].map(cp16_labels["ind_use"])
    sut_use_controls["bn_dkk"] = sut_use_controls["value"] * DKK_PER_EUR / 1000.0
    sut_use_controls["mapping_status"] = (
        "Calibration control only: CPA bundles cannot identify GREU energy products"
    )
    sut_use_controls = sut_use_controls[
        [
            "prd_ava", "product_label", "ind_use", "user_label", "unit", "value",
            "bn_dkk", "obs_status", "mapping_status",
        ]
    ]

    oil_path = RAW / "Weekly_Oil_Bulletin_Prices_History_2026-07-30.xlsx"
    oil_prices = oil_annual_average(oil_path)
    oil_rates = oil_rate_snapshot(oil_path)

    # ---------------------- Coverage and model interface ---------------------
    prior = pd.read_excel(
        DATA / "eurostat_energy_emissions_dk2020_reconciliation.xlsx",
        sheet_name="energy_totals",
    )
    physical_row = prior[
        prior["metric"].str.contains("concept-adjusted PEFA boundary", na=False)
    ].iloc[0]
    physical_share = physical_row["eurostat"] / physical_row["dk_greu"] * 100

    exact_control_products = {"power", "gasol_transp", "diesel_transp"}
    near_control_products = exact_control_products | {"natgas_incl_biongas"}
    exact_pj = use.loc[use["product"].isin(exact_control_products), "pj"].sum()
    near_pj = use.loc[use["product"].isin(near_control_products), "pj"].sum()
    exact_purch = use.loc[use["product"].isin(exact_control_products), "purch"].sum()
    near_purch = use.loc[use["product"].isin(near_control_products), "purch"].sum()
    total_use_pj = use["pj"].sum()
    total_purch = use["purch"].sum()

    coverage_summary = pd.DataFrame(
        [
            {
                "measure": "Physical PEFA comparable flow",
                "numerator": physical_row["eurostat"],
                "denominator": physical_row["dk_greu"],
                "unit": "PJ",
                "coverage_pct": physical_share,
                "interpretation": "Direct physical source at PEFA product×NACE, but no GREU purpose split",
            },
            {
                "measure": "Monetary cells directly source-complete",
                "numerator": 0,
                "denominator": int((use["pj"].abs() > 1e-12).sum()),
                "unit": "nonzero GREU use rows",
                "coverage_pct": 0.0,
                "interpretation": "No official source jointly supplies product×user basic, 3 margins, 5 taxes, VAT",
            },
            {
                "measure": "Exact-family retail/price-control envelope",
                "numerator": exact_pj,
                "denominator": total_use_pj,
                "unit": "PJ benchmark",
                "coverage_pct": exact_pj / total_use_pj * 100,
                "interpretation": "Electricity, road gasoline, road diesel; controls only, not cell values",
            },
            {
                "measure": "Near-family price-control envelope incl. mixed natural gas/biogas",
                "numerator": near_pj,
                "denominator": total_use_pj,
                "unit": "PJ benchmark",
                "coverage_pct": near_pj / total_use_pj * 100,
                "interpretation": "Natural-gas statistics do not separate the biogas in GREU's combined product",
            },
            {
                "measure": "Exact-family purchaser-value benchmark share",
                "numerator": exact_purch,
                "denominator": total_purch,
                "unit": "bn DKK target benchmark",
                "coverage_pct": exact_purch / total_purch * 100,
                "interpretation": "Danish target used only to quantify control-envelope importance",
            },
            {
                "measure": "Near-family purchaser-value benchmark share",
                "numerator": near_purch,
                "denominator": total_purch,
                "unit": "bn DKK target benchmark",
                "coverage_pct": near_purch / total_purch * 100,
                "interpretation": "Not direct source coverage and not used to fill values",
            },
        ]
    )

    field_requirements = pd.DataFrame(
        [
            ["Physical product×user flows", "pj", "YES", "EnergyBalance and model quantities", "PEFA direct at product×NACE; coarser and no purpose", "PARTIAL"],
            ["Basic values", "basic", "YES", "EnergyBalance; IO production/import tests", "SUT/FIGARO broad CPA controls only", "GAP"],
            ["Wholesale margin", "ws_marg / EAV", "YES", "EnergyBalance; rerouted to industry 46000", "SUT has combined trade+transport margin, not this split", "GAP"],
            ["Retail margin", "ret_marg / DAV", "YES", "EnergyBalance; rerouted to industry 47000", "SUT has combined trade+transport margin, not this split", "GAP"],
            ["Motor-vehicle-sales margin", "mvs_marg / CAV", "YES", "EnergyBalance; rerouted to industry 45000", "No corresponding EU-wide valuation field", "GAP"],
            ["Energy tax", "ener_tax", "YES", "etaxes sum and IO TaxSub identity", "Tax accounts payer total; price/rate controls for selected products", "GAP"],
            ["CO2 tax", "co2_tax", "YES", "etaxes sum and total-price set", "env_ac_taxind2 NRG_CO2 payer control, not product rate/cell", "GAP"],
            ["SO2 tax", "so2_tax", "YES", "etaxes sum and total-price set", "No EU-wide product×user field", "GAP"],
            ["NOx tax", "nox_tax", "YES", "etaxes sum and total-price set", "No EU-wide product×user field", "GAP"],
            ["PSO tax/levy", "pso_tax", "YES", "etaxes sum and total-price set", "Electricity price component is broader and not GREU-equivalent", "GAP"],
            ["VAT", "vat", "YES", "EnergyBalance and IO Moms identity", "Rate is public; taxable base/recovery/exemptions are not a cell matrix", "GAP"],
            ["Purchaser value", "purch", "NO (stored)", "Stacked by read_data.py but excluded from GAMS total-price set", "Can be derived only after all components exist", "DERIVED"],
            ["Energy-only IO layout", "io_energy_long_format", "YES", "Subtracts energy IO from full IO and exports energy IO parameters", "Must be generated from the complete component layer", "GAP"],
        ],
        columns=[
            "required_output", "target_field", "required_by_model_interface",
            "code_role", "best_public_source_status", "verdict",
        ],
    )

    probes = pd.DataFrame(
        json.loads((RAW / "eu27_coverage_probe_2020.json").read_text(encoding="utf-8"))
    )
    coverage_probe_summary = (
        probes.groupby("dataset", as_index=False)
        .agg(
            eu27_countries_tested=("geo", "size"),
            anonymous_http_200=("anonymous_http_access", "sum"),
            countries_with_2020_observations=("available", "sum"),
            min_observations=("observations", "min"),
            max_observations=("observations", "max"),
        )
    )
    coverage_probe_summary["passes_complete_eu27_2020"] = (
        coverage_probe_summary["countries_with_2020_observations"] == 27
    )
    nearest_probes = pd.DataFrame(
        json.loads(
            (RAW / "nearest_year_probe_2015_2024.json").read_text(encoding="utf-8")
        )
    )
    nearest_summary = (
        nearest_probes.groupby(["dataset", "geo"], as_index=False)
        .agg(
            years_tested=("year", "size"),
            years_with_observations=("available", "sum"),
            maximum_observations=("observations", "max"),
        )
    )
    nearest_summary["verdict"] = np.where(
        nearest_summary["years_with_observations"] > 0,
        "Nearby year available",
        "No observations in 2015-2024 live probe",
    )

    source_register = pd.DataFrame(
        [
            ["PEFA env_ac_pefasu", "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/env_ac_pefasu?geo=DK&time=2020&lang=en", "Anonymous API", "Eurostat reuse policy", "EU-27; annual; DK 2020", "TJ; product×NACE", "Physical backbone; 99.389% comparable aggregate", "No prices or purpose"],
            ["National supply table naio_10_cp15", "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/naio_10_cp15?geo=DK&time=2020&lang=en", "Anonymous API", "Eurostat reuse policy", "26/27 with 2020 observations; BG absent also in 2015-2024 live probe", "MIO_EUR/MIO_NAC; CPA×industry/valuation item", "Broad product totals: basic, purchaser, combined margins, net product taxes", "No energy-product×user decomposition; combined margins/net taxes"],
            ["National use table naio_10_cp16", "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/naio_10_cp16?geo=DK&time=2020&lang=en", "Anonymous API", "Eurostat reuse policy", "26/27 with 2020 observations; BG absent also in 2015-2024 live probe", "MIO_EUR/MIO_NAC; CPA×user", "Broad CPA purchaser-value controls", "No valuation-component split"],
            ["Environmental taxes env_ac_taxind2", "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/env_ac_taxind2?geo=DK&time=2020&lang=en", "Anonymous API", "Eurostat reuse policy", "27/27 with 2020 observations", "MIO_EUR/MIO_NAC; tax category×payer NACE", "Energy/CO2 payer controls", "Not product rates; categories do not equal five GREU taxes"],
            ["Gas price components nrg_pc_202_c/203_c", "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/nrg_pc_202_c?geo=DK&time=2020&lang=en", "Anonymous API", "Eurostat reuse policy", "HH 24/27; non-HH 25/27 in 2020; missing countries also absent 2015-2024", "EUR/PPS/NAC per kWh or GJ; band×component", "Natural-gas price/tax controls by household status and band", "Not all EU-27; no NACE; one product only"],
            ["Electricity price components nrg_pc_204_c/205_c", "https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/nrg_pc_204_c?geo=DK&time=2020&lang=en", "Anonymous API", "Eurostat reuse policy", "27/27 with 2020 observations", "currency/kWh; band×component", "Electricity supply/network/tax/VAT controls", "No NACE or purpose; components differ from GREU taxes"],
            ["EC Weekly Oil Bulletin", "https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en", "Anonymous XLSX", "EC legal notice/reuse; no dataset-specific licence found", "EU-27; weekly history 2005 onward incl. 2020", "Consumer prices/taxes per 1000 L or tonne", "Road fuels and selected petroleum controls", "Limited products/market segments; no NACE/purpose"],
            ["DG TAXUD energy excise tables", "https://taxation-customs.ec.europa.eu/system/files/2021-09/excise_duties-part_ii_energy_products_en.pdf", "Anonymous PDF", "EC legal notice/reuse; no dataset-specific licence found", "All EU countries; situation 2021-07-01 (nearest preserved rate table)", "National rates by energy product/use", "Legal-rate reference", "PDF/manual structure; nearest not exact 2020; exemptions/use rules required"],
            ["DG TAXUD VAT rates", "https://taxation-customs.ec.europa.eu/document/download/82a38bdb-d724-472d-8e02-325b271e0d88_en?filename=vat_rates_en.pdf", "Anonymous PDF", "EC legal notice/reuse; no dataset-specific licence found", "EU Member States; situation 2020-01-01", "Percent rates and product/service treatment", "Exact-date legal-rate control; DK standard 25%", "Does not supply taxable base, recovery or cell allocation"],
            ["FIGARO naio_10_fcp_s3/u3/ii3", "https://joint-research-centre.ec.europa.eu/projects-and-activities/trade-and-industrial-policy-analysis/input-output-accounts/figaro-tables_en", "Anonymous SDMX API", "Eurostat/JRC reuse policy", "EU-27 and partners; 2010-2024", "MIO_EUR; CPA/NACE A64 basic prices", "Economy-wide IO/basic-price control", "Not an energy-only monetary account"],
        ],
        columns=[
            "source", "official_url", "access", "licensing_reuse", "coverage_years",
            "units_dimensions", "usable_role", "hard_limitation",
        ],
    )
    source_register.insert(2, "retrieval_date", RETRIEVAL_DATE)

    assumptions_gaps = pd.DataFrame(
        [
            ["DIRECT", "PEFA physical flow values", "Official product×NACE physical observations; prior pilot reconciliation", "Usable with published classification loss"],
            ["DIRECT", "SUT broad CPA valuation totals", "Official basic/purchaser/net-tax/combined-margin controls", "Calibration controls only"],
            ["DIRECT", "Energy/environmental tax payer totals", "Official NACE payer aggregates", "Cannot identify energy product or tax rate"],
            ["DIRECT", "Electricity/gas/oil price observations", "Official averages by band/market segment", "Cannot assign to each NACE/purpose"],
            ["DERIVED", "Purchaser price", "basic + three margins + five taxes + VAT", "Valid only after every component is supplied"],
            ["TRANSPARENT RULE NEEDED", "PEFA product → CPA price control", "No official concordance resolves every GREU energy product", "Owner-approved bridge required"],
            ["TRANSPARENT RULE NEEDED", "Industry/purpose price differentiation", "Retail price bands are not NACE/purpose", "Explicit model/rate rule required"],
            ["CALIBRATION", "Product and payer controls", "RAS/other balancing could reconcile SUT and tax totals", "Allocation method must be approved and residuals retained"],
            ["UNAVAILABLE", "Three separate GREU trade margins per energy cell", "SUT publishes combined trade+transport margins", "Cannot be observed EU-wide"],
            ["UNAVAILABLE", "Five Danish-style taxes per energy cell", "Tax accounts and price components use different, broader categories", "Redesign tax interface or build country tax engine"],
            ["UNAVAILABLE", "VAT amount per energy cell", "Rate is public but taxable base/recovery/exemption status is not", "Requires modelled incidence"],
            ["PROHIBITED IN THIS PILOT", "Danish target shares as fill keys", "Would make the pipeline Danish-dependent", "Targets are validation benchmarks only"],
        ],
        columns=["status", "concept", "evidence_or_rule", "implication"],
    )

    validation = pd.DataFrame(
        [
            ["Target physical supply-use identity", energy.loc[energy.bal == "sup", "pj"].sum() - energy.loc[energy.bal == "use", "pj"].sum(), "PJ", 1e-5],
            ["Target max purchaser identity residual", identity_cells["purch_identity_residual"].abs().max(), "bn DKK", 1e-8],
            ["Target total purchaser identity residual", identity_cells["purch_identity_residual"].sum(), "bn DKK", 1e-8],
            ["Target IO total residual max", io_total_recon["io_minus_expected_bn_dkk"].abs().max(), "bn DKK", 0.0021],
            ["Target IO production-row residual max", io_row_recon["io_minus_expected_bn_dkk"].abs().max(), "bn DKK", 0.0021],
            ["Target IO column residual max", io_column_recon["io_minus_energy_purch_bn_dkk"].abs().max(), "bn DKK", 0.0021],
            ["SUT selected-product valuation identity residual max", sut_identity["supply_valuation_residual_mio_eur"].abs().max(), "MIO_EUR", 0.2],
            ["SUT selected-product supply-use residual max", sut_identity["supply_minus_use_mio_eur"].abs().max(), "MIO_EUR", 0.2],
        ],
        columns=["check", "absolute_residual", "unit", "tolerance"],
    )
    validation["status"] = np.where(
        validation["absolute_residual"].abs() <= validation["tolerance"], "PASS", "CHECK"
    )

    metadata = pd.DataFrame(
        [
            ["created", RETRIEVAL_DATE],
            ["country_year", "Denmark (DK), 2020"],
            ["artifact_type", "Evidence-backed feasibility/gap workbook; not model-ready input"],
            ["headline_verdict", "NO: a complete energy-money input cannot be produced from direct EU-wide public fields without explicit allocation/tax modelling"],
            ["direct_monetary_cell_coverage", "0%; available monetary sources are controls at incompatible margins"],
            ["physical_coverage", f"{physical_share:.6f}% of comparable PJ in the prior PEFA pilot"],
            ["target_use", "Danish workbooks are validation benchmarks only; no target share fills a source gap"],
            ["inputs_inspected", "energy_and_emissions.xlsx/ems_energy; io_energy_long_format.xlsx/io; io_energy_matrix_format.xlsx/2020; metadata.xlsx/all sheets"],
            ["script", "data/preprocessing/scripts/reconcile_energy_money_dk_2020.py"],
            ["raw_folder", "data/preprocessing/data/eurostat_energy_money_raw/"],
            ["output", "data/preprocessing/data/energy_money_dk2020_feasibility_gap.xlsx"],
            ["currency_conversion", f"7.4542 DKK/EUR, 2020 annual average (existing official Eurostat FIGARO pilot pull)"],
            ["important_boundary", "SUT CPA controls are product bundles; they are not PEFA/GREU energy-product values"],
        ],
        columns=["key", "value"],
    )

    sheets = OrderedDict(
        [
            ("metadata", metadata),
            ("field_requirements", field_requirements),
            ("coverage_summary", coverage_summary),
            ("source_register", source_register),
            ("eu27_coverage", coverage_probe_summary),
            ("eu27_probe_detail", probes),
            ("nearest_year_gaps", nearest_summary),
            ("nearest_probe_detail", nearest_probes),
            ("target_benchmarks", target_benchmarks),
            ("target_identity_cells", identity_cells),
            ("io_total_recon", io_total_recon),
            ("io_row_recon", io_row_recon),
            ("io_column_recon", io_column_recon),
            ("sut_energy_controls", sut_controls),
            ("sut_use_controls", sut_use_controls),
            ("sut_identities", sut_identity),
            ("tax_by_payer", tax_by_payer),
            ("price_components", price_components),
            ("oil_2020_prices", oil_prices),
            ("oil_rate_snapshot", oil_rates),
            ("assumptions_gaps", assumptions_gaps),
            ("validation", validation),
        ]
    )
    with pd.ExcelWriter(OUT, engine="openpyxl") as writer:
        for name, frame in sheets.items():
            frame.to_excel(writer, sheet_name=name, index=False)
    style_workbook(OUT)
    print(f"wrote {OUT}")
    print(coverage_summary.to_string(index=False))
    print(validation.to_string(index=False))


if __name__ == "__main__":
    main()
