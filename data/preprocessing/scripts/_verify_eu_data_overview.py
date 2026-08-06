"""One-off check that docs/EU_data_overview.xlsx is well-formed."""
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

path = Path(__file__).resolve().parents[3] / "docs" / "EU_data_overview.xlsx"
legal = {"MATCHES", "CLOSE MATCH", "COARSER", "CONSTRUCTED", "GAP", "KEPT"}

xl = pd.ExcelFile(path)
assert xl.sheet_names == ["README", "Summary", "Detail"], xl.sheet_names

summary = pd.read_excel(path, sheet_name="Summary")
detail = pd.read_excel(path, sheet_name="Detail")
print("Summary:", summary.shape, "| Detail:", detail.shape)
assert list(summary.columns)[4] == "Status" and list(detail.columns)[4] == "Status"

bad_s = set(summary["Status"]) - legal
bad_d = set(detail["Status"]) - legal
assert not bad_s and not bad_d, (bad_s, bad_d)
assert summary.notna().all().all(), summary.isna().sum()
assert detail.notna().all().all(), detail.isna().sum()

# Every Summary input (except the module row) appears in Detail.
detail_inputs = set(detail["Danish input"])
missing = [i for i in summary["Danish input"]
           if not i.startswith("Financial-accounts module") and i not in detail_inputs]
assert not missing, missing

print("Status counts (Summary):", summary["Status"].value_counts().to_dict())
print("Status counts (Detail): ", detail["Status"].value_counts().to_dict())

# Check the fills were applied to status cells.
wb = load_workbook(path)
fills = {ws.title: {row[4].fill.fgColor.rgb for row in ws.iter_rows(min_row=2)}
         for ws in (wb["Summary"], wb["Detail"])}
print("Distinct status fills:", {k: len(v) for k, v in fills.items()})
print("ALL CHECKS PASS")
